from operator import lt
import re
import matplotlib.pyplot as plt
import os
import seaborn
import ast
import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime

def normalize_answer(text: str) -> str:
        """Normalize answer format for comparison."""
        text = str(text).strip().lower()
        
        # Remove </think> if present
        text = text.replace('</think>', '')
        
        # Remove ". " (period followed by space) after word characters
        text = re.sub(r'(\w)\.', r'\1 ', text)
        
        # Remove special characters like %&!#$
        text = re.sub(r'[%&!#$?]', '', text)
        
        # Remove spaces ONLY inside parentheses: "(x - 2)" → "(x-2)"
        text = re.sub(r'\(\s*([^)]+?)\s*\)', lambda m: '(' + m.group(1).replace(' ', '') + ')', text)
    
        
        # Normalize spacing around commas: "1 , 6" or "1, 6" → "1,6"
        text = ','.join([part.strip() for part in text.split(',')])

        if ',' in text and re.match(r'^[(\s]*[\d\s,.-]+[)\s]*$', text):
            # This is coordinate-like, safe to remove parentheses and normalize spaces
            text = text.replace('(', '').replace(')', '')
        
        return text
    
# def extract_answers(file_path, num_agents=3, num_rounds=3):
def extract_answers(file_path):
    """
    Extract initial, round and final answers, and the correct answer from task transcript files.
    Strip, lowercase, and return all answers for consistent comparison.
    
    Args:
        file_path: Path to the directory containing transcript files.
        num_agents: Number of agents in each task (default: 3).
        num_rounds: Number of rounds in each task (default: 3).
    """
    answer_regex = r"(?m)^\s*Person \d+ answered:[ \t]*(.+?)(?=(?:\.\s|\n|$))" #r"(?<=answered: )(.+?)(?=(?:\.\s|\n|$))" 
    correct_answer_regex = r"(?<=Correct Answer: )(.+)"
    round_answer_regex = r"(?<=The answer is: )(.+?)(?=(?:\.\s|\n|\[SHARED\]|$|\.$))"
    task_id_regex = r"(?<=Task ID: )(\d+)"

    answers = dict()
    round_answers = dict()
    task_files = [f for f in os.listdir(file_path) if "task" in f and f.endswith(".txt")]
    
    # Summary information
    summary_file = os.path.join(file_path, "summary.txt")
    with open(summary_file, 'r') as f:
        summary_text = f.read()
    summary_text.split()

    num_agents_regex = r"(?:Agents: )(\d+)"
    num_rounds_regex = r"(?:Rounds: )(\d+)"
    malicious_regex = r"(?:Malicious agent\(s\): )(.+)(?:\s+\|)"
    es_regex = r"(?:Kept reasoning agent: )(\[[^\]]+\])"
    share_mode_regex = r"(?:Shared content mode: )(\w+)"
    dataset_name_regex = r"(?:Dataset: \w+\/)(\w+)"
    model_name_regex = r"(?:Model: .+\/)(.+)"
    
    num_agents = int(re.findall(num_agents_regex, summary_text)[0])
    num_rounds = int(re.findall(num_rounds_regex, summary_text)[0])
    malicious_agents = ast.literal_eval(re.findall(malicious_regex, summary_text)[0]) if re.findall(malicious_regex, summary_text) else []
    es_agents = ast.literal_eval(re.findall(es_regex, summary_text)[0]) if re.findall(es_regex, summary_text) else []
    share_mode = re.findall(share_mode_regex, summary_text)[0]
    dataset_filename = re.findall(dataset_name_regex, summary_text)[0]
    system = "MAS" if num_agents > 1 else "SAS"
    model_name = re.findall(model_name_regex, summary_text)[0]

    for _, filename in enumerate(sorted(task_files), 1): 
        rounds_a = dict()
        # Read the entire file as a string
        with open(os.path.join(file_path, filename), 'r') as file:
            text = file.read()
        
        text.split()

        idx = re.findall(task_id_regex, text)[0]
        result = re.findall(answer_regex, text)
        init_answer = [normalize_answer(ans) for ans in result[0:num_agents]]
        final_answer = [normalize_answer(ans) for ans in result[num_agents:]]
        correct_answer = [normalize_answer(ans) for ans in re.findall(correct_answer_regex, text)]


        answers[idx] = (init_answer, final_answer, correct_answer)

        # rounds = re.split(r'\n\s*(?=Round \d+:)', text)
        rounds = re.split(r'\n  (?=Round \d+:)', text)

        # keep only actual rounds
        rounds = [r for r in rounds if r.startswith("Round")]
        
        for round_num in range(1, num_rounds + 1):
            person_round = re.split(r'\n\s*(?=Person \d+:)', rounds[round_num - 1])
            person_round = [pr for pr in person_round if pr.startswith("Person")]
            person_round_answer = []
        
            for agent_num in range(1, num_agents + 1):
                round_answer = re.findall(round_answer_regex, person_round[agent_num - 1])
                person_round_answer.append(normalize_answer(round_answer[0]) if round_answer else '')
                
            
            rounds_a[round_num] = person_round_answer
        
        
        rounds_a[num_rounds + 1] = final_answer
        round_answers[idx] = (init_answer, rounds_a, correct_answer)

    summary_info = {
        'num_agents': num_agents,
        'num_rounds': num_rounds,
        'malicious_agents': malicious_agents,
        'es_agents': es_agents,
        'share_mode': share_mode,
        'dataset_filename': dataset_filename,
        'system': system,
        'model_name': model_name
    }
    return answers, round_answers, summary_info




# def calculate_pattern_statistics(tasks_answers, skip_agents=[], original_task_answers=None):
#     """Calculate statistics WITHOUT creating a plot. Returns raw counts."""
#     wrong_correct = 0
#     wrong_different_wrong = 0
#     wrong_same_wrong = 0
#     wrong_malicious_wrong = 0
#     correct_wrong = 0
#     correct_malicious_wrong = 0
#     correct_correct = 0
#     num_tasks = len(tasks_answers)
#     missed_patterns = 0
#     missed_patterns_dict = dict()

#     basline_and_malicious_equal = 0
#     task_doesnt_exists_in_baseline = 0
#     malicious_agent_answers_groundtruth = 0 
    
#     # print(f"num_tasks before: {num_tasks}")
#     # malicious_final_answers = []  # Track original answers from skipped agents that changed
#     # num_agents = len(tasks_answers.items().__iter__().__next__()[1][0]) - len(skip_agents)  # Get number of agents from the first task's init_answer
    
#     for task_num, (init_answer, final_answer, correct_answer) in tasks_answers.items():
#         malicious_round1_answers = []  # Track original answers from skipped agents that changed
#         skip_task = False
#         if original_task_answers is not None and task_num in original_task_answers[1]:
#             round_answers, baseline_round_answers = original_task_answers
#             # print(round_answers[task_num][1])
#             # print(round_answers[task_num][1][1])

#             round1_answers = round_answers[task_num][1][1]  # Get round 1 answers for this task
#             baseline_round1_answers = baseline_round_answers[task_num][1][1]  # Get round 1 answers for this task
      
#             # First pass: identify if any malicious agent changed their answer
#             # mal_agent_changes = []  # Track which malicious agents changed
#             for agent_idx in skip_agents:
#                 original_agent_round1 = baseline_round1_answers[agent_idx]
#                 current_agent_round1 = round1_answers[agent_idx]
               
#                 if current_agent_round1 != correct_answer[0]:
#                     if current_agent_round1 != original_agent_round1:
#                         #print(f"Malicious agent {agent_idx + 1} malicious answer is {current_agent_round1} in task {task_num}")
#                         malicious_round1_answers.append(current_agent_round1)
#                     else:
#                         basline_and_malicious_equal += 1
#                         skip_task = True
#                 else:
#                     malicious_agent_answers_groundtruth += 1 
#                     skip_task = True

        
#         if skip_task:
#             num_tasks -= 1  # Exclude this task from total count since we're skipping it
#             print(f"Skipping task {task_num}")
#             continue
#         elif skip_agents != [] and malicious_round1_answers == []:
#             task_doesnt_exists_in_baseline += 1
#             num_tasks -= 1  # Exclude this task from total count since we're skipping it
#             print(f"Skipping task {task_num}")
#             continue
#         for agent_idx, agent_answer in enumerate(zip(init_answer, final_answer)):
#             wrong_init_answer = agent_answer[0] != correct_answer[0]
#             correct_final_answer = agent_answer[1] == correct_answer[0]
            
#             # Skip this agent if in skip list
#             if agent_idx in skip_agents:
#                 continue
#             #print(f"Task {task_num} - Agent answer is {agent_answer[1]}")
#             if wrong_init_answer and correct_final_answer:
#                 wrong_correct += 1
#             elif wrong_init_answer and not correct_final_answer and agent_answer[1] in malicious_round1_answers:
#                 wrong_malicious_wrong += 1
#                 #print(f"Agent {agent_idx + 1} in task {task_num} changed from wrong to malicious wrong: {agent_answer[1]}")
#             elif wrong_init_answer and not correct_final_answer and init_answer != final_answer:
#                 wrong_different_wrong += 1
#             elif wrong_init_answer and init_answer == final_answer:
#                 wrong_same_wrong += 1 
#             elif not wrong_init_answer and not correct_final_answer and agent_answer[1] in malicious_round1_answers:
#                 correct_malicious_wrong += 1
#                 #print(f"Agent {agent_idx + 1} in task {task_num} changed from correct to malicious wrong: {agent_answer[1]}")
#             elif not wrong_init_answer and not correct_final_answer:
#                 correct_wrong += 1
#             elif not wrong_init_answer and correct_final_answer:
#                 correct_correct += 1
#             else:
#                 missed_patterns += 1
#                 missed_patterns_dict[task_num] = (init_answer, final_answer, correct_answer)
#                 print(f"Missed pattern for task {task_num}: init_answer={init_answer}, final_answer={final_answer}, correct_answer={correct_answer}")

#     #print(f"num_tasks after: {num_tasks}")
#     # Return counts dictionary instead of plotting
#     return {
#         'wrong_correct': wrong_correct,
#         'wrong_different_wrong': wrong_different_wrong,
#         'wrong_same_wrong': wrong_same_wrong,
#         'wrong_malicious_wrong': wrong_malicious_wrong,
#         'correct_wrong': correct_wrong,
#         'correct_malicious_wrong': correct_malicious_wrong,
#         'correct_correct': correct_correct,
#         'num_tasks': num_tasks,
#         #'num_agents': num_agents,
#         'missed_patterns': missed_patterns,
#         'missed_patterns_dict': missed_patterns_dict,
#         'basline_and_malicious_equal': basline_and_malicious_equal,
#         'task_doesnt_exists_in_baseline': task_doesnt_exists_in_baseline,
#         'malicious_agent_answers_groundtruth': malicious_agent_answers_groundtruth
#     }


def calculate_pattern_statistics(tasks_answers, skip_agents=[], original_task_answers=None, rounds=3):
    """Calculate statistics WITHOUT creating a plot. Returns raw counts."""
    wrong_correct = 0
    wrong_different_wrong = 0
    wrong_same_wrong = 0
    wrong_malicious_wrong = 0
    correct_wrong = 0
    correct_malicious_wrong = 0
    correct_correct = 0
    num_tasks = len(tasks_answers)
    missed_patterns = 0
    missed_patterns_dict = dict()

    basline_and_malicious_equal = 0
    task_doesnt_exists_in_baseline = 0
    malicious_agent_answers_groundtruth = 0 

    prompt_injection_experiment = True

    different_malicious_answers = 0
    dfa_correct_to_correct = 0
    dfa_propagation = 0
    dfa_malicious_propagation = 0
    
    # print(f"num_tasks before: {num_tasks}")
    # malicious_final_answers = []  # Track original answers from skipped agents that changed
    # num_agents = len(tasks_answers.items().__iter__().__next__()[1][0]) - len(skip_agents)  # Get number of agents from the first task's init_answer
    
    for task_num, (init_answer, final_answer, correct_answer) in tasks_answers.items():
        malicious_round_answers = []  # Track original answers from skipped agents that changed
        skip_task = False

        if original_task_answers is not None and task_num in original_task_answers[1]:
            prompt_injection_experiment = False
            round_answers, baseline_round_answers = original_task_answers
            # print(round_answers[task_num][1])
            # print(round_answers[task_num][1][1])

            # round1_answers = round_answers[task_num][1][1]  # Get round 1 answers for this task
            # round2_answers = round_answers[task_num][1][2]  # Get round 2 answers for this task
            # round3_answers = round_answers[task_num][1][3]  # Get round 3 answers for this task
            # baseline_round1_answers = baseline_round_answers[task_num][1][1]  # Get round 1 answers for this task
            # baseline_round2_answers = baseline_round_answers[task_num][1][2]  # Get round 2 answers for this task
            # baseline_round3_answers = baseline_round_answers[task_num][1][3]  # Get round 3 answers for this task

            round_answers = [round_answers[task_num][1][r] for r in range(1, rounds + 1)]
            baseline_round_answers = [baseline_round_answers[task_num][1][r] for r in range(1, rounds + 1)]
      
            # First pass: identify if any malicious agent changed their answer
            # mal_agent_changes = []  # Track which malicious agents changed
            for agent_idx in skip_agents:
                # original_agent_rounds = [baseline_round1_answers[agent_idx], baseline_round2_answers[agent_idx], baseline_round3_answers[agent_idx]]
                # current_agent_rounds = [round1_answers[agent_idx], round2_answers[agent_idx], round3_answers[agent_idx]]
               
                original_agent_rounds = [baseline_round_answers[r-1][agent_idx] for r in range(1, rounds + 1)]
                current_agent_rounds = [round_answers[r-1][agent_idx] for r in range(1, rounds + 1)]

                if all(ca == correct_answer[0] for ca in current_agent_rounds) == False:  # If the agent's current answers are not all correct
                    for r in range(rounds):
                        if current_agent_rounds[r] != correct_answer[0] and current_agent_rounds[r] != original_agent_rounds[r]:  # If any of the agent's round answers is not correct and not the same as baseline, consider it as malicious wrong for pattern analysis
                            malicious_round_answers.append(current_agent_rounds[r])
                            break
                    else:  # If all the agent's round answers are either correct or the same as baseline, consider it as equal for pattern analysis
                        basline_and_malicious_equal += 1
                        skip_task = True
                    # if current_agent_rounds[0] != correct_answer[0] and current_agent_rounds[0] != original_agent_rounds[0]:  # If the agent's round 1 answer is not correct, consider it as malicious wrong for pattern analysis
                    # #print(f"Malicious agent {agent_idx + 1} malicious answer is {current_agent_round1} in task {task_num}")
                    #     malicious_round_answers.append(current_agent_rounds[0])
                    # elif current_agent_rounds[1] != correct_answer[0] and current_agent_rounds[1] != original_agent_rounds[1]:
                    #     malicious_round_answers.append(current_agent_rounds[1])
                    # elif current_agent_rounds[2] != correct_answer[0] and current_agent_rounds[2] != original_agent_rounds[2]:
                    #     malicious_round_answers.append(current_agent_rounds[2])
                    # else:
                    #     basline_and_malicious_equal += 1
                    #     skip_task = True
                else:
                    malicious_agent_answers_groundtruth += 1 
                    skip_task = True
        
        if skip_task:
            num_tasks -= 1  # Exclude this task from total count since we're skipping it
            # print(f"Skipping task {task_num}")
            continue
        elif skip_agents != [] and malicious_round_answers == [] and prompt_injection_experiment == False:
            task_doesnt_exists_in_baseline += 1
            num_tasks -= 1  # Exclude this task from total count since we're skipping it
            # print(f"Skipping task {task_num}")
            continue

        if len(malicious_round_answers) > 1 and malicious_round_answers[0] != malicious_round_answers[1]:
            different_malicious_answers += 1
        
        for agent_idx, agent_answer in enumerate(zip(init_answer, final_answer)):
            wrong_init_answer = agent_answer[0] != correct_answer[0]
            correct_final_answer = agent_answer[1] == correct_answer[0]
            
            # Skip this agent if in skip list
            if agent_idx in skip_agents:
                continue
            #print(f"Task {task_num} - Agent answer is {agent_answer[1]}")
            if wrong_init_answer and correct_final_answer:
                wrong_correct += 1
                if len(malicious_round_answers) > 1 and malicious_round_answers[0] != malicious_round_answers[1]:
                    dfa_propagation += 1
            elif wrong_init_answer and not correct_final_answer and agent_answer[1] in malicious_round_answers:
                wrong_malicious_wrong += 1
                #print(f"Agent {agent_idx + 1} in task {task_num} changed from wrong to malicious wrong: {agent_answer[1]}")
            elif wrong_init_answer and not correct_final_answer and init_answer != final_answer:
                wrong_different_wrong += 1
            elif wrong_init_answer and init_answer == final_answer:
                wrong_same_wrong += 1 
            elif not wrong_init_answer and not correct_final_answer and agent_answer[1] in malicious_round_answers:
                correct_malicious_wrong += 1
                if len(malicious_round_answers) > 1 and malicious_round_answers[0] != malicious_round_answers[1]:
                    dfa_malicious_propagation += 1
                #print(f"Agent {agent_idx + 1} in task {task_num} changed from correct to malicious wrong: {agent_answer[1]}")
            elif not wrong_init_answer and not correct_final_answer:
                correct_wrong += 1
            elif not wrong_init_answer and correct_final_answer:
                correct_correct += 1
                if len(malicious_round_answers) > 1 and malicious_round_answers[0] != malicious_round_answers[1]:
                    dfa_correct_to_correct += 1
            else:
                missed_patterns += 1
                missed_patterns_dict[task_num] = (init_answer, final_answer, correct_answer)
                print(f"Missed pattern for task {task_num}: init_answer={init_answer}, final_answer={final_answer}, correct_answer={correct_answer}")

    print(f"Total times across task that have different malicious answers: {different_malicious_answers}")
    print(f"    Different malicious answers and correct propagated to wrong: {dfa_propagation}")
    print(f"    Different malicious answers and correct propagated to malicious wrong: {dfa_malicious_propagation}")
    print(f"    Different malicious answers and stayed to correct: {dfa_correct_to_correct}")
    #print(f"num_tasks after: {num_tasks}")
    # Return counts dictionary instead of plotting
    return {
        'wrong_correct': wrong_correct,
        'wrong_different_wrong': wrong_different_wrong,
        'wrong_same_wrong': wrong_same_wrong,
        'wrong_malicious_wrong': wrong_malicious_wrong,
        'correct_wrong': correct_wrong,
        'correct_malicious_wrong': correct_malicious_wrong,
        'correct_correct': correct_correct,
        'num_tasks': num_tasks,
        #'num_agents': num_agents,
        'missed_patterns': missed_patterns,
        'missed_patterns_dict': missed_patterns_dict,
        'basline_and_malicious_equal': basline_and_malicious_equal,
        'task_doesnt_exists_in_baseline': task_doesnt_exists_in_baseline,
        'malicious_agent_answers_groundtruth': malicious_agent_answers_groundtruth
    }


def plot_pattern_statistics_comparison(datasets_data, model_name, dataset_filename='', share_mode='Both', system='MAS', num_agents=3, num_malicious_agents=0, comparison_type='dataset'):
    """
    Plot statistics across multiple datasets side-by-side.
    
    Args:
        datasets_data: Dict like {'openai-gsm8k': stats_dict, 'cais-mmlu': stats_dict, ...}
        model_name: Name of the model (for title/filename)
        dataset_filename: Filename-friendly version of dataset name (for title/filename)
        share_mode: 'Both', 'Reasoning', or 'Answer'
        system: 'MAS' or 'SAS'
        num_agents: Number of agents
        comparison_type: Type of comparison to plot; 'share-mode' for share mode comparison, 'malicious' for malicious comparison, or 'dataset' for dataset comparison
    """
    if comparison_type == 'malicious' or comparison_type == 'early_stopping' or comparison_type == 'IR-CA' or comparison_type == 'CR-IA':
        categories = ['Wrong→Correct', 'Wrong→Different Wrong', 'Wrong→Same Wrong', 'Correct→Wrong', 'Correct→Correct', 'Wrong→Malicious Wrong', 'Correct→Malicious Wrong']
    elif comparison_type == 'sharemode_malicious' or comparison_type == 'malicious_convergence':
        categories = ['Wrong→Correct', 'Wrong→Wrong', 'Correct→Wrong', 'Correct→Correct', 'Malicious Wrong Convergence']

        datasets_data = {
            name: stats
            for name, stats in datasets_data.items()
            if ("No Mal Agent" not in name and "No Malicious Agent" not in name)
        }
    else:
        categories = ['Wrong→Correct', 'Wrong→Different Wrong', 'Wrong→Same Wrong', 'Correct→Wrong', 'Correct→Correct']
       

    
    if comparison_type == 'sharemode_malicious' or comparison_type == 'malicious_convergence':
        fig, ax = plt.subplots(figsize=(22, 7))
        x = range(len(categories))
        bar_width = 0.1

        def dataset_family(name):
            # Example:
            # "openai/gsm8k: Mal Agent 1 - Answer" -> "openai/gsm8k"
            return name.split(':', 1)[0].strip()

        family_order = []
        for name in datasets_data.keys():
            fam = dataset_family(name)
            if fam not in family_order:
                family_order.append(fam)

        family_colors = seaborn.color_palette("deep", n_colors=max(1, len(family_order)))
        family_color_map = {fam: family_colors[i] for i, fam in enumerate(family_order)}
    else:
        fig, ax = plt.subplots(figsize=(14, 6))
        x = range(len(categories))
        bar_width = 0.15 
        colors = seaborn.color_palette("deep", n_colors=len(datasets_data))
    
    
    for idx, (dataset_name, payload) in enumerate(datasets_data.items()):
        if isinstance(payload, dict) and "stats" in payload:
            stats = payload["stats"]
            num_honest_agents = payload.get("honest_agents", num_agents)
            # malicious_agents = payload["malicious_agents"]
        else:
            stats = payload
            num_honest_agents = num_agents
        if comparison_type == 'malicious' or comparison_type == 'early_stopping' or comparison_type == 'IR-CA' or comparison_type == 'CR-IA':
            counts = [
                stats['wrong_correct'],
                stats['wrong_different_wrong'],
                stats['wrong_same_wrong'],
                stats['correct_wrong'],
                stats['correct_correct'],
                stats['wrong_malicious_wrong'], 
                stats['correct_malicious_wrong'],
            ]
            print(f"num_tasks {stats['num_tasks']}, num_honest_agents {num_honest_agents}")
            percentages = [(count / (stats['num_tasks'] * num_honest_agents)) * 100 for count in counts]
            # Offset bars for side-by-side display
            offset = bar_width * idx
            bars = ax.bar([i + offset for i in x], percentages, bar_width, 
                        label=dataset_name, color=colors[idx], alpha=0.8)  
        elif comparison_type == 'sharemode_malicious' or comparison_type == 'malicious_convergence':
            counts = [
                stats['wrong_correct'],
                stats['wrong_different_wrong'] + stats['wrong_same_wrong'],
                stats['correct_wrong'],
                stats['correct_correct'],
                stats['wrong_malicious_wrong'] + stats['correct_malicious_wrong']
            ]
            print(f"Malicious Wrong Convergence: {stats['wrong_malicious_wrong'] + stats['correct_malicious_wrong']}")

    
            percentages = [(count / (stats['num_tasks'] * num_honest_agents)) * 100 for count in counts]
            # Offset bars for side-by-side display
            offset = bar_width * idx

            fam = dataset_family(dataset_name)
            bar_color = family_color_map[fam]

            name_lower = dataset_name.lower()
            if 'answer' in name_lower or num_honest_agents == 1:
                bar_alpha = 0.45
            elif 'both' in name_lower or num_honest_agents == 2:
                bar_alpha = 0.80
            else:
                bar_alpha = 0.70
                
            bars = ax.bar([i + offset for i in x], percentages, bar_width, 
                        label=dataset_name, color=bar_color, alpha=bar_alpha)  
        else:
            if "SAS" in dataset_name:
                num_agents_update = 1  # For SAS, only 1 agent's answer is relevant for percentage calculation
            else:
                num_agents_update = num_agents  # For MAS, all agents' answers are relevant
            counts = [
                stats['wrong_correct'],
                stats['wrong_different_wrong'],
                stats['wrong_same_wrong'],
                stats['correct_wrong'],
                stats['correct_correct']
            ]
            percentages = [(count / (stats['num_tasks'] * num_agents_update)) * 100 for count in counts]
            # Offset bars for side-by-side display
            offset = bar_width * idx
            bars = ax.bar([i + offset for i in x], percentages, bar_width, 
                        label=dataset_name, color=colors[idx], alpha=0.8)
            
        
        # Add percentage labels
        for bar, percentage in zip(bars, percentages):
            height = bar.get_height()
            ax.text(bar.get_x() + bar.get_width()/2., height,
                   f'{percentage:.1f}%',
                   ha='center', va='bottom', fontsize=7)
    
    ax.set_ylabel('Percentage (%) of Tasks')
    ax.set_xlabel('Answer Pattern')
    ax.set_ylim(0, 105)
    
    if comparison_type == 'share-mode':
        ax.set_title(f'{num_agents} Agents {system}, {model_name}:\n Share-Modes Impact on Answer Patterns')
    elif comparison_type == 'malicious':
        ax.set_title(f'Share-Mode: {share_mode} | {num_agents} Agents {system} ({num_malicious_agents} Malicious), {model_name}:\n Malicious Agent Impact on Answer Patterns')
    elif comparison_type == 'CR-IA':
        ax.set_title(f'Share-Mode: {share_mode} | {num_agents} Agents {system} ({num_malicious_agents} CR-IA Agents), {model_name}:\n CR-IA Agent Impact on Answer Patterns')
    elif comparison_type == 'sharemode_malicious':
        ax.set_title(f'{num_agents} Agents {system} ({num_malicious_agents} Malicious), {model_name}:\n Share-Mode and Malicious Agent Impact on Answer Patterns')
    elif comparison_type == 'malicious_convergence':
        ax.set_title(f'Share-Mode: {share_mode} | {num_agents} Agents {system}, {model_name}:\n Malicious Agent Impact on Answer Pattern Convergence')
    elif comparison_type == 'early_stopping':
        ax.set_title(f'Share-Mode: {share_mode} | {num_agents} Agents {system} ({num_malicious_agents} Early Stopping Agents), {model_name}, {dataset_filename} dataset:\n Sharing Early Stopping Reasoning Impact on Answer Pattern Convergence')
    elif comparison_type == 'IR-CA':
        ax.set_title(f'Share-Mode: {share_mode} | {num_agents} Agents {system} ({num_malicious_agents} IR-CA Agents), {model_name}, {dataset_filename} dataset:\n Sharing IR-CA Reasoning Impact on Answer Pattern Convergence')
    elif comparison_type == 'model':
        ax.set_title(f'Share-Mode: {share_mode} | {num_agents} Agents {system}, {dataset_filename} dataset:\n Answer Patterns Across Models')    
    elif comparison_type == 'system':
        ax.set_title(f'Share-Mode: {share_mode} |  {model_name}, {dataset_filename} dataset:\n Answer Patterns Across Systems')    
    else:
        ax.set_title(f'Share-Mode: {share_mode} | {num_agents} Agents {system}, {model_name}:\n Answer Patterns Comparison Across Datasets')
    

    ax.set_xticks([i + bar_width * (len(datasets_data) - 1) / 2 for i in x])
    ax.set_xticklabels(categories, rotation=45, ha='right')
    ax.legend()
        
    plt.tight_layout()
        
    script_dir = os.path.dirname(os.path.abspath(__file__))
    output_folder = os.path.join(script_dir, 'statistic_images')
    os.makedirs(output_folder, exist_ok=True)
    
    dataset_suffix = '_'.join(datasets_data.keys())

    if comparison_type == 'share-mode': 
        plot_image_name = f"{model_name}_{dataset_filename}_{system}_{comparison_type}_comparison_statistics.png"
    elif comparison_type == 'malicious':
        plot_image_name = f"{model_name}_{dataset_filename}_{system}_{num_malicious_agents}_comparison_statistics_{share_mode}.png"
    elif comparison_type == 'CR-IA':
        plot_image_name = f"{model_name}_{dataset_filename}_{system}_{num_malicious_agents}_comparison_statistics_{share_mode}.png"
    elif comparison_type == 'sharemode_malicious':
        plot_image_name = f"{model_name}_{system}_{num_malicious_agents}_{comparison_type}_comparison_statistics.png"
    elif comparison_type == 'malicious_convergence':
        plot_image_name = f"{model_name}_{system}_{comparison_type}_comparison_statistics_{share_mode}.png"
    elif comparison_type == 'early_stopping' or comparison_type == 'IR-CA':
        plot_image_name = f"{model_name}_{system}_{num_malicious_agents}_{comparison_type}_comparison_statistics_{share_mode}.png"
    elif comparison_type == 'model':
        plot_image_name = f"{dataset_filename}_{system}_{comparison_type}_comparison_statistics_{share_mode}.png"
    elif comparison_type == 'system':
        plot_image_name = f"{model_name}_{dataset_filename}_{comparison_type}_comparison_statistics_{share_mode}.png"
    else:
        plot_image_name = f"{model_name}_{system}_{comparison_type}_comparison_statistics_{share_mode}.png"
    
    filepath = os.path.join(output_folder, plot_image_name)
    plt.savefig(filepath)
    print(f"Comparison plot saved as {plot_image_name}")

    # plt.clf()
    plt.close()


def plot_horizontal_pattern_statistics_comparison(
    datasets_data,
    model_name,
    dataset_filename='',
    share_mode='Both',
    system='MAS',
    num_agents=3,
    num_malicious_agents=0,
    comparison_type='dataset'
):
    """
    Horizontal version:
    - Percentage on x-axis
    - Answer patterns on y-axis
    - Better for narrow paper columns
    """

    # Categories + optional filtering for malicious comparison families
    if comparison_type == 'malicious':
        categories = [
            'Wrong→Correct',
            'Wrong→Different Wrong',
            'Wrong→Same Wrong',
            'Correct→Wrong',
            'Correct→Correct',
            'Wrong→Malicious Wrong',
            'Correct→Malicious Wrong'
        ]
    elif comparison_type in ('sharemode_malicious', 'malicious_convergence'):
        categories = [
            'Wrong→Correct',
            'Wrong→Wrong',
            'Correct→Wrong',
            'Correct→Correct',
            'Malicious Wrong Convergence'
        ]
        datasets_data = {
            name: payload
            for name, payload in datasets_data.items()
            if ("No Mal Agent" not in name and "No Malicious Agent" not in name)
        }
    else:
        categories = [
            'Wrong→Correct',
            'Wrong→Different Wrong',
            'Wrong→Same Wrong',
            'Correct→Wrong',
            'Correct→Correct'
        ]

    num_sets = max(1, len(datasets_data))
    y_base = list(range(len(categories)))

    # Compact width for paper, adaptive height for readability
    fig_height = max(4.8, 0.85 * len(categories) + 0.18 * num_sets * len(categories))
    fig, ax = plt.subplots(figsize=(10.5, fig_height))

    # Bar thickness per dataset within each category group
    bar_h = min(0.22, 0.8 / num_sets)

    # Color setup
    if comparison_type in ('sharemode_malicious', 'malicious_convergence'):
        def dataset_family(name):
            return name.split(':', 1)[0].strip()

        family_order = []
        for name in datasets_data.keys():
            fam = dataset_family(name)
            if fam not in family_order:
                family_order.append(fam)

        family_colors = seaborn.color_palette("deep", n_colors=max(1, len(family_order)))
        family_color_map = {fam: family_colors[i] for i, fam in enumerate(family_order)}
    else:
        colors = seaborn.color_palette("deep", n_colors=num_sets)

    for idx, (dataset_name, payload) in enumerate(datasets_data.items()):
        if isinstance(payload, dict) and "stats" in payload:
            stats = payload["stats"]
            num_honest_agents = payload.get("honest_agents", num_agents)
        else:
            stats = payload
            num_honest_agents = num_agents

        if comparison_type == 'malicious':
            counts = [
                stats['wrong_correct'],
                stats['wrong_different_wrong'],
                stats['wrong_same_wrong'],
                stats['correct_wrong'],
                stats['correct_correct'],
                stats['wrong_malicious_wrong'],
                stats['correct_malicious_wrong'],
            ]
            denom_agents = num_honest_agents

        elif comparison_type in ('sharemode_malicious', 'malicious_convergence'):
            counts = [
                stats['wrong_correct'],
                stats['wrong_different_wrong'] + stats['wrong_same_wrong'],
                stats['correct_wrong'],
                stats['correct_correct'],
                stats['wrong_malicious_wrong'] + stats['correct_malicious_wrong'],
            ]
            denom_agents = num_honest_agents

        else:
            counts = [
                stats['wrong_correct'],
                stats['wrong_different_wrong'],
                stats['wrong_same_wrong'],
                stats['correct_wrong'],
                stats['correct_correct'],
            ]
            denom_agents = 1 if "SAS" in dataset_name else num_agents

        denom = max(1, stats['num_tasks'] * denom_agents)
        percentages = [(count / denom) * 100 for count in counts]

        # Vertical offset per dataset within each pattern row
        y_offset = (idx - (num_sets - 1) / 2.0) * bar_h
        y_pos = [y + y_offset for y in y_base]

        if comparison_type in ('sharemode_malicious', 'malicious_convergence'):
            fam = dataset_family(dataset_name)
            bar_color = family_color_map[fam]
            name_lower = dataset_name.lower()
            if 'answer' in name_lower or num_honest_agents == 1:
                bar_alpha = 0.45
            elif 'both' in name_lower or num_honest_agents == 2:
                bar_alpha = 0.80
            else:
                bar_alpha = 0.70
            bars = ax.barh(y_pos, percentages, height=bar_h, label=dataset_name, color=bar_color, alpha=bar_alpha)
        else:
            bars = ax.barh(y_pos, percentages, height=bar_h, label=dataset_name, color=colors[idx], alpha=0.8)

        # Value labels at bar end
        for bar, pct in zip(bars, percentages):
            ax.text(
                bar.get_width() + 0.8,
                bar.get_y() + bar.get_height() / 2.0,
                f'{pct:.1f}%',
                va='center',
                ha='left',
                fontsize=6
            )

    ax.set_xlabel('Percentage (%) of Tasks')
    ax.set_ylabel('Answer Pattern')
    ax.set_xlim(0, 105)
    ax.set_yticks(y_base)
    ax.set_yticklabels(categories)
    ax.invert_yaxis()

    if comparison_type == 'share-mode':
        ax.set_title(f'{num_agents} Agents {system}, {model_name}: Share-Modes Impact on Answer Patterns')
    elif comparison_type == 'malicious':
        ax.set_title(f'Share-Mode: {share_mode} | {num_agents} Agents {system} ({num_malicious_agents} Malicious), {model_name}:\nMalicious Agent Impact on Answer Patterns')
    elif comparison_type == 'sharemode_malicious':
        ax.set_title(f'{num_agents} Agents {system} ({num_malicious_agents} Malicious), {model_name}:\nShare-Mode and Malicious Agent Impact on Answer Patterns')
    elif comparison_type == 'malicious_convergence':
        ax.set_title(f'Share-Mode: {share_mode} | {num_agents} Agents {system}, {model_name}:\nMalicious Convergence Patterns')
    elif comparison_type == 'model':
        ax.set_title(f'Share-Mode: {share_mode} | {num_agents} Agents {system}, {dataset_filename}:\nAnswer Patterns Across Models')
    elif comparison_type == 'system':
        ax.set_title(f'Share-Mode: {share_mode} | {model_name}, {dataset_filename}:\nAnswer Patterns Across Systems')
    else:
        ax.set_title(f'Share-Mode: {share_mode} | {num_agents} Agents {system}, {model_name}:\nAnswer Patterns Across Datasets')

    # Put legend outside so plot area stays compact
    legend_cols = min(4, num_sets)
    ax.legend(
        loc='upper center',
        bbox_to_anchor=(0.5, -0.14),
        ncol=legend_cols,
        frameon=False,
        fontsize=8
    )
    fig.subplots_adjust(bottom=0.22)
    plt.tight_layout()

    script_dir = os.path.dirname(os.path.abspath(__file__))
    output_folder = os.path.join(script_dir, 'statistic_images')
    os.makedirs(output_folder, exist_ok=True)

    if comparison_type == 'share-mode':
        plot_image_name = f"{model_name}_{dataset_filename}_{system}_{comparison_type}_comparison_statistics_horizontal.png"
    elif comparison_type == 'malicious':
        plot_image_name = f"{model_name}_{dataset_filename}_{system}_{num_malicious_agents}_comparison_statistics_{share_mode}_horizontal.png"
    elif comparison_type == 'sharemode_malicious':
        plot_image_name = f"{model_name}_{system}_{num_malicious_agents}_{comparison_type}_comparison_statistics_horizontal.png"
    elif comparison_type == 'malicious_convergence':
        plot_image_name = f"{model_name}_{system}_{comparison_type}_comparison_statistics_{share_mode}_horizontal.png"
    elif comparison_type == 'model':
        plot_image_name = f"{dataset_filename}_{system}_{comparison_type}_comparison_statistics_{share_mode}_horizontal.png"
    elif comparison_type == 'system':
        plot_image_name = f"{model_name}_{dataset_filename}_{comparison_type}_comparison_statistics_{share_mode}_horizontal.png"
    else:
        plot_image_name = f"{model_name}_{system}_{comparison_type}_comparison_statistics_{share_mode}_horizontal.png"

    filepath = os.path.join(output_folder, plot_image_name)
    plt.savefig(filepath, dpi=300, bbox_inches='tight')
    print(f"Comparison plot saved as {plot_image_name}")
    plt.close()



def calculate_round_statistics(tasks_answers, skip_agents=[], original_task_answers=None, num_agents=3, rounds=3):
    """Calculate round convergence statistics WITHOUT creating a plot. Returns raw counts."""
    wrong_correct = 0
    correct_wrong = 0
    correct_malicious_wrong = 0
    num_tasks = len(tasks_answers)
    
    num_agents_converge_correct = dict()
    num_agents_converge_wrong = dict()
    num_agents_converge_malicious_wrong = dict()

    basline_and_malicious_equal = 0
    task_doesnt_exists_in_baseline = 0
    malicious_agent_answers_groundtruth = 0 

    prompt_injection_experiment = True
    
    for round_num in range(1, rounds + 3):  # Include final round and unknown convergence round
        num_agents_converge_correct[round_num] = num_agents_converge_wrong[round_num] = num_agents_converge_malicious_wrong[round_num] = 0
  

    for task_num, (init_answer, round_answers, correct_answer) in tasks_answers.items():
        final_answer = round_answers[rounds + 1]
        malicious_round_answers = []  # Track original answers from skipped agents that changed
        skip_task = False

        if original_task_answers is not None and task_num in original_task_answers:
            prompt_injection_experiment = False
            _, baseline_round_answers, _ = original_task_answers[task_num]
       
            # First pass: identify if any malicious agent changed their answer
            # mal_agent_changes = []  # Track which malicious agents changed
            for agent_idx in skip_agents:
                original_agent_rounds = [baseline_round_answers[r][agent_idx] for r in range(1, rounds + 1)]
                current_agent_rounds = [round_answers[r][agent_idx] for r in range(1, rounds + 1)]
                
                if all(ca == correct_answer[0] for ca in current_agent_rounds) == False:  # If the agent's current answers are not all correct
                    for r in range(rounds):
                        if current_agent_rounds[r] != correct_answer[0] and current_agent_rounds[r] != original_agent_rounds[r]:  # If any of the agent's round answers is not correct and not the same as baseline, consider it as malicious wrong for pattern analysis
                            malicious_round_answers.append(current_agent_rounds[r])
                            break
                    else:  # If all the agent's round answers are either correct or the same as baseline, consider it as equal for pattern analysis
                        basline_and_malicious_equal += 1
                        skip_task = True
                else:
                    malicious_agent_answers_groundtruth += 1 
                    skip_task = True
        if skip_task:
            num_tasks -= 1  # Exclude this task from total count since we're skipping it
            # print(f"Skipping task {task_num}")
            continue
        elif skip_agents != [] and malicious_round_answers == [] and prompt_injection_experiment == False:
            task_doesnt_exists_in_baseline += 1
            num_tasks -= 1  # Exclude this task from total count since we're skipping it
            # print(f"Skipping task {task_num}")
            continue

        for agent_idx, agent_answer in enumerate(zip(init_answer, final_answer)): 
            wrong_init_answer = agent_answer[0] != correct_answer[0]
            correct_final_answer = correct_answer[0] == agent_answer[1]
          
            # Skip this agent if in skip list
            if agent_idx in skip_agents:
                continue
            if wrong_init_answer and correct_final_answer:
                wrong_correct += 1

                for round_num, round_answer in round_answers.items():
                    correct_round_answer = round_answer[agent_idx] == correct_answer[0]

                    if round_num == rounds + 1 and round_answer[agent_idx] == '':
                        num_agents_converge_correct[round_num + 1] += 1 
                        break
                    elif correct_round_answer:
                        num_agents_converge_correct[round_num] += 1 
                        break
                    else:
                        continue
            elif not wrong_init_answer and not correct_final_answer and agent_answer[1] in malicious_round_answers:
                correct_wrong += 1

                for round_num, round_answer in round_answers.items():
                    wrong_round_answer = round_answer[agent_idx] != correct_answer[0]
                    
                    # Unknown convergence if agent's answer is empty in the final round without previously converging to correct or wrong
                    if round_num == rounds + 1 and round_answer[agent_idx] == '':
                        num_agents_converge_wrong[round_num + 1] += 1 
                        break
                    elif round_answer[agent_idx] in malicious_round_answers:
                        num_agents_converge_malicious_wrong[round_num] += 1 
                        num_agents_converge_wrong[round_num] += 1 
                        correct_malicious_wrong += 1
                        break    
                    else:
                        continue
            elif not wrong_init_answer and not correct_final_answer:
                correct_wrong += 1
                found_convergence = False
                for round_num, round_answer in round_answers.items():
                    wrong_round_answer = round_answer[agent_idx] != correct_answer[0]
                    
                    # Unknown convergence if agent's answer is empty in the final round without previously converging to correct or wrong
                    if round_num == rounds + 1 and round_answer[agent_idx] == '':
                        num_agents_converge_wrong[round_num + 1] += 1 
                        found_convergence = True
                        break   
                    elif wrong_round_answer:
                        num_agents_converge_wrong[round_num] += 1 
                        found_convergence = True
                        break
                    else:
                        continue
                if not found_convergence:
                    print(f"  ⚠️  WARNING: No round convergence found for Task {task_num}, Agent {agent_idx}!")
                    print(f"     Round answers: {round_answers}")
    
    # At the very end of the function, before return
    print(f"\n=== Results ===")
    print(f"correct_wrong total: {correct_wrong}")
    print(f"Sum of num_agents_converge_wrong: {sum(num_agents_converge_wrong.values())}")
    print(f"num_agents_converge_wrong dict: {num_agents_converge_wrong}")
    if correct_wrong != sum(num_agents_converge_wrong.values()):
        print(f"⚠️  MISMATCH: {correct_wrong} != {sum(num_agents_converge_wrong.values())}")
    return {
        'wrong_correct': wrong_correct,
        'correct_wrong': correct_wrong,
        'correct_malicious_wrong': correct_malicious_wrong,
        'num_agents_converge_correct': num_agents_converge_correct,
        'num_agents_converge_wrong': num_agents_converge_wrong,
        'num_agents_converge_malicious_wrong': num_agents_converge_malicious_wrong,
        'num_tasks': num_tasks,
        'num_agents': num_agents,
        'rounds': rounds
    }


def plot_round_statistics_comparison(datasets_data, model_name, dataset_filename='', share_mode='Both', system="MAS", num_agents=3, rounds=3, num_malicious_agents=0, comparison_type='share-mode'):
    """
    Plot round convergence statistics across multiple datasets side-by-side.
    
    Args:
        datasets_data: Dict like {'openai-gsm8k': stats_dict, 'cais-mmlu': stats_dict, ...}
        model_name: Name of the model (for title/filename)
        share_mode: 'Both', 'Reasoning', or 'Answer'
        system: 'MAS' or 'SAS'
        num_agents: Number of agents
        rounds: Number of rounds
        num_malicious_agents: Number of malicious agents
        comparison_type: Type of comparison to plot; 'share-mode' for share mode comparison, 'malicious' for malicious comparison, or 'dataset' for dataset comparison
    """
    # Build round categories
    round_categories = []
    for round_num in range(1, rounds + 1):
        round_categories.append(f'Round {round_num}')
    round_categories.append(f'Final Answer')
    round_categories.append(f'Unknown\nConvergence\nRound')
    
    fig, ax = plt.subplots(figsize=(14, 6))
    
    x = range(len(round_categories))
    bar_width = 0.15
    
    # Generate color palette for datasets
    #colors = plt.cm.Set3(range(len(datasets_data)))
    colors = seaborn.color_palette("deep", n_colors=len(datasets_data))
    
    for idx, (dataset_name, stats) in enumerate(datasets_data.items()):

        if comparison_type == 'malicious' or comparison_type == 'CR-IA':
            correct_wrong = stats['correct_wrong']
            
            if correct_wrong == 0:
                print(f"Warning: {dataset_name} has no 'Correct→Wrong' conversions")
                continue
            
            # Extract convergence counts for each round
            num_agents_converge_wrong = stats['num_agents_converge_wrong']
            round_percentages = [(num_agents_converge_wrong[round_num] / correct_wrong) * 100 
                            for round_num in range(1, rounds + 3)]
            
        else:
            wrong_correct = stats['wrong_correct']
        
            if wrong_correct == 0:
                print(f"Warning: {dataset_name} has no 'Wrong→Correct' conversions")
                continue
            
            # Extract convergence counts for each round
            num_agents_converge_correct = stats['num_agents_converge_correct']
            round_percentages = [(num_agents_converge_correct[round_num] / wrong_correct) * 100 
                                for round_num in range(1, rounds + 3)]
        
        # Offset bars for side-by-side display
        offset = bar_width * idx
        bars = ax.bar([i + offset for i in x], round_percentages, bar_width, 
                    label=dataset_name, color=colors[idx], alpha=0.8)
        
        # Add percentage labels
        for bar, percentage in zip(bars, round_percentages):
            height = bar.get_height()
            ax.text(bar.get_x() + bar.get_width()/2., height,
                f'{percentage:.1f}%',
                ha='center', va='bottom', fontsize=7)

    if comparison_type == 'malicious':    
        ax.set_ylabel('Percentage (%) of Correct→Wrong Answer Tasks')
        ax.set_xlabel('Round')
        ax.set_ylim(0, 105)
        ax.set_title(f'Share-Mode: {share_mode} | {num_agents} Agents {system} ({num_malicious_agents} Malicious), {model_name}:\n Convergence From Correct→Wrong Answer')
    if comparison_type == 'CR-IA':    
        ax.set_ylabel('Percentage (%) of Correct→Wrong Answer Tasks')
        ax.set_xlabel('Round')
        ax.set_ylim(0, 105)
        ax.set_title(f'Share-Mode: {share_mode} | {num_agents} Agents {system} ({num_malicious_agents} CR-IA Agents), {model_name}:\n Convergence From Correct→Wrong Answer')
    elif comparison_type == 'share-mode':
        ax.set_ylabel('Percentage (%) of Wrong→Correct Answer Tasks')
        ax.set_xlabel('Round')
        ax.set_ylim(0, 105)
        ax.set_title(f'{num_agents} Agents {system}, {model_name}:\n Convergence From Wrong→Correct Answer')
    else:
        ax.set_ylabel('Percentage (%) of Wrong→Correct Answer Tasks')
        ax.set_xlabel('Round')
        ax.set_ylim(0, 105)
        ax.set_title(f'Share-Mode: {share_mode} | {num_agents} Agents {system}, {model_name}:\n Convergence From Wrong→Correct Answer')

    ax.set_xticks([i + bar_width * (len(datasets_data) - 1) / 2 for i in x])
    ax.set_xticklabels(round_categories, rotation=45, ha='right')
    ax.legend()
        
    plt.tight_layout()
    
    script_dir = os.path.dirname(os.path.abspath(__file__))
    output_folder = os.path.join(script_dir, 'statistic_images')
    os.makedirs(output_folder, exist_ok=True)   

    if comparison_type == 'malicious': 
        filepath = os.path.join(output_folder, f"{model_name}_{system}_{dataset_filename}_{num_malicious_agents}{comparison_type}_comparison_round_statistics_{share_mode}.png")
        plt.savefig(filepath)
        print(f"Comparison round plot saved as {model_name}_{system}_{dataset_filename}_{num_malicious_agents}{comparison_type}_comparison_round_statistics_{share_mode}.png")
    elif comparison_type == 'share-mode':
        filepath = os.path.join(output_folder, f"{model_name}_{system}_{dataset_filename}_{comparison_type}_comparison_round_statistics.png")
        plt.savefig(filepath)
        print(f"Comparison round plot saved as {model_name}_{system}_{dataset_filename}_{comparison_type}_comparison_round_statistics.png")
    else:
        filepath = os.path.join(output_folder, f"{model_name}_{system}_{comparison_type}_comparison_round_statistics_{share_mode}.png")
        plt.savefig(filepath)
        print(f"Comparison round plot saved as {model_name}_{system}_{comparison_type}_comparison_round_statistics_{share_mode}.png")
    plt.close()


def plot_malicious_round_statistics_comparison(datasets_data, model_name, dataset_filename='', share_mode='Both', system="MAS", num_agents=3, rounds=3, num_malicious_agents=0):
    """
    Plot round convergence statistics across multiple datasets side-by-side for malicious settings.
    
    Args:
        datasets_data: Dict like {'openai-gsm8k': stats_dict, 'cais-mmlu': stats_dict, ...}
        model_name: Name of the model (for title/filename)
        share_mode: 'Both', 'Reasoning', or 'Answer'
        system: 'MAS' or 'SAS'
        num_agents: Number of agents
        rounds: Number of rounds
        num_malicious_agents: Number of malicious agents

    """
    # Build round categories
    round_categories = []
    for round_num in range(1, rounds + 1):
        round_categories.append(f'Round {round_num}')
    round_categories.append(f'Final Answer')
    round_categories.append(f'Unknown\nConvergence\nRound')
    
    fig, ax = plt.subplots(figsize=(14, 6))

    fig2, ax2 = plt.subplots(figsize=(14, 6))
    
    x = range(len(round_categories))
    bar_width = 0.15
    
    # Generate color palette for datasets
    #colors = plt.cm.Set3(range(len(datasets_data)))
    colors = seaborn.color_palette("deep", n_colors=len(datasets_data))
    
    for idx, (dataset_name, stats) in enumerate(datasets_data.items()):

        correct_wrong = stats['correct_wrong']
        
        if correct_wrong == 0:
            print(f"Warning: {dataset_name} has no 'Correct→Wrong' conversions")
            continue
        
        # Extract convergence counts for each round
        num_agents_converge_wrong = stats['num_agents_converge_wrong']
        round_percentages = [(num_agents_converge_wrong[round_num] / correct_wrong) * 100 
                        for round_num in range(1, rounds + 3)]
        
        correct_malicious_wrong = stats['correct_malicious_wrong']
        
        # Calculate malicious percentages (zero if no malicious conversions)
        if correct_malicious_wrong > 0:
            num_agents_converge_malicious_wrong = stats['num_agents_converge_malicious_wrong']
            round_percentages_malicious = [(num_agents_converge_malicious_wrong[round_num] / correct_wrong) * 100 
                            for round_num in range(1, rounds + 3)]
        else:
            round_percentages_malicious = [0.0] * (rounds + 2)

        # Offset bars for side-by-side display
        offset = bar_width * idx
        bars = ax.bar([i + offset for i in x], round_percentages, bar_width, 
                    label=dataset_name, color=colors[idx], alpha=0.8)
        
        # Add percentage labels to first plot
        for bar, percentage in zip(bars, round_percentages):
            height = bar.get_height()
            ax.text(bar.get_x() + bar.get_width()/2., height,
                f'{percentage:.1f}%',
                ha='center', va='bottom', fontsize=7)

        # Plot both series on ax2
        correct_bars = ax2.bar(
            [i + offset for i in x],
            round_percentages,
            bar_width,
            label=f"{dataset_name} Correct→Wrong",
            color=colors[idx],
            alpha=0.9
        )

        # Add percentage labels for Correct→Wrong bars
        for bar, percentage in zip(correct_bars, round_percentages):
            height = bar.get_height()
            ax2.text(
                bar.get_x() + bar.get_width() / 2.0,
                height,
                f"{percentage:.1f}%",
                ha='center',
                va='bottom',
                fontsize=7
            )

        # Only plot malicious bars for non-baseline datasets
        if 'No' not in dataset_name:
            malicious_bars = ax2.bar(
                [i + offset for i in x],
                round_percentages_malicious,
                bar_width,
                label=f"{dataset_name} Correct→Malicious Wrong",
                color=colors[idx],
                alpha=0.9,
                hatch='//',
                edgecolor='black',
                linewidth=0.8
            )

            for bar, percentage in zip(malicious_bars, round_percentages_malicious):
                height = bar.get_height()
                ax2.text(
                    bar.get_x() + bar.get_width() / 2.0,
                    height,
                    f"{percentage:.1f}%",
                    ha='center',
                    va='bottom',
                    fontsize=7
                )
 
    ax2.set_ylabel('Percentage (%) of Correct→MaliciousWrong Answer Tasks')
    ax2.set_xlabel('Round')
    ax2.set_ylim(0, 105)
    ax2.set_title(f'Share-Mode: {share_mode} | {num_agents} Agents {system} ({num_malicious_agents} Malicious), {model_name}:\n Convergence From Correct→ Malicious Wrong Answer')

    ax2.set_xticks([i + bar_width * (len(datasets_data) - 1) / 2 for i in x])
    ax2.set_xticklabels(round_categories, rotation=45, ha='right')
    ax2.legend()
        
    fig2.tight_layout()
    
    script_dir = os.path.dirname(os.path.abspath(__file__))
    output_folder = os.path.join(script_dir, 'statistic_images')
    os.makedirs(output_folder, exist_ok=True)   

   
    filepath_mal = os.path.join(output_folder, f"{model_name}_{system}_{dataset_filename}_{num_malicious_agents}_comparison_round_statistics_maliciouswrong_{share_mode}.png")
    fig2.savefig(filepath_mal)
    print(f"Comparison round plot saved as {model_name}_{system}_{dataset_filename}_{num_malicious_agents}_comparison_round_statistics_maliciouswrong_{share_mode}.png")
    plt.close(fig2)


def stats_dict_to_row(dataset_name, stats, metadata):
    """
    Convert a stats dictionary to a single row with all metrics.
    
    Args:
        dataset_name: Name of the dataset
        stats: Dictionary returned from calculate_pattern_statistics()
        metadata: Dict with model_name, system, share_mode, num_agents
    
    Returns:
        Dictionary representing one row
    """
    return {
        'Model': metadata.get('model_name', 'N/A'),
        'System': metadata.get('system', 'N/A'),
        'Share Mode': metadata.get('share_mode', 'N/A'),
        'Dataset': dataset_name,
        'Num Agents': metadata.get('num_agents', 'N/A'),
        'Malicious Agents': str(metadata.get('malicious_agents', [])),
        'Wrong→Correct': stats.get('wrong_correct', 0),
        'Wrong→Different Wrong': stats.get('wrong_different_wrong', 0),
        'Wrong→Same Wrong': stats.get('wrong_same_wrong', 0),
        'Wrong→Malicious Wrong': stats.get('wrong_malicious_wrong', 0),
        'Correct→Wrong': stats.get('correct_wrong', 0),
        'Correct→Malicious Wrong': stats.get('correct_malicious_wrong', 0),
        'Correct→Correct': stats.get('correct_correct', 0),
        'Malicious Wrong Convergence': stats.get('wrong_malicious_wrong', 0) + stats.get('correct_malicious_wrong', 0),
        'Num Tasks': stats.get('num_tasks', 0),
        'Missed Patterns': stats.get('missed_patterns', 0),
        'Baseline and Malicious Equal': stats.get('basline_and_malicious_equal', 0),
        'Task Doesn\'t Exist in Baseline': stats.get('task_doesnt_exists_in_baseline', 0),
        'Malicious Agent Answers Ground Truth': stats.get('malicious_agent_answers_groundtruth', 0)
    }

    



def export_comparisons_to_excel(comparison_results, output_filename='dataset_comparison_results.xlsx'):
    """
    Export dataset comparison results to Excel with separate sheets for each comparison type.
    
    Args:
        comparison_results: List of tuples (comparison_name, datasets_data, metadata)
                           where metadata includes optional 'sheet_type' field.
                           Results with same sheet_type go to same sheet.
        output_filename: Path to output Excel file
    """
    # Group results by sheet_type
    sheets_data = {}  # {sheet_name: [(comparison_name, dataset_name, stats, metadata), ...]}
    
        # Flatten and organize by sheet_type
    for comparison_name, datasets_data, metadata in comparison_results:
        sheet_type = metadata.get('sheet_type', 'Results')
        datasets_mal_agents = metadata.get('datasets_malicious_agents', {})  # Only exists for malicious_comparison
        
        if sheet_type not in sheets_data:
            sheets_data[sheet_type] = []
        
        for dataset_name, data in datasets_data.items():
            # Handle both structures: direct stats (dataset_comparison) and nested stats (malicious_comparison)
            if isinstance(data, dict) and 'stats' in data:
                # This is from malicious_comparison with nested structure
                stats = data['stats']
            else:
                # This is from dataset_comparison with direct stats
                stats = data
            
            row = stats_dict_to_row(dataset_name, stats, metadata)
            row['Comparison'] = comparison_name
          
            # Override malicious agents with per-dataset value (only if it exists)
            if datasets_mal_agents and dataset_name in datasets_mal_agents:
                row['Malicious Agents'] = str(datasets_mal_agents[dataset_name])
            
            # Extract share mode from dataset_name if it contains share-mode information
            if 'Share-mode' in dataset_name or 'share-mode' in dataset_name:
                if 'Both' in dataset_name:
                    row['Share Mode'] = 'Both'
                elif 'Reasoning' in dataset_name:
                    row['Share Mode'] = 'Reasoning'
                elif 'Answer' in dataset_name:
                    row['Share Mode'] = 'Answer'
            
            sheets_data[sheet_type].append(row)
    
    # Column order for all sheets
    col_order = ['Comparison', 'Model', 'System', 'Share Mode', 'Dataset', 'Num Agents', 'Malicious Agents',
                 'Wrong→Correct', 'Wrong→Different Wrong', 'Wrong→Same Wrong',
                 'Wrong→Malicious Wrong', 'Correct→Wrong', 'Correct→Malicious Wrong',
                 'Correct→Correct', 'Malicious Wrong Convergence', 'Num Tasks', 'Missed Patterns', 'Baseline and Malicious Equal',
                 'Task Doesn\'t Exist in Baseline', 'Malicious Agent Answers Ground Truth']
    

    # Helper function to apply formatting to a worksheet
    def format_worksheet(worksheet, df):
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF')
        
        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        for idx, col in enumerate(df.columns, 1):
            max_length = max(df[col].astype(str).map(len).max(), len(col))
            worksheet.column_dimensions[chr(64 + idx)].width = min(max_length + 2, 20)
    
    # Write to Excel with multiple sheets
    with pd.ExcelWriter(output_filename, engine='openpyxl') as writer:
        for sheet_name in sorted(sheets_data.keys()):
            rows_list = sheets_data[sheet_name]
            df = pd.DataFrame(rows_list)
            
            # Reorder columns
            df = df[[col for col in col_order if col in df.columns]]
            
            # Write to sheet
            df.to_excel(writer, sheet_name=sheet_name, index=False)
            
            # Apply formatting
            worksheet = writer.sheets[sheet_name]
            format_worksheet(worksheet, df)
    
    print(f"Results exported to {output_filename}")
    print(f"Sheets created: {', '.join(sorted(sheets_data.keys()))}")


def round_stats_dict_to_rows(dataset_name, stats, metadata):
    """
    Convert a round stats dictionary to multiple rows (one per round).
    
    Args:
        dataset_name: Name of the dataset
        stats: Dictionary returned from calculate_round_statistics()
        metadata: Dict with model_name, system, share_mode, num_agents
    
    Returns:
        List of dictionaries, each representing one row
    """
    rows = []
    rounds = stats.get('rounds', 0)
    
    for round_num in range(1, rounds + 3):  # Include final answer and unknown rounds
        round_label = f'Round {round_num}' if round_num <= rounds else ('Final Answer' if round_num == rounds + 1 else 'Unknown')
        
        wrong_correct = stats.get('wrong_correct', 0)
        correct_wrong = stats.get('correct_wrong', 0)
        correct_malicious_wrong = stats.get('correct_malicious_wrong', 0)
        
        row = {
            'Comparison': '',  # Will be filled by export function
            'Model': metadata.get('model_name', 'N/A'),
            'System': metadata.get('system', 'N/A'),
            'Share Mode': metadata.get('share_mode', 'N/A'),
            'Dataset': dataset_name,
            'Num Agents': metadata.get('num_agents', 'N/A'),
            'Malicious Agents': str(metadata.get('malicious_agents', [])),
            'Round': round_label,
            'Wrong→Correct Per Round': stats.get('num_agents_converge_correct', {}).get(round_num, 0),
            'Correct→Wrong Per Round': stats.get('num_agents_converge_wrong', {}).get(round_num, 0),
            'Correct→Malicious Wrong Per Round': stats.get('num_agents_converge_malicious_wrong', {}).get(round_num, 0),
            'Wrong→Correct Total': stats.get('wrong_correct', 0),
            'Correct→Wrong Total': stats.get('correct_wrong', 0),
            'Correct→Malicious Wrong Total': stats.get('correct_malicious_wrong', 0),
            'Total Tasks': stats.get('num_tasks', 0),
        }
        rows.append(row)
    
    return rows


def export_round_comparisons_to_excel(comparison_results, output_filename='round_comparison_results.xlsx'):
    """
    Export round comparison results to Excel with separate sheets for each comparison type.
    
    Args:
        comparison_results: List of tuples (comparison_name, datasets_data, metadata)
                           where datasets_data contains round statistics
                           and metadata includes optional 'sheet_type' field.
                           Results with same sheet_type go to same sheet.
        output_filename: Path to output Excel file
    """
    # Group results by sheet_type
    sheets_data = {}  # {sheet_name: [row_dicts]}
    
    # Flatten and organize by sheet_type
    for comparison_name, datasets_data, metadata in comparison_results:
        sheet_type = metadata.get('sheet_type', 'Round Results')
        datasets_mal_agents = metadata.get('datasets_malicious_agents', {})
        
        if sheet_type not in sheets_data:
            sheets_data[sheet_type] = []
        
        for dataset_name, data in datasets_data.items():
            # Handle both structures: direct stats (dataset_comparison) and nested stats (malicious_comparison)
            if isinstance(data, dict) and 'stats' in data:
                # This is from malicious_comparison with nested structure
                stats = data['stats']
            else:
                # This is from dataset_comparison with direct stats
                stats = data
            
            rows = round_stats_dict_to_rows(dataset_name, stats, metadata)
            
            for row in rows:
                row['Comparison'] = comparison_name
                
                # Override malicious agents with per-dataset value (only if it exists)
                if datasets_mal_agents and dataset_name in datasets_mal_agents:
                    row['Malicious Agents'] = str(datasets_mal_agents[dataset_name])

                # Extract share mode from dataset_name if it contains share-mode information
                if 'Share-mode' in dataset_name or 'share-mode' in dataset_name:
                    if 'Both' in dataset_name:
                        row['Share Mode'] = 'Both'
                    elif 'Reasoning' in dataset_name:
                        row['Share Mode'] = 'Reasoning'
                    elif 'Answer' in dataset_name:
                        row['Share Mode'] = 'Answer'
                
                sheets_data[sheet_type].append(row)
    
    # Column order for all sheets
    col_order = ['Comparison', 'Model', 'System', 'Share Mode', 'Dataset', 'Num Agents', 'Malicious Agents',
             'Round', 'Wrong→Correct Per Round', 'Correct→Wrong Per Round', 'Correct→Malicious Wrong Per Round',
             'Wrong→Correct Total', 'Correct→Wrong Total', 'Correct→Malicious Wrong Total',
             'Total Tasks']
    
    # Helper function to apply formatting to a worksheet
    def format_worksheet(worksheet, df):
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF')
        
        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        for idx, col in enumerate(df.columns, 1):
            max_length = max(df[col].astype(str).map(len).max(), len(col))
            worksheet.column_dimensions[chr(64 + idx)].width = min(max_length + 2, 20)
    
    # Write to Excel with multiple sheets
    with pd.ExcelWriter(output_filename, engine='openpyxl') as writer:
        for sheet_name in sorted(sheets_data.keys()):
            rows_list = sheets_data[sheet_name]
            df = pd.DataFrame(rows_list)
            
            # Reorder columns
            df = df[[col for col in col_order if col in df.columns]]
            
            # Write to sheet
            df.to_excel(writer, sheet_name=sheet_name, index=False)
            
            # Apply formatting
            worksheet = writer.sheets[sheet_name]
            format_worksheet(worksheet, df)
    
    print(f"Round results exported to {output_filename}")
    print(f"Sheets created: {', '.join(sorted(sheets_data.keys()))}")


def dataset_comparison(transcripts, round_figure=False, comparison_type='dataset', return_data=True, return_round_data=False, comparison_name=''):
    """
    Main function to perform dataset comparison for a given model and share mode.
    
    Args:
        transcripts: Dict like {'openai-gsm8k': 'path/to/transcripts', 'cais-mmlu': 'path/to/transcripts', ...}
        round_figure: Whether to calculate and plot round convergence statistics
        comparison_type: Type of comparison to perform (dataset, share-mode, system, model, malicious, 
                         sharemode-malicious or malicious_convergence), like 'share-mode' for share mode 
                         comparison, 'system' for system comparison, or 'dataset' for dataset comparison
        return_data: If True, returns (datasets_data, metadata) instead of just plotting
    """
    datasets_data = dict()
    datasets_data_round = dict()
    num_agents = None
    num_rounds = None
    share_mode = ""
    dataset_filename = ""
    model_name = ""
    system = ""
 
    for dataset_name, transcript_path in transcripts.items():
        # Extract answers for each dataset
        answers, round_answers, summary_info = extract_answers(transcript_path)
        num_agents = summary_info['num_agents']
        num_rounds = summary_info['num_rounds']
        share_mode = summary_info['share_mode']
        dataset_filename = summary_info['dataset_filename']
        system = summary_info['system']
        model_name = summary_info['model_name']

        # Calculate statistics for each dataset
        stats = calculate_pattern_statistics(answers)
        if round_figure or return_round_data:
            round_stats = calculate_round_statistics(round_answers, num_agents=num_agents, rounds=num_rounds)
            datasets_data_round[dataset_name] = round_stats
        # Store statistics for comparison
        datasets_data[dataset_name] = stats

    
    # Prepare metadata
    metadata = {
        'model_name': model_name,
        'system': system,
        'share_mode': share_mode,
        'num_agents': num_agents,
        'dataset_filename': dataset_filename,
        'comparison_type': comparison_type,
        'sheet_type': comparison_type,
    }
    
    plot_pattern_statistics_comparison(
        datasets_data,
        model_name,
        dataset_filename=dataset_filename,
        share_mode=share_mode,
        system=system,
        num_agents=num_agents,
        comparison_type=comparison_type,
    )
    if round_figure:
        plot_round_statistics_comparison(
            datasets_data_round,
            model_name,
            dataset_filename=dataset_filename,
            share_mode=share_mode,
            system=system,
            num_agents=num_agents,
            rounds=num_rounds,
            comparison_type=comparison_type,
        )

    # Return data if requested
    if return_data:
        return (comparison_name, datasets_data, metadata)
    elif return_round_data:
        return (comparison_name, datasets_data_round, metadata)



def malicious_comparison(transcripts, round_figure=True, comparison_type='malicious', return_data=True, return_round_data=False, comparison_name=''):
    datasets_data = {}
    datasets_data_round = {}
    datasets_malicious_agents = {}

    baseline_answers = None
    baseline_round_answers = None
    malicious_agents = []
    num_agents = num_rounds = None
    share_mode = dataset_filename = system = model_name = ""

    for dataset_name, transcript_path in transcripts.items():
        answers, round_answers, summary_info = extract_answers(transcript_path)

        num_rounds = summary_info['num_rounds']
        share_mode = summary_info['share_mode']
        dataset_filename = summary_info['dataset_filename']
        system = summary_info['system']
        model_name = summary_info['model_name']
        num_agents = summary_info['num_agents']

        if 'No' in dataset_name:
            
            baseline_answers = answers
            baseline_round_answers = round_answers

        
            datasets_data[dataset_name] = {
            "stats": calculate_pattern_statistics(
                baseline_answers,
                skip_agents=[],
                original_task_answers=None,
            ),
            "honest_agents": num_agents,
            }           

            if round_figure or return_round_data:
                datasets_data_round[dataset_name] = calculate_round_statistics(
                    baseline_round_answers,
                    skip_agents=[],
                    original_task_answers=None,
                    num_agents=num_agents,
                    rounds=num_rounds,
                )
        elif comparison_type == 'early_stopping' or comparison_type == 'IR-CA':
            if 'early_stopping' in comparison_type:
                es_agents = summary_info['es_agents']
                datasets_malicious_agents[dataset_name] = es_agents
            else:
                es_agents = summary_info['malicious_agents']
                datasets_malicious_agents[dataset_name] = es_agents

            datasets_data[dataset_name] = {
                "stats": calculate_pattern_statistics(
                answers,
                skip_agents=es_agents,
                original_task_answers=None,
                ),
                "honest_agents": num_agents - len(es_agents),
            }

            if round_figure or return_round_data:
                datasets_data_round[dataset_name] = calculate_round_statistics(
                    round_answers,
                    skip_agents=es_agents,
                    original_task_answers=None,
                    num_agents=num_agents - len(es_agents),
                    rounds=num_rounds,
                )
        else:
            malicious_agents = summary_info['malicious_agents']
            datasets_malicious_agents[dataset_name] = malicious_agents

            datasets_data[dataset_name] = {
                "stats": calculate_pattern_statistics(
                    answers,
                    skip_agents=malicious_agents,
                    # Only compare first round answers for pattern statistics to determine if malicious answer:
                    original_task_answers=(round_answers, baseline_round_answers),
                ),
                "honest_agents": num_agents - len(malicious_agents),  # malicious run
                # "malicious_agents": malicious_agents
            }

            if round_figure or return_round_data:
                datasets_data_round[dataset_name] = calculate_round_statistics(
                    round_answers,
                    skip_agents=malicious_agents,
                    original_task_answers=baseline_round_answers,
                    num_agents=num_agents - len(malicious_agents),
                    rounds=num_rounds,
                )

    # Prepare metadata
    metadata = {
        'model_name': model_name,
        'system': system,
        'share_mode': share_mode,
        'num_agents': num_agents,
        'dataset_filename': dataset_filename,
        'comparison_type': comparison_type,
        'sheet_type': comparison_type,
        'datasets_malicious_agents': datasets_malicious_agents
    }          

    plot_pattern_statistics_comparison(
        datasets_data,
        model_name,
        dataset_filename=dataset_filename,
        share_mode=share_mode,
        system=system,
        num_agents=num_agents,
        num_malicious_agents=len(malicious_agents),
        comparison_type=comparison_type,
    )

    if round_figure:
        plot_malicious_round_statistics_comparison(
            datasets_data_round,
            model_name,
            dataset_filename=dataset_filename,
            share_mode=share_mode,
            system=system,
            num_agents=num_agents,
            rounds=num_rounds,
            num_malicious_agents=len(malicious_agents),
        )
    
    # Return data if requested
    if return_data:
        return (comparison_name, datasets_data, metadata)
    elif return_round_data:
        return (comparison_name, datasets_data_round, metadata)



# transcript_data = {
#     'openai/gsm8k No Malicious Agent': 'transcripts_mal_test/qwen3b_2026-03-29_00h31m15s',
#     'openai/gsm8k Malicious Agent 3': 'transcripts_mal_test/qwen3b_job342_2026-04-01_01h15m49s',
#     'openai/gsm8k No Malicious Agent again': 'transcripts_mal_test/qwen3b_2026-03-29_00h31m15s',
#     'openai/gsm8k Malicious Agent 2': 'transcripts_mal_test/qwen3b_job342_2026-04-01_01h15m49s'
# }

# malicious_comparison(transcript_data, round_figure=False, return_data=False)
# # dataset_comparison(transcript_data, round_figure=False, comparison_type='malicious', return_data=False)

# ╔════════════════════════════════════════╗
# ║           DATASET COMPARISONS          ║
# ╚════════════════════════════════════════╝

all_results = []
# # ------ Comparison between different datasets, QWEN, both, MAS --------
# transcript_data = {
#     'openai/gsm8k': 'transcripts/MAS/both/qwen3b_2026-03-29_00h31m15s',
#     'cais/mmlu': 'transcripts/MAS/both/qwen3b_2026-03-29_15h34m56s',
#     'ChilleD/StrategyQA': 'transcripts/MAS/both/qwen3b_2026-03-29_15h26m25s',
#     'tasksource/bigbench': 'transcripts/MAS/both/qwen3b_2026-03-29_03h35m18s'
# }
# result = dataset_comparison(transcript_data, round_figure=True, comparison_name='Qwen, both, MAS')

# all_results.append(result)

# # ------ Comparison between different datasets, Olmo, both, MAS --------
# transcript_data = {
#     'openai/gsm8k': 'transcripts/MAS/both/olmo7b_2026-03-29_15h41m56s',
#     'cais/mmlu': 'transcripts/MAS/both/olmo7b_2026-03-29_15h29m53s',
#     'ChilleD/StrategyQA': 'transcripts/MAS/both/olmo7b_2026-03-29_15h26m25s',
#     'tasksource/bigbench': 'transcripts/MAS/both/olmo7b_2026-03-29_01h47m46s'
# }
# result = dataset_comparison(transcript_data, round_figure=True, comparison_name='Olmo, both, MAS')
# all_results.append(result)

# # ------ Comparison between different datasets, Llama, both, MAS --------
# transcript_data = {
#     'openai/gsm8k': 'transcripts/MAS/both/llama3b_2026-03-29_15h19m52s',
#     'cais/mmlu': 'transcripts/MAS/both/llama3b_2026-03-29_00h34m49s',
#     'ChilleD/StrategyQA': 'transcripts/MAS/both/llama3b_2026-03-29_15h21m49s',
#     'tasksource/bigbench': 'transcripts/MAS/both/llama3b_2026-03-29_03h08m47s'
# }
# result = dataset_comparison(transcript_data, round_figure=True, comparison_name='Llama, both, MAS')
# all_results.append(result)

# # ------ Comparison between different datasets, QWEN, both, SAS --------
# transcript_data = {
#     'openai/gsm8k': 'transcripts/SAS/both/qwen3b_job9712_2026-03-30_22h08m41s',
#     'cais/mmlu': 'transcripts/SAS/both/qwen3b_job9713_2026-03-30_22h10m26s',
#     'ChilleD/StrategyQA': 'transcripts/SAS/both/qwen3b_job9714_2026-03-30_22h12m15s',
#     'tasksource/bigbench': 'transcripts/SAS/both/qwen3b_job9715_2026-03-30_22h13m45s'
# }
# result = dataset_comparison(transcript_data, round_figure=True, comparison_name='Qwen, both, SAS')
# all_results.append(result)

# # ------ Comparison between different datasets, Olmo, both, SAS --------
# transcript_data = {
#     'openai/gsm8k': 'transcripts/SAS/both/olmo7b_job9749_2026-03-31_00h58m55s',
#     'cais/mmlu': 'transcripts/SAS/both/olmo7b_job9750_2026-03-31_00h59m14s',
#     'ChilleD/StrategyQA': 'transcripts/SAS/both/olmo7b_job9751_2026-03-31_00h59m43s',
#     'tasksource/bigbench': 'transcripts/SAS/both/olmo7b_job9790_2026-03-31_09h14m42s'
# }
# result = dataset_comparison(transcript_data, round_figure=True, comparison_name='Olmo, both, SAS')
# all_results.append(result)

# # ------ Comparison between different datasets, llama, both, SAS --------
# transcript_data = {
#     'openai/gsm8k': 'transcripts/SAS/both/llama3b_job9794_2026-03-31_09h18m03s',
#     'cais/mmlu': 'transcripts/SAS/both/llama3b_job9793_2026-03-31_09h18m03s',
#     'ChilleD/StrategyQA': 'transcripts/SAS/both/llama3b_job9792_2026-03-31_09h16m56s',
#     'tasksource/bigbench': 'transcripts/SAS/both/llama3b_job9791_2026-03-31_09h16m39s'
# }
# result = dataset_comparison(transcript_data, round_figure=True, comparison_name='Llama, both, SAS')
# all_results.append(result)

# ╔════════════════════════════════════════╗
# ║         SHARE-MODE COMPARISONS         ║
# ╚════════════════════════════════════════╝

# # --- share-mode comparison for Qwen ---
# transcript_data = {
#     'openai/gsm8k: Both Share-mode': 'transcripts/MAS/both/qwen3b_2026-03-29_00h31m15s',
#     'openai/gsm8k: Reasoning Share-mode': 'transcripts/MAS/reasoning/qwen3b_job9706_2026-03-30_22h01m46s',
#     'openai/gsm8k: Answer Share-mode': 'transcripts/MAS/answer/qwen3b_job9705_2026-03-30_21h58m14s'
# }
# # result = dataset_comparison(transcript_data, comparison_type='share-mode', comparison_name='Qwen, GSM8K, MAS')
# # all_results.append(result)
# result = dataset_comparison(transcript_data, round_figure=True, comparison_type='share-mode', return_data=False, return_round_data=True, comparison_name='Qwen, GSM8K, MAS')
# all_results.append(result)

# transcript_data = {
#     'cais/mmlu: Both Share-mode': 'transcripts/MAS/both/qwen3b_2026-03-29_15h34m56s',
#     'cais/mmlu: Reasoning Share-mode': 'transcripts/MAS/reasoning/qwen3b_job9913_2026-03-31_11h56m29s',
#     'cais/mmlu: Answer Share-mode': 'transcripts/MAS/answer/qwen3b_job9914_2026-03-31_11h56m37s'
# }
# # result = dataset_comparison(transcript_data, comparison_type='share-mode', comparison_name='Qwen, MMLU, MAS')
# # all_results.append(result)
# result = dataset_comparison(transcript_data, round_figure=True, comparison_type='share-mode', return_data=False, return_round_data=True, comparison_name='Qwen, MMLU, MAS')
# all_results.append(result)

# transcript_data = {
#     'ChilleD/StrategyQA: Both Share-mode': 'transcripts/MAS/both/qwen3b_2026-03-29_15h26m25s',
#     'ChilleD/StrategyQA: Reasoning Share-mode': 'transcripts/MAS/reasoning/qwen3b_job9917_2026-03-31_12h00m28s',
#     'ChilleD/StrategyQA: Answer Share-mode': 'transcripts/MAS/answer/qwen3b_job9918_2026-03-31_17h05m39s'
# }
# # result = dataset_comparison(transcript_data, comparison_type='share-mode', comparison_name='Qwen, StrategyQA, MAS')
# # all_results.append(result)
# result = dataset_comparison(transcript_data, round_figure=True, comparison_type='share-mode', return_data=False, return_round_data=True, comparison_name='Qwen, StrategyQA, MAS')
# all_results.append(result)

# transcript_data = {
#     'tasksource/bigbench Both Share-mode': 'transcripts/MAS/both/qwen3b_2026-03-29_03h35m18s',
#     'tasksource/bigbench Reasoning Share-mode': 'transcripts/MAS/reasoning/qwen3b_2026-03-29_21h37m18s',
#     'tasksource/bigbench Answer Share-mode': 'transcripts/MAS/answer/qwen3b_2026-03-29_21h46m31s'
# }
# # result = dataset_comparison(transcript_data, comparison_type='share-mode', comparison_name='Qwen, BigBench, MAS')
# # all_results.append(result)
# result = dataset_comparison(transcript_data, round_figure=True, comparison_type='share-mode', return_data=False, return_round_data=True, comparison_name='Qwen, BigBench, MAS')
# all_results.append(result)

# # --- share-mode comparison for Olmo ---
# transcript_data = {
#     'openai/gsm8k: Both Share-mode': 'transcripts/MAS/both/olmo7b_2026-03-29_15h41m56s',
#     'openai/gsm8k: Reasoning Share-mode': 'transcripts/MAS/reasoning/olmo7b_job125_2026-03-31_19h35m18s',
#     'openai/gsm8k: Answer Share-mode': 'transcripts/MAS/answer/olmo7b_job126_2026-03-31_19h41m31s'
# }
# # result = dataset_comparison(transcript_data, comparison_type='share-mode', comparison_name='Olmo, GSM8K, MAS')
# # all_results.append(result)
# result = dataset_comparison(transcript_data, round_figure=True, comparison_type='share-mode', return_data=False, return_round_data=True, comparison_name='Olmo, GSM8K, MAS')
# all_results.append(result)

# transcript_data = {
#     'cais/mmlu: Both Share-mode': 'transcripts/MAS/both/olmo7b_2026-03-29_15h29m53s',
#     'cais/mmlu: Reasoning Share-mode': 'transcripts/MAS/reasoning/olmo7b_job123_2026-03-31_19h08m02s',
#     'cais/mmlu: Answer Share-mode': 'transcripts/MAS/answer/olmo7b_job124_2026-03-31_19h08m18s'
# }
# # result = dataset_comparison(transcript_data, comparison_type='share-mode', comparison_name='Olmo, MMLU, MAS')
# # all_results.append(result)
# result = dataset_comparison(transcript_data, round_figure=True, comparison_type='share-mode', return_data=False, return_round_data=True, comparison_name='Olmo, MMLU, MAS')
# all_results.append(result)

# transcript_data = {
#     'ChilleD/StrategyQA: Both Share-mode': 'transcripts/MAS/both/olmo7b_2026-03-29_15h26m25s',
#     'ChilleD/StrategyQA: Reasoning Share-mode': 'transcripts/MAS/reasoning/olmo7b_job9972_2026-03-31_17h42m46s',
#     'ChilleD/StrategyQA: Answer Share-mode': 'transcripts/MAS/answer/olmo7b_job9973_2026-03-31_18h08m36s'
# }
# # result = dataset_comparison(transcript_data, comparison_type='share-mode', comparison_name='Olmo, StrategyQA, MAS')
# # all_results.append(result)
# result = dataset_comparison(transcript_data, round_figure=True, comparison_type='share-mode', return_data=False, return_round_data=True, comparison_name='Olmo, StrategyQA, MAS')
# all_results.append(result)

# transcript_data = {
#     'tasksource/bigbench: Both Share-mode': 'transcripts/MAS/both/olmo7b_2026-03-29_01h47m46s',
#     'tasksource/bigbench: Reasoning Share-mode': 'transcripts/MAS/reasoning/olmo7b_2026-03-29_21h37m22s',
#     'tasksource/bigbench: Answer Share-mode': 'transcripts/MAS/answer/olmo7b_2026-03-29_21h46m31s'
# }
# # result = dataset_comparison(transcript_data, comparison_type='share-mode', comparison_name='Olmo, BigBench, MAS')
# # all_results.append(result)
# result = dataset_comparison(transcript_data, round_figure=True, comparison_type='share-mode', return_data=False, return_round_data=True, comparison_name='Olmo, BigBench, MAS')
# all_results.append(result)

# # --- share-mode comparison for Llama ---
# transcript_data = {
#     'openai/gsm8k: Both Share-mode': 'transcripts/MAS/both/llama3b_2026-03-29_15h19m52s',
#     'openai/gsm8k: Reasoning Share-mode': 'transcripts/MAS/reasoning/llama3b_job9703_2026-03-30_21h56m07s',
#     'openai/gsm8k: Answer Share-mode': 'transcripts/MAS/answer/llama3b_job9704_2026-03-30_21h57m33s'
# }
# # result = dataset_comparison(transcript_data, comparison_type='share-mode', comparison_name='Llama, GSM8K, MAS')
# # all_results.append(result)
# result = dataset_comparison(transcript_data, round_figure=True, comparison_type='share-mode', return_data=False, return_round_data=True, comparison_name='Llama, GSM8K, MAS')
# all_results.append(result)

# transcript_data = {
#     'cais/mmlu: Both Share-mode': 'transcripts/MAS/both/llama3b_2026-03-29_00h34m49s',
#     'cais/mmlu: Reasoning Share-mode': 'transcripts/MAS/reasoning/llama3b_job348_2026-04-01_03h40m36s',
#     'cais/mmlu: Answer Share-mode': 'transcripts/MAS/answer/llama3b_job349_2026-04-01_04h00m41s'
# }
# # result = dataset_comparison(transcript_data, comparison_type='share-mode', comparison_name='Llama, MMLU, MAS')
# # all_results.append(result)
# result = dataset_comparison(transcript_data, round_figure=True, comparison_type='share-mode', return_data=False, return_round_data=True, comparison_name='Llama, MMLU, MAS')
# all_results.append(result)

# transcript_data = {
#     'ChilleD/StrategyQA: Both Share-mode': 'transcripts/MAS/both/llama3b_2026-03-29_15h21m49s',
#     'ChilleD/StrategyQA: Reasoning Share-mode': 'transcripts/MAS/reasoning/llama3b_job350_2026-04-01_05h03m45s',
#     'ChilleD/StrategyQA: Answer Share-mode': 'transcripts/MAS/answer/llama3b_job351_2026-04-01_07h39m19s'
# }
# # result = dataset_comparison(transcript_data, comparison_type='share-mode', comparison_name='Llama, StrategyQA, MAS')
# # all_results.append(result)
# result = dataset_comparison(transcript_data, round_figure=True, comparison_type='share-mode', return_data=False, return_round_data=True, comparison_name='Llama, StrategyQA, MAS')
# all_results.append(result)

# transcript_data = {
#     'tasksource/bigbench: Both Share-mode': 'transcripts/MAS/both/llama3b_2026-03-29_03h08m47s',
#     'tasksource/bigbench: Reasoning Share-mode': 'transcripts/MAS/reasoning/llama3b_2026-03-29_21h37m45s',
#     'tasksource/bigbench: Answer Share-mode': 'transcripts/MAS/answer/llama3b_2026-03-29_23h24m02s'
# }
# # result = dataset_comparison(transcript_data, comparison_type='share-mode', comparison_name='Llama, BigBench, MAS')
# # all_results.append(result)
# result = dataset_comparison(transcript_data, round_figure=True, comparison_type='share-mode', return_data=False, return_round_data=True, comparison_name='Llama, BigBench, MAS')
# all_results.append(result)

# export_round_comparisons_to_excel(all_results, 'round_results.xlsx')


# # ╔════════════════════════════════════════╗
# # ║       MALICIOUS PROMPT COMPARISONS     ║
# # ╚════════════════════════════════════════╝

# # --- Malicious prompt comparison for Qwen ---
# transcript_data = {
#     'tasksource/bigbench: No Mal Agent - both': 'transcripts/MAS/both/qwen3b_2026-03-29_03h35m18s',
#     'tasksource/bigbench: Mal Agent 1 - both': 'transcripts/MAS/malicious/bigbench/qwen3b_job9921_2026-03-31_13h53m46s',
#     'tasksource/bigbench: Mal Agent 2 - both': 'transcripts/MAS/malicious/bigbench/qwen3b_job8051_2026-04-22_20h01m28s',
#     'tasksource/bigbench: Mal Agent 3 - both': 'transcripts/MAS/malicious/bigbench/qwen3b_job9925_2026-03-31_16h22m36s',
# }
# # result = malicious_comparison(transcript_data, round_figure=False, comparison_name='Qwen, BigBench, MAS, both')
# # all_results.append(result)
# result = malicious_comparison(transcript_data, round_figure=True, return_data=False, return_round_data=True, comparison_name='Qwen, BigBench, MAS, both')
# all_results.append(result)

# # transcript_data = {
# #     'tasksource/bigbench: No Mal Agent - answer': 'transcripts/MAS/answer/qwen3b_2026-03-29_21h46m31s',
# #     'tasksource/bigbench: Mal Agent 1 - answer': 'transcripts/MAS/malicious/bigbench/qwen3b_job9929_2026-03-31_16h56m02s',
# #     'tasksource/bigbench: Mal Agent 2 - answer': 'transcripts/MAS/malicious/bigbench/qwen3b_job8462_2026-04-23_04h29m30s',
# #     'tasksource/bigbench: Mal Agent 3 - answer': 'transcripts/MAS/malicious/bigbench/qwen3b_job337_2026-04-01_01h00m42s',
# # }
# # result = malicious_comparison(transcript_data, round_figure=False, comparison_name='Qwen, BigBench, MAS, answer')
# # all_results.append(result)


# # --- Malicious prompt comparison for Olmo ---
# transcript_data = {
#     'tasksource/bigbench: No Mal Agent - both': 'transcripts/MAS/both/olmo7b_2026-03-29_01h47m46s',
#     'tasksource/bigbench: Mal Agent 1 - both': 'transcripts/MAS/malicious/bigbench/olmo7b_job9922_2026-03-31_15h39m15s',
#     'tasksource/bigbench: Mal Agent 2 - both': 'transcripts/MAS/malicious/bigbench/olmo7b_job8052_2026-04-22_20h02m33s',
#     'tasksource/bigbench: Mal Agent 3 - both': 'transcripts/MAS/malicious/bigbench/olmo7b_job347_2026-04-01_03h38m14s',
# }
# # result = malicious_comparison(transcript_data, round_figure=False, comparison_name='Olmo, BigBench, MAS, both')
# # all_results.append(result)
# result = malicious_comparison(transcript_data, round_figure=True, return_data=False, return_round_data=True, comparison_name='Olmo, BigBench, MAS, both')
# all_results.append(result)

# # transcript_data = {
# #     'tasksource/bigbench: No Mal Agent - answer': 'transcripts/MAS/answer/olmo7b_job126_2026-03-31_19h41m31s',
# #     'tasksource/bigbench: Mal Agent 1 - answer': 'transcripts/MAS/malicious/bigbench/olmo7b_job9930_2026-03-31_16h57m17s',
# #     'tasksource/bigbench: Mal Agent 2 - answer': 'transcripts/MAS/malicious/bigbench/olmo7b_job8463_2026-04-23_04h30m06s',
# #     'tasksource/bigbench: Mal Agent 3 - answer': 'transcripts/MAS/malicious/bigbench/olmo7b_job338_2026-04-01_01h02m15s',
# # }
# # result = malicious_comparison(transcript_data, round_figure=False, comparison_name='Olmo, BigBench, MAS, answer')
# # all_results.append(result)


# # --- Malicious prompt comparison for Olmo Think ---
# transcript_data = {
#     'tasksource/bigbench: No Mal Agent - both': 'transcripts/MAS/both/olmo7bThink_job3967_2026-04-25_15h20m06s',
#     'tasksource/bigbench: Mal Agent 1 - both': 'transcripts/MAS/malicious/bigbench/olmo7bThink_job5357_2026-04-13_15h40m32s',
#     'tasksource/bigbench: Mal Agent 2 - both': 'transcripts/MAS/malicious/bigbench/olmo7bThink_job7176_2026-04-27_14h41m04s',
#     'tasksource/bigbench: Mal Agent 3 - both': 'transcripts/MAS/malicious/bigbench/olmo7bThink_job7180_2026-04-27_18h24m20s',
# }
# # result = malicious_comparison(transcript_data, round_figure=False, comparison_name='Olmo Think, BigBench, MAS, both')
# # all_results.append(result)
# result = malicious_comparison(transcript_data, round_figure=True, return_data=False, return_round_data=True, comparison_name='Olmo Think, BigBench, MAS, both')
# all_results.append(result)



# # # --- Malicious prompt comparison for Llama ---
# transcript_data = {
#     'tasksource/bigbench: No Mal Agent - both': 'transcripts/MAS/both/llama3b_2026-03-29_03h08m47s',
#     'tasksource/bigbench: Mal Agent 1 - both': 'transcripts/MAS/malicious/bigbench/llama3b_job9920_2026-03-31_12h08m05s',
#     'tasksource/bigbench: Mal Agent 2 - both': 'transcripts/MAS/malicious/bigbench/llama3b_job8050_2026-04-22_15h53m20s',
#     'tasksource/bigbench: Mal Agent 3 - both': 'transcripts/MAS/malicious/bigbench/llama3b_job9924_2026-03-31_16h17m31s',
# }
# # result = malicious_comparison(transcript_data, round_figure=False, comparison_name='Llama, BigBench, MAS, both')
# # all_results.append(result)
# result = malicious_comparison(transcript_data, round_figure=True, return_data=False, return_round_data=True, comparison_name='Llama, BigBench, MAS, both')
# all_results.append(result)

# # transcript_data = {
# #     'tasksource/bigbench: No Mal Agent - answer': 'transcripts/MAS/answer/llama3b_2026-03-29_23h24m02s',
# #     'tasksource/bigbench: Mal Agent 1 - answer': 'transcripts/MAS/malicious/bigbench/llama3b_job9928_2026-03-31_16h55m27s',
# #     'tasksource/bigbench: Mal Agent 2 - answer': 'transcripts/MAS/malicious/bigbench/llama3b_job8461_2026-04-23_03h23m10s',
# #     'tasksource/bigbench: Mal Agent 3 - answer': 'transcripts/MAS/malicious/bigbench/llama3b_job336_2026-04-01_01h00m20s',
# # }
# # result = malicious_comparison(transcript_data, round_figure=False, comparison_name='Llama, BigBench, MAS, answer')
# # all_results.append(result)


# # --- Malicious prompt comparison for Qwen ---
# transcript_data = {
#     'openai/gsm8k: No Mal Agent - both': 'transcripts/MAS/both/qwen3b_2026-03-29_00h31m15s',
#     'openai/gsm8k: Mal Agent 1 - both': 'transcripts/MAS/malicious/gsm8k/qwen3b_job345_2026-04-01_01h35m32s',
#     'openai/gsm8k: Mal Agent 2 - both': 'transcripts/MAS/malicious/gsm8k/qwen3b_job7484_2026-04-21_17h30m01s',
#     'openai/gsm8k: Mal Agent 3 - both': 'transcripts/MAS/malicious/gsm8k/qwen3b_job342_2026-04-01_01h15m49s',
# }
# # result = malicious_comparison(transcript_data, round_figure=False, comparison_name='Qwen, GSM8K, MAS, both')
# # all_results.append(result)
# result = malicious_comparison(transcript_data, round_figure=True, return_data=False, return_round_data=True, comparison_name='Qwen, GSM8K, MAS, both')
# all_results.append(result)

# # transcript_data = {
# #     'openai/gsm8k: No Mal Agent - answer': 'transcripts/MAS/answer/qwen3b_job9705_2026-03-30_21h58m14s',
# #     'openai/gsm8k: Mal Agent 1 - answer': 'transcripts/MAS/malicious/gsm8k/qwen3b_job3963_2026-04-02_02h14m14s',
# #     'openai/gsm8k: Mal Agent 2 - answer': 'transcripts/MAS/malicious/gsm8k/qwen3b_job7490_2026-04-22_00h14m22s',
# #     'openai/gsm8k: Mal Agent 3 - answer': 'transcripts/MAS/malicious/gsm8k/qwen3b_job3960_2026-04-01_16h46m12s',
# # }
# # result = malicious_comparison(transcript_data, round_figure=False, comparison_name='Qwen, GSM8K, MAS, answer')
# # all_results.append(result)



# # --- Malicious prompt comparison for Olmo ---
# transcript_data = {
#     'openai/gsm8k: No Mal Agent - both': 'transcripts/MAS/both/olmo7b_2026-03-29_15h41m56s',
#     'openai/gsm8k: Mal Agent 1 - both': 'transcripts/MAS/malicious/gsm8k/olmo7b_job346_2026-04-01_02h18m18s',
#     'openai/gsm8k: Mal Agent 2 - both': 'transcripts/MAS/malicious/gsm8k/olmo7b_job7485_2026-04-21_18h06m31s',
#     'openai/gsm8k: Mal Agent 3 - both': 'transcripts/MAS/malicious/gsm8k/olmo7b_job343_2026-04-01_01h15m49s',
# }
# # result = malicious_comparison(transcript_data, round_figure=False, comparison_name='Olmo, GSM8K, MAS, both')
# # all_results.append(result)
# result = malicious_comparison(transcript_data, round_figure=True, return_data=False, return_round_data=True, comparison_name='Olmo, GSM8K, MAS, both')
# all_results.append(result)

# # transcript_data = {
# #     'openai/gsm8k: No Mal Agent - answer': 'transcripts/MAS/answer/olmo7b_job126_2026-03-31_19h41m31s',
# #     'openai/gsm8k: Mal Agent 1 - answer': 'transcripts/MAS/malicious/gsm8k/olmo7b_job3964_2026-04-02_04h05m45s',
# #     'openai/gsm8k: Mal Agent 2 - answer': 'transcripts/MAS/malicious/gsm8k/olmo7b_job7491_2026-04-22_00h14m19s',
# #     'openai/gsm8k: Mal Agent 3 - answer': 'transcripts/MAS/malicious/gsm8k/olmo7b_job3961_2026-04-01_23h51m40s',
# # }
# # result = malicious_comparison(transcript_data, round_figure=False, comparison_name='Olmo, GSM8K, MAS, answer')
# # all_results.append(result)

# # --- Malicious prompt comparison for Olmo Think ---
# transcript_data = {
#     'openai/gsm8k: No Mal Agent - both': 'transcripts/MAS/both/olmo7bThink_job3961_2026-04-25_15h14m16s',
#     'openai/gsm8k: Mal Agent 1 - both': 'transcripts/MAS/malicious/gsm8k/olmo7bThink_job5354_2026-04-13_15h40m36s', 
#     'openai/gsm8k: Mal Agent 2 - both': 'transcripts/MAS/malicious/gsm8k/olmo7bThink_job7173_2026-04-27_14h21m36s',
#     'openai/gsm8k: Mal Agent 3 - both': 'transcripts/MAS/malicious/gsm8k/olmo7bThink_job7177_2026-04-27_14h54m01s',

# }
# # result = malicious_comparison(transcript_data, round_figure=False, comparison_name='Olmo Think, GSM8K, MAS, both')
# # all_results.append(result)
# result = malicious_comparison(transcript_data, round_figure=True, return_data=False, return_round_data=True, comparison_name='Olmo Think, GSM8K, MAS, both')
# all_results.append(result)



# # --- Malicious prompt comparison for Llama ---
# transcript_data = {
#     'openai/gsm8k: No Mal Agent - both': 'transcripts/MAS/both/llama3b_2026-03-29_15h19m52s',
#     'openai/gsm8k: Mal Agent 1 - both': 'transcripts/MAS/malicious/gsm8k/llama3b_job344_2026-04-01_01h21m12s',
#     'openai/gsm8k: Mal Agent 2 - both': 'transcripts/MAS/malicious/gsm8k/llama3b_job7483_2026-04-21_17h12m09s',
#     'openai/gsm8k: Mal Agent 3 - both': 'transcripts/MAS/malicious/gsm8k/llama3b_job341_2026-04-01_01h15m43s',
# }
# # result = malicious_comparison(transcript_data, round_figure=False, comparison_name='Llama, GSM8K, MAS, both')
# # all_results.append(result)
# result = malicious_comparison(transcript_data, round_figure=True, return_data=False, return_round_data=True, comparison_name='Llama, GSM8K, MAS, both')
# all_results.append(result)

# # transcript_data = {
# #     'openai/gsm8k: No Mal Agent - answer': 'transcripts/MAS/answer/llama3b_job9704_2026-03-30_21h57m33s',
# #     'openai/gsm8k: Mal Agent 1 - answer': 'transcripts/MAS/malicious/gsm8k/llama3b_job3962_2026-04-02_01h48m51s',
# #     'openai/gsm8k: Mal Agent 2 - answer': 'transcripts/MAS/malicious/gsm8k/llama3b_job1628_2026-04-24_14h22m13s',
# #     'openai/gsm8k: Mal Agent 3 - answer': 'transcripts/MAS/malicious/gsm8k/llama3b_job3959_2026-04-01_15h52m13s',
# # }
# # result = malicious_comparison(transcript_data, round_figure=False, comparison_name='Llama, GSM8K, MAS, answer')
# # all_results.append(result)


# # --- Malicious prompt comparison for Qwen ---
# transcript_data = {
#     'cais/mmlu: No Mal Agent - both': 'transcripts/MAS/both/qwen3b_2026-03-29_15h34m56s',
#     'cais/mmlu: Mal Agent 1 - both': 'transcripts/MAS/malicious/mmlu/qwen3b_job4470_2026-04-02_08h10m34s',
#     'cais/mmlu: Mal Agent 2 - both': 'transcripts/MAS/malicious/mmlu/qwen3b_job7421_2026-04-21_15h44m18s',
#     'cais/mmlu: Mal Agent 3 - both': 'transcripts/MAS/malicious/mmlu/qwen3b_job4466_2026-04-02_04h47m13s',
# }
# # result = malicious_comparison(transcript_data, round_figure=False, comparison_name='Qwen, MMLU, MAS, both')
# # all_results.append(result)
# result = malicious_comparison(transcript_data, round_figure=True, return_data=False, return_round_data=True, comparison_name='Qwen, MMLU, MAS, both')
# all_results.append(result)

# # transcript_data = {
# #     'cais/mmlu: No Mal Agent - answer': 'transcripts/MAS/answer/qwen3b_job9914_2026-03-31_11h56m37s',
# #     'cais/mmlu: Mal Agent 1 - answer': 'transcripts/MAS/malicious/mmlu/qwen3b_job7400_2026-04-03_01h36m45s',
# #     'cais/mmlu: Mal Agent 2 - answer': 'transcripts/MAS/malicious/mmlu/qwen3b_job7476_2026-04-21_17h07m22s',
# #     'cais/mmlu: Mal Agent 3 - answer': 'transcripts/MAS/malicious/mmlu/qwen3b_job7403_2026-04-03_06h42m02s',
# # }
# # result = malicious_comparison(transcript_data, round_figure=False, comparison_name='Qwen, MMLU, MAS, answer')
# # all_results.append(result)


# # --- Malicious prompt comparison for Olmo ---
# transcript_data = {
#     'cais/mmlu: No Mal Agent - both': 'transcripts/MAS/both/olmo7b_2026-03-29_15h29m53s',
#     'cais/mmlu: Mal Agent 1 - both': 'transcripts/MAS/malicious/mmlu/olmo7b_job4471_2026-04-02_08h10m34s',
#     'cais/mmlu: Mal Agent 2 - both': 'transcripts/MAS/malicious/mmlu/olmo7b_job7423_2026-04-21_16h41m29s',
#     'cais/mmlu: Mal Agent 3 - both': 'transcripts/MAS/malicious/mmlu/olmo7b_job4467_2026-04-02_08h10m34s',
# }
# # result = malicious_comparison(transcript_data, round_figure=False, comparison_name='Olmo, MMLU, MAS, both')
# # all_results.append(result)
# result = malicious_comparison(transcript_data, round_figure=True, return_data=False, return_round_data=True, comparison_name='Qwen, MMLU, MAS, both')
# all_results.append(result)

# # transcript_data = {
# #     'cais/mmlu: No Mal Agent - answer': 'transcripts/MAS/answer/olmo7b_job124_2026-03-31_19h08m18s',
# #     'cais/mmlu: Mal Agent 1 - answer': 'transcripts/MAS/malicious/mmlu/olmo7b_job7401_2026-04-03_05h19m55s',
# #     'cais/mmlu: Mal Agent 2 - answer': 'transcripts/MAS/malicious/mmlu/olmo7b_job7477_2026-04-21_17h07m22s',
# #     'cais/mmlu: Mal Agent 3 - answer': 'transcripts/MAS/malicious/mmlu/olmo7b_job7404_2026-04-03_07h25m56s',
# # }
# # result = malicious_comparison(transcript_data, round_figure=False, comparison_name='Olmo, MMLU, MAS, answer')
# # all_results.append(result)


# # --- Malicious prompt comparison for Olmo Think ---
# transcript_data = {
#     'cais/mmlu: No Mal Agent - both': 'transcripts/MAS/both/olmo7bThink_job3963_2026-04-25_15h15m10s',
#     'cais/mmlu: Mal Agent 1 - both': 'transcripts/MAS/malicious/mmlu/olmo7bThink_job5355_2026-04-13_15h40m11s',
#     'cais/mmlu: Mal Agent 2 - both': 'transcripts/MAS/malicious/mmlu/olmo7bThink_job7174_2026-04-27_14h22m26s',
#     'cais/mmlu: Mal Agent 3 - both': 'transcripts/MAS/malicious/mmlu/olmo7bThink_job7178_2026-04-27_16h14m44s',
# }
# # result = malicious_comparison(transcript_data, round_figure=False, comparison_name='Olmo Think, MMLU, MAS, both')
# # all_results.append(result)
# result = malicious_comparison(transcript_data, round_figure=True, return_data=False, return_round_data=True, comparison_name='Olmo Think, MMLU, MAS, both')
# all_results.append(result)


# # --- Malicious prompt comparison for Llama ---
# transcript_data = {
#     'cais/mmlu: No Mal Agent - both': 'transcripts/MAS/both/llama3b_2026-03-29_00h34m49s',
#     'cais/mmlu: Mal Agent 1 - both': 'transcripts/MAS/malicious/mmlu/llama3b_job4469_2026-04-02_08h10m34s',
#     'cais/mmlu: Mal Agent 2 - both': 'transcripts/MAS/malicious/mmlu/llama3b_job7420_2026-04-21_15h24m18s',
#     'cais/mmlu: Mal Agent 3 - both': 'transcripts/MAS/malicious/mmlu/llama3b_job4465_2026-04-02_04h34m10s',
# }
# # result = malicious_comparison(transcript_data, round_figure=False, comparison_name='Llama, MMLU, MAS, both')
# # all_results.append(result)
# result = malicious_comparison(transcript_data, round_figure=True, return_data=False, return_round_data=True, comparison_name='Llama, MMLU, MAS, both')
# all_results.append(result)

# # # --- Malicious prompt comparison for Llama ---
# # transcript_data = {
# #     'cais/mmlu: No Mal Agent - answer': 'transcripts/MAS/answer/llama3b_job349_2026-04-01_04h00m41s',
# #     'cais/mmlu: Mal Agent 1 - answer': 'transcripts/MAS/malicious/mmlu/llama3b_job7399_2026-04-03_01h36m45s',
# #     'cais/mmlu: Mal Agent 2 - answer': 'transcripts/MAS/malicious/mmlu/llama3b_job7475_2026-04-21_17h07m22s',
# #     'cais/mmlu: Mal Agent 3 - answer': 'transcripts/MAS/malicious/mmlu/llama3b_job7402_2026-04-03_05h21m13s',
# # }
# # result = malicious_comparison(transcript_data, round_figure=False, comparison_name='Llama, MMLU, MAS, answer')
# # all_results.append(result)


# # --- Malicious prompt comparison for Qwen ---
# transcript_data = {
#     'ChilleD/StrategyQA: No Mal Agent - both': 'transcripts/MAS/both/qwen3b_2026-03-29_15h26m25s',
#     'ChilleD/StrategyQA: Mal Agent 1 - both': 'transcripts/MAS/malicious/strategyQA/qwen3b_job7394_2026-04-03_01h31m22s',
#     'ChilleD/StrategyQA: Mal Agent 2 - both': 'transcripts/MAS/malicious/strategyQA/qwen3b_job7487_2026-04-21_17h11m38s',
#     'ChilleD/StrategyQA: Mal Agent 3 - both': 'transcripts/MAS/malicious/strategyQA/qwen3b_job7384_2026-04-03_01h14m27s',
# }
# # result = malicious_comparison(transcript_data, round_figure=False, comparison_name='Qwen, StrategyQA, MAS, both')
# # all_results.append(result)
# result = malicious_comparison(transcript_data, round_figure=True, return_data=False, return_round_data=True, comparison_name='Qwen, StrategyQA, MAS, both')
# all_results.append(result)

# # transcript_data = {
# #     'ChilleD/StrategyQA: No Mal Agent - answer': 'transcripts/MAS/answer/qwen3b_job9918_2026-03-31_17h05m39s',
# #     'ChilleD/StrategyQA: Mal Agent 1 - answer': 'transcripts/MAS/malicious/strategyQA/qwen3b_job7387_2026-04-03_01h15m56s',
# #     'ChilleD/StrategyQA: Mal Agent 2 - answer': 'transcripts/MAS/malicious/strategyQA/qwen3b_job8048_2026-04-22_15h48m00s',
# #     'ChilleD/StrategyQA: Mal Agent 3 - answer': 'transcripts/MAS/malicious/strategyQA/qwen3b_job7904_2026-04-04_02h02m06s',
# # }
# # result = malicious_comparison(transcript_data, round_figure=False, comparison_name='Qwen, StrategyQA, MAS, answer')
# # all_results.append(result)


# # --- Malicious prompt comparison for Olmo ---
# transcript_data = {
#     'ChilleD/StrategyQA: No Mal Agent - both': 'transcripts/MAS/both/olmo7b_2026-03-29_15h26m25s',
#     'ChilleD/StrategyQA: Mal Agent 1 - both': 'transcripts/MAS/malicious/strategyQA/olmo7b_job7395_2026-04-03_01h30m38s',
#     'ChilleD/StrategyQA: Mal Agent 2 - both': 'transcripts/MAS/malicious/strategyQA/olmo7b_job1627_2026-04-24_13h31m57s',
#     'ChilleD/StrategyQA: Mal Agent 3 - both': 'transcripts/MAS/malicious/strategyQA/olmo7b_job7385_2026-04-03_01h15m54s',
# }
# # result = malicious_comparison(transcript_data, round_figure=False, comparison_name='Olmo, StrategyQA, MAS, both')
# # all_results.append(result)
# result = malicious_comparison(transcript_data, round_figure=True, return_data=False, return_round_data=True, comparison_name='Olmo, StrategyQA, MAS, both')
# all_results.append(result)

# # transcript_data = {
# #     'ChilleD/StrategyQA: No Mal Agent - answer': 'transcripts/MAS/answer/olmo7b_job9973_2026-03-31_18h08m36s',
# #     'ChilleD/StrategyQA: Mal Agent 1 - answer': 'transcripts/MAS/malicious/strategyQA/olmo7b_job7388_2026-04-03_01h15m54s',
# #     'ChilleD/StrategyQA: Mal Agent 2 - answer': 'transcripts/MAS/malicious/strategyQA/olmo7b_job8049_2026-04-22_15h53m17s',
# #     'ChilleD/StrategyQA: Mal Agent 3 - answer': 'transcripts/MAS/malicious/strategyQA/olmo7b_job7905_2026-04-04_02h02m06s',
# # }
# # result = malicious_comparison(transcript_data, round_figure=False, comparison_name='Olmo, StrategyQA, MAS, answer')
# # all_results.append(result)


# # --- Malicious prompt comparison for Olmo Think ---
# transcript_data = {
#     'ChilleD/StrategyQA: No Mal Agent - both': 'transcripts/MAS/both/olmo7bThink_job3965_2026-04-25_15h20m06s',
#     'ChilleD/StrategyQA: Mal Agent 1 - both': 'transcripts/MAS/malicious/strategyQA/olmo7bThink_job5356_2026-04-13_15h40m32s',
#     'ChilleD/StrategyQA: Mal Agent 2 - both': 'transcripts/MAS/malicious/strategyQA/olmo7bThink_job7175_2026-04-27_14h41m04s',
#     'ChilleD/StrategyQA: Mal Agent 3 - both': 'transcripts/MAS/malicious/strategyQA/olmo7bThink_job7179_2026-04-27_16h15m33s',
# }
# # result = malicious_comparison(transcript_data, round_figure=False, comparison_name='Olmo Think, StrategyQA, MAS, both')
# # all_results.append(result)
# result = malicious_comparison(transcript_data, round_figure=True, return_data=False, return_round_data=True, comparison_name='Olmo Think, StrategyQA, MAS, both')
# all_results.append(result)


# # --- Malicious prompt comparison for Llama ---
# transcript_data = {
#     'ChilleD/StrategyQA: No Mal Agent - both': 'transcripts/MAS/both/llama3b_2026-03-29_15h21m49s',
#     'ChilleD/StrategyQA: Mal Agent 1 - both': 'transcripts/MAS/malicious/strategyQA/llama3b_job7393_2026-04-03_01h31m22s',
#     'ChilleD/StrategyQA: Mal Agent 2 - both': 'transcripts/MAS/malicious/strategyQA/llama3b_job7486_2026-04-21_17h11m38s',
#     'ChilleD/StrategyQA: Mal Agent 3 - both': 'transcripts/MAS/malicious/strategyQA/llama3b_job7396_2026-04-03_01h30m38s',
# }
# # result = malicious_comparison(transcript_data, round_figure=False, comparison_name='Llama, StrategyQA, MAS, both')
# # all_results.append(result)
# result = malicious_comparison(transcript_data, round_figure=True, return_data=False, return_round_data=True, comparison_name='Llama, StrategyQA, MAS, both')
# all_results.append(result)

# # transcript_data = {
# #     'ChilleD/StrategyQA: No Mal Agent - answer': 'transcripts/MAS/answer/llama3b_job351_2026-04-01_07h39m19s',
# #     'ChilleD/StrategyQA: Mal Agent 1 - answer': 'transcripts/MAS/malicious/strategyQA/llama3b_job7386_2026-04-03_01h15m52s',
# #     'ChilleD/StrategyQA: Mal Agent 2 - answer': 'transcripts/MAS/malicious/strategyQA/llama3b_job8047_2026-04-22_15h48m00s',
# #     'ChilleD/StrategyQA: Mal Agent 3 - answer': 'transcripts/MAS/malicious/strategyQA/llama3b_job7903_2026-04-04_02h02m06s',
# # }
# # result = malicious_comparison(transcript_data, round_figure=False, comparison_name='Llama, StrategyQA, MAS, answer')
# # all_results.append(result)


# # ╔═════════════════════════════════════════════════╗
# # ║       2 MALICIOUS AGENTS PROMPT COMPARISONS     ║
# # ╚═════════════════════════════════════════════════╝

# # --- Malicious prompt comparison for Llama ---
# transcript_data = {
#     'tasksource/bigbench: No Mal Agent': 'transcripts/MAS/both/llama3b_2026-03-29_03h08m47s',
#     'tasksource/bigbench: Mal Agent 1 & 2': 'transcripts/MAS/malicious/2maliciousAgents/agents1_2/bigbench/llama3b_job7928_2026-04-04_06h09m05s',
#     'tasksource/bigbench: Mal Agent 1 & 3': 'transcripts/MAS/malicious/2maliciousAgents/agents1_3/bigbench/llama3b_job1759_2026-04-10_15h44m06s',
#     'tasksource/bigbench: Mal Agent 2 & 3': 'transcripts/MAS/malicious/2maliciousAgents/agents2_3/bigbench/llama3b_job1756_2026-04-10_15h29m06s',
#     # 'tasksource/bigbench: No Mal Agent - answer': 'transcripts/MAS/answer/llama3b_2026-03-29_23h24m02s',
#     # 'tasksource/bigbench: Mal Agent 1 & 2 - answer': 'transcripts/MAS/malicious/2maliciousAgents/agents1_2/bigbench/llama3b_job8617_2026-04-05_12h39m22s',
# }
# result =malicious_comparison(transcript_data, round_figure=False, comparison_name='Llama, BigBench, MAS, both')
# # all_results.append(result)
# result = malicious_comparison(transcript_data, round_figure=True, return_data=False, return_round_data=True, comparison_name='Llama, BigBench, MAS, both')
# all_results.append(result)


# --- Malicious prompt comparison for Olmo ---
# transcript_data = {
#     'tasksource/bigbench: No Mal Agent': 'transcripts/MAS/both/olmo7b_2026-03-29_01h47m46s',
#     'tasksource/bigbench: Mal Agent 1 & 2': 'transcripts/MAS/malicious/2maliciousAgents/agents1_2/bigbench/olmo7b_job7930_2026-04-04_11h15m30s',
#     'tasksource/bigbench: Mal Agent 1 & 3': 'transcripts/MAS/malicious/2maliciousAgents/agents1_3/bigbench/olmo7b_job1761_2026-04-10_15h46m50s',
#     'tasksource/bigbench: Mal Agent 2 & 3': 'transcripts/MAS/malicious/2maliciousAgents/agents2_3/bigbench/olmo7b_job1758_2026-04-10_15h44m06s',
#     # 'tasksource/bigbench: No Mal Agent - answer': 'transcripts/MAS/answer/olmo7b_job126_2026-03-31_19h41m31s',
#     # 'tasksource/bigbench: Mal Agent 1 & 2 - answer': 'transcripts/MAS/malicious/2maliciousAgents/agents1_2/bigbench/olmo7b_job8619_2026-04-05_12h39m23s',
# }
# result = malicious_comparison(transcript_data, round_figure=False, comparison_name='Olmo, BigBench, MAS, both')
# # all_results.append(result)
# result = malicious_comparison(transcript_data, round_figure=True, return_data=False, return_round_data=True, comparison_name='Olmo, BigBench, MAS, both')
# all_results.append(result)


# # --- Malicious prompt comparison for Olmo Think ---
# transcript_data = {
#     'tasksource/bigbench: No Mal Agent - both': 'transcripts/MAS/both/olmo7bThink_job3967_2026-04-25_15h20m06s',
#     'tasksource/bigbench: Mal Agent 1 & 2 - both': 'transcripts/MAS/malicious/2maliciousAgents/agents1_2/bigbench/olmo7bThink_job5362_2026-04-13_16h25m34s',
#     'tasksource/bigbench: Mal Agent 1 & 3 - both': 'transcripts/MAS/malicious/2maliciousAgents/agents1_3/bigbench/olmo7bThink_job3990_2026-05-01_18h04m53s',
#     'tasksource/bigbench: Mal Agent 2 & 3 - both': 'transcripts/MAS/malicious/2maliciousAgents/agents2_3/bigbench/olmo7bThink_job5451_2026-05-03_16h38m13s',
# }
# # result = malicious_comparison(transcript_data, round_figure=False, comparison_name='Olmo Think, BigBench, MAS, both')
# # all_results.append(result)
# result = malicious_comparison(transcript_data, round_figure=True, return_data=False, return_round_data=True, comparison_name='Olmo Think, BigBench, MAS, both')
# all_results.append(result)


# --- Malicious prompt comparison for Qwen ---
# transcript_data = {
#     'tasksource/bigbench: No Mal Agent': 'transcripts/MAS/both/qwen3b_2026-03-29_03h35m18s',
#     'tasksource/bigbench: Mal Agent 1 & 2': 'transcripts/MAS/malicious/2maliciousAgents/agents1_2/bigbench/qwen3b_job7929_2026-04-04_07h15m35s',
#     'tasksource/bigbench: Mal Agent 1 & 3': 'transcripts/MAS/malicious/2maliciousAgents/agents1_3/bigbench/qwen3b_job1760_2026-04-10_15h44m06s',
#     'tasksource/bigbench: Mal Agent 2 & 3': 'transcripts/MAS/malicious/2maliciousAgents/agents2_3/bigbench/qwen3b_job1757_2026-04-10_15h29m23s',
#     # 'tasksource/bigbench: No Mal Agent - answer': 'transcripts/MAS/answer/qwen3b_2026-03-29_21h46m31s',
#     # 'tasksource/bigbench: Mal Agent 1 & 2 - answer': 'transcripts/MAS/malicious/2maliciousAgents/agents1_2/bigbench/qwen3b_job8618_2026-04-05_12h39m16s',
# }
# result = malicious_comparison(transcript_data, round_figure=False, comparison_name='Qwen, BigBench, MAS, both')
# # all_results.append(result)
# result = malicious_comparison(transcript_data, round_figure=True, return_data=False, return_round_data=True, comparison_name='Qwen, BigBench, MAS, both')
# all_results.append(result)


# # --- Malicious prompt comparison for Llama ---
# transcript_data = {
#     'openai/gsm8k: No Mal Agent': 'transcripts/MAS/both/llama3b_2026-03-29_15h19m52s',
#     'openai/gsm8k: Mal Agent 1 & 2': 'transcripts/MAS/malicious/2maliciousAgents/agents1_2/gsm8k/llama3b_job2569_2026-04-11_14h37m10s',
#     'openai/gsm8k: Mal Agent 1 & 3': 'transcripts/MAS/malicious/2maliciousAgents/agents1_3/gsm8k/llama3b_job5414_2026-04-08_13h29m50s',
#     'openai/gsm8k: Mal Agent 2 & 3': 'transcripts/MAS/malicious/2maliciousAgents/agents2_3/gsm8k/llama3b_job5376_2026-04-08_13h27m42s',
#     # 'openai/gsm8k: No Mal Agent - answer': 'transcripts/MAS/answer/llama3b_job9704_2026-03-30_21h57m33s',
#     # 'openai/gsm8k: Mal Agent 1 & 2 - answer': 'transcripts/MAS/malicious/2maliciousAgents/agents1_2/gsm8k/llama3b_job8078_2026-04-04_12h52m39s',
# }
# result = malicious_comparison(transcript_data, round_figure=False, comparison_name='Llama, GSM8K, MAS, both')
# # all_results.append(result)
# result = malicious_comparison(transcript_data, round_figure=True, return_data=False, return_round_data=True, comparison_name='Llama, GSM8K, MAS, both')
# all_results.append(result)


# # --- Malicious prompt comparison for Olmo ---
# transcript_data = {
#     'openai/gsm8k: No Mal Agent': 'transcripts/MAS/both/olmo7b_2026-03-29_15h41m56s',
#     'openai/gsm8k: Mal Agent 1 & 2': 'transcripts/MAS/malicious/2maliciousAgents/agents1_2/gsm8k/olmo7b_job7914_2026-04-04_02h03m27s',
#     'openai/gsm8k: Mal Agent 1 & 3': 'transcripts/MAS/malicious/2maliciousAgents/agents1_3/gsm8k/olmo7b_job5418_2026-04-08_23h19m55s',
#     'openai/gsm8k: Mal Agent 2 & 3': 'transcripts/MAS/malicious/2maliciousAgents/agents2_3/gsm8k/olmo7b_job5378_2026-04-08_13h27m42s',
# #     'openai/gsm8k: No Mal Agent - answer': 'transcripts/MAS/answer/olmo7b_job126_2026-03-31_19h41m31s',
# #     'openai/gsm8k: Mal Agent 1 & 2 - answer': 'transcripts/MAS/malicious/2maliciousAgents/agents1_2/gsm8k/olmo7b_job8080_2026-04-04_12h53m27s',
# }
# result = malicious_comparison(transcript_data, round_figure=False, comparison_name='Olmo, GSM8K, MAS, both')
# # all_results.append(result)
# result = malicious_comparison(transcript_data, round_figure=True, return_data=False, return_round_data=True, comparison_name='Olmo, GSM8K, MAS, both')
# all_results.append(result)


# # --- Malicious prompt comparison for Olmo Think ---
# transcript_data = {
#     'openai/gsm8k: No Mal Agent - both': 'transcripts/MAS/both/olmo7bThink_job3961_2026-04-25_15h14m16s',
#     'openai/gsm8k: Mal Agent 1 & 2 - both': 'transcripts/MAS/malicious/2maliciousAgents/agents1_2/gsm8k/olmo7bThink_job5359_2026-04-13_15h41m03s',
#     'openai/gsm8k: Mal Agent 1 & 3 - both': 'transcripts/MAS/malicious/2maliciousAgents/agents1_3/gsm8k/olmo7bThink_job3984_2026-05-01_18h04m37s',
#     'openai/gsm8k: Mal Agent 2 & 3 - both': 'transcripts/MAS/malicious/2maliciousAgents/agents2_3/gsm8k/olmo7bThink_job3991_2026-05-01_19h22m20s'
# }
# # result = malicious_comparison(transcript_data, round_figure=False, comparison_name='Olmo Think, GSM8K, MAS, both')
# # all_results.append(result)
# result = malicious_comparison(transcript_data, round_figure=True, return_data=False, return_round_data=True, comparison_name='Olmo Think, GSM8K, MAS, both')
# all_results.append(result)


# # --- Malicious prompt comparison for Qwen ---
# transcript_data = {
#     'openai/gsm8k: No Mal Agent': 'transcripts/MAS/both/qwen3b_2026-03-29_00h31m15s',
#     'openai/gsm8k: Mal Agent 1 & 2': 'transcripts/MAS/malicious/2maliciousAgents/agents1_2/gsm8k/qwen3b_job7913_2026-04-04_02h04m02s',
#     'openai/gsm8k: Mal Agent 1 & 3': 'transcripts/MAS/malicious/2maliciousAgents/agents1_3/gsm8k/qwen3b_job5417_2026-04-08_18h19m29s',
#     'openai/gsm8k: Mal Agent 2 & 3': 'transcripts/MAS/malicious/2maliciousAgents/agents2_3/gsm8k/qwen3b_job5377_2026-04-08_13h27m42s',
#     # 'openai/gsm8k: No Mal Agent - answer': 'transcripts/MAS/answer/qwen3b_job9705_2026-03-30_21h58m14s',
#     # 'openai/gsm8k: Mal Agent 1 & 2 - answer': 'transcripts/MAS/malicious/2maliciousAgents/agents1_2/gsm8k/qwen3b_job8079_2026-04-04_12h53m27s',
# }
# result = malicious_comparison(transcript_data, round_figure=False, comparison_name='Qwen, GSM8K, MAS, both')
# # all_results.append(result)
# result = malicious_comparison(transcript_data, round_figure=True, return_data=False, return_round_data=True, comparison_name='Qwen, GSM8K, MAS, both')
# all_results.append(result)


# # --- Malicious prompt comparison for Llama ---
# transcript_data = {
#     'cais/mmlu: No Mal Agent': 'transcripts/MAS/both/llama3b_2026-03-29_00h34m49s',
#     'cais/mmlu: Mal Agent 1 & 2': 'transcripts/MAS/malicious/2maliciousAgents/agents1_2/mmlu/llama3b_job7915_2026-04-04_02h03m27s',
#     'cais/mmlu: Mal Agent 1 & 3': 'transcripts/MAS/malicious/2maliciousAgents/agents1_3/mmlu/llama3b_job5415_2026-04-08_13h29m51s',
#     'cais/mmlu: Mal Agent 2 & 3': 'transcripts/MAS/malicious/2maliciousAgents/agents2_3/mmlu/llama3b_job5383_2026-04-08_13h27m42s',
#     # 'cais/mmlu: No Mal Agent - answer': 'transcripts/MAS/answer/llama3b_job349_2026-04-01_04h00m41s',
#     # 'cais/mmlu: Mal Agent 1 & 2 - answer': 'transcripts/MAS/malicious/2maliciousAgents/agents1_2/mmlu/llama3b_job8624_2026-04-05_12h38m21s',
# }
# result = malicious_comparison(transcript_data, round_figure=False, comparison_name='Llama, MMLU, MAS, both')
# # all_results.append(result)
# result = malicious_comparison(transcript_data, round_figure=True, return_data=False, return_round_data=True, comparison_name='Llama, MMLU, MAS, both')
# all_results.append(result)


# --- Malicious prompt comparison for Olmo ---
# transcript_data = {
#     'cais/mmlu: No Mal Agent': 'transcripts/MAS/both/olmo7b_2026-03-29_15h29m53s',
#     'cais/mmlu: Mal Agent 1 & 2': 'transcripts/MAS/malicious/2maliciousAgents/agents1_2/mmlu/olmo7b_job7917_2026-04-04_02h03m27s',
#     'cais/mmlu: Mal Agent 1 & 3': 'transcripts/MAS/malicious/2maliciousAgents/agents1_3/mmlu/olmo7b_job5419_2026-04-09_06h17m09s',
#     'cais/mmlu: Mal Agent 2 & 3': 'transcripts/MAS/malicious/2maliciousAgents/agents2_3/mmlu/olmo7b_job5385_2026-04-08_13h27m42s',
#     # 'cais/mmlu: No Mal Agent - answer': 'transcripts/MAS/answer/olmo7b_job124_2026-03-31_19h08m18s',
#     # 'cais/mmlu: Mal Agent 1 & 2 - answer': 'transcripts/MAS/malicious/2maliciousAgents/agents1_2/mmlu/olmo7b_job8626_2026-04-05_12h38m21s',
# }
# result = malicious_comparison(transcript_data, round_figure=False, comparison_name='Olmo, MMLU, MAS, both')
# # all_results.append(result)
# result = malicious_comparison(transcript_data, round_figure=True, return_data=False, return_round_data=True, comparison_name='Olmo, MMLU, MAS, both')
# all_results.append(result)


# # --- Malicious prompt comparison for Olmo Think ---
# transcript_data = {
#     'cais/mmlu: No Mal Agent - both': 'transcripts/MAS/both/olmo7bThink_job3963_2026-04-25_15h15m10s',
#     'cais/mmlu: Mal Agent 1 & 2 - both': 'transcripts/MAS/malicious/2maliciousAgents/agents1_2/mmlu/olmo7bThink_job5360_2026-04-13_16h01m29s',
#     'cais/mmlu: Mal Agent 1 & 3 - both': 'transcripts/MAS/malicious/2maliciousAgents/agents1_3/mmlu/olmo7bThink_job3986_2026-05-01_18h04m37s',
#     'cais/mmlu: Mal Agent 2 & 3 - both': 'transcripts/MAS/malicious/2maliciousAgents/agents2_3/mmlu/olmo7bThink_job3994_2026-05-01_20h40m31s',
# }
# # result = malicious_comparison(transcript_data, round_figure=False, comparison_name='Olmo Think, MMLU, MAS, both')
# # all_results.append(result)
# result = malicious_comparison(transcript_data, round_figure=True, return_data=False, return_round_data=True, comparison_name='Olmo Think, MMLU, MAS, both')
# all_results.append(result)


# --- Malicious prompt comparison for Qwen ---
# transcript_data = {
#     'cais/mmlu: No Mal Agent': 'transcripts/MAS/both/qwen3b_2026-03-29_15h34m56s',
#     'cais/mmlu: Mal Agent 1 & 2': 'transcripts/MAS/malicious/2maliciousAgents/agents1_2/mmlu/qwen3b_job7916_2026-04-04_02h03m27s',
#     'cais/mmlu: Mal Agent 1 & 3': 'transcripts/MAS/malicious/2maliciousAgents/agents1_3/mmlu/qwen3b_job5416_2026-04-08_13h29m51s',
#     'cais/mmlu: Mal Agent 2 & 3': 'transcripts/MAS/malicious/2maliciousAgents/agents2_3/mmlu/qwen3b_job5384_2026-04-08_13h27m42s',
#     # 'cais/mmlu: No Mal Agent - answer': 'transcripts/MAS/answer/qwen3b_job9914_2026-03-31_11h56m37s',
#     # 'cais/mmlu: Mal Agent 1 & 2 - answer': 'transcripts/MAS/malicious/2maliciousAgents/agents1_2/mmlu/qwen3b_job8625_2026-04-05_12h38m21s',
# }
# result = malicious_comparison(transcript_data, round_figure=False, comparison_name='Qwen, MMLU, MAS, both')
# # all_results.append(result)
# result = malicious_comparison(transcript_data, round_figure=True, return_data=False, return_round_data=True, comparison_name='Qwen, MMLU, MAS, both')
# all_results.append(result)


# # --- Malicious prompt comparison for Llama ---
# transcript_data = {
#     'ChilleD/StrategyQA: No Mal Agent': 'transcripts/MAS/both/llama3b_2026-03-29_15h21m49s',
#     'ChilleD/StrategyQA: Mal Agent 1 & 2': 'transcripts/MAS/malicious/2maliciousAgents/agents1_2/strategyQA/llama3b_job7919_2026-04-04_02h08m17s',
#     'ChilleD/StrategyQA: Mal Agent 1 & 3': 'transcripts/MAS/malicious/2maliciousAgents/agents1_3/strategyQA/llama3b_job5420_2026-04-09_06h36m40s',
#     'ChilleD/StrategyQA: Mal Agent 2 & 3': 'transcripts/MAS/malicious/2maliciousAgents/agents2_3/strategyQA/llama3b_job5386_2026-04-08_13h27m42s',
#     # 'ChilleD/StrategyQA: No Mal Agent - answer': 'transcripts/MAS/answer/llama3b_job351_2026-04-01_07h39m19s',
#     # 'ChilleD/StrategyQA: Mal Agent 1 & 2 - answer': 'transcripts/MAS/malicious/2maliciousAgents/agents1_2/strategyQA/llama3b_job8620_2026-04-05_12h39m16s',
# }
# result = malicious_comparison(transcript_data, round_figure=False, comparison_name='Llama, StrategyQA, MAS, both')
# # all_results.append(result)
# result = malicious_comparison(transcript_data, round_figure=True, return_data=False, return_round_data=True, comparison_name='Llama, StrategyQA, MAS, both')
# all_results.append(result)


# # --- Malicious prompt comparison for Olmo ---
# transcript_data = {
#     'ChilleD/StrategyQA: No Mal Agent': 'transcripts/MAS/both/olmo7b_2026-03-29_15h26m25s',
#     'ChilleD/StrategyQA: Mal Agent 1 & 2': 'transcripts/MAS/malicious/2maliciousAgents/agents1_2/strategyQA/olmo7b_job7921_2026-04-04_02h09m58s',
#     'ChilleD/StrategyQA: Mal Agent 1 & 3': 'transcripts/MAS/malicious/2maliciousAgents/agents1_3/strategyQA/olmo7b_job5422_2026-04-09_09h45m17s',
#     'ChilleD/StrategyQA: Mal Agent 2 & 3': 'transcripts/MAS/malicious/2maliciousAgents/agents2_3/strategyQA/olmo7b_job5388_2026-04-09_15h24m11s',
#     # 'ChilleD/StrategyQA: No Mal Agent - answer': 'transcripts/MAS/answer/olmo7b_job9973_2026-03-31_18h08m36s',
#     # 'ChilleD/StrategyQA: Mal Agent 1 & 2 - answer': 'transcripts/MAS/malicious/2maliciousAgents/agents1_2/strategyQA/olmo7b_job8622_2026-04-05_12h38m00s',
# }
# result = malicious_comparison(transcript_data, round_figure=False, comparison_name='Olmo, StrategyQA, MAS, both')
# # all_results.append(result)
# result = malicious_comparison(transcript_data, round_figure=True, return_data=False, return_round_data=True, comparison_name='Olmo, StrategyQA, MAS, both')
# all_results.append(result)


# # --- Malicious prompt comparison for Olmo Think ---
# transcript_data = {
#     'ChilleD/StrategyQA: No Mal Agent - both': 'transcripts/MAS/both/olmo7bThink_job3965_2026-04-25_15h20m06s',
#     'ChilleD/StrategyQA: Mal Agent 1 & 2 - both': 'transcripts/MAS/malicious/2maliciousAgents/agents1_2/strategyQA/olmo7bThink_job5361_2026-04-13_16h18m06s',
#     'ChilleD/StrategyQA: Mal Agent 1 & 3 - both': 'transcripts/MAS/malicious/2maliciousAgents/agents1_3/strategyQA/olmo7bThink_job4597_2026-05-02_23h55m31s',
#     'ChilleD/StrategyQA: Mal Agent 2 & 3 - both': 'transcripts/MAS/malicious/2maliciousAgents/agents2_3/strategyQA/olmo7bThink_job5449_2026-05-03_16h01m04s',
# }
# # result = malicious_comparison(transcript_data, round_figure=False, comparison_name='Olmo Think, StrategyQA, MAS, both')
# # all_results.append(result)
# result = malicious_comparison(transcript_data, round_figure=True, return_data=False, return_round_data=True, comparison_name='Olmo Think, StrategyQA, MAS, both')
# all_results.append(result)


# # --- Malicious prompt comparison for Qwen ---
# transcript_data = {
#     'ChilleD/StrategyQA: No Mal Agent': 'transcripts/MAS/both/qwen3b_2026-03-29_15h26m25s',
#     'ChilleD/StrategyQA: Mal Agent 1 & 2': 'transcripts/MAS/malicious/2maliciousAgents/agents1_2/strategyQA/qwen3b_job7920_2026-04-04_02h08m17s',
#     'ChilleD/StrategyQA: Mal Agent 1 & 3': 'transcripts/MAS/malicious/2maliciousAgents/agents1_3/strategyQA/qwen3b_job5421_2026-04-09_08h09m11s',
#     'ChilleD/StrategyQA: Mal Agent 2 & 3': 'transcripts/MAS/malicious/2maliciousAgents/agents2_3/strategyQA/qwen3b_job5387_2026-04-09_14h42m27s',
#     # 'ChilleD/StrategyQA: No Mal Agent - answer': 'transcripts/MAS/answer/qwen3b_job9918_2026-03-31_17h05m39s',
#     # 'ChilleD/StrategyQA: Mal Agent 1 & 2 - answer': 'transcripts/MAS/malicious/2maliciousAgents/agents1_2/strategyQA/qwen3b_job8621_2026-04-05_12h38m00s',
# }
# result = malicious_comparison(transcript_data, round_figure=False, comparison_name='Qwen, StrategyQA, MAS, both')
# # all_results.append(result)
# result = malicious_comparison(transcript_data, round_figure=True, return_data=False, return_round_data=True, comparison_name='Qwen, StrategyQA, MAS, both')
# all_results.append(result)


# export_round_comparisons_to_excel(all_results, 'round_results_mal.xlsx')

# ╔═════════════════════════════════════════════════╗
# ║       EARLY STOPPING ONE AGENT COMPARISON       ║
# ╚═════════════════════════════════════════════════╝

# # --- Early Stopping comparison for Qwen ---
# transcript_data = {
#     'openai/gsm8k: No Mal Agent': 'transcripts/MAS/both/qwen3b_2026-03-29_00h31m15s',
#     'openai/gsm8k: Mal Agent 1': 'transcripts/MAS/early_stopping/1ESagent/gsm8k/agent0/qwen3b_job3201_2026-04-18_16h04m28s',
#     'openai/gsm8k: Mal Agent 2': 'transcripts/MAS/early_stopping/1ESagent/gsm8k/agent1/qwen3b_job4764_2026-05-12_12h18m32s',
#     'openai/gsm8k: Mal Agent 3': 'transcripts/MAS/early_stopping/1ESagent/gsm8k/agent2/qwen3b_job6017_2026-05-14_07h57m11s',
# }
# # result = malicious_comparison(transcript_data, round_figure=False, comparison_type='early_stopping', return_data=True, comparison_name='Qwen, GSM8K, MAS, both')
# # all_results.append(result)
# result = malicious_comparison(transcript_data, round_figure=True, comparison_type='early_stopping',return_data=False, return_round_data=True, comparison_name='Qwen, GSM8K, MAS, both')
# all_results.append(result)


# transcript_data = {
#     'cais/mmlu: No Mal Agent': 'transcripts/MAS/both/qwen3b_2026-03-29_15h34m56s',
#     'cais/mmlu: Mal Agent 1': 'transcripts/MAS/early_stopping/1ESagent/mmlu/agent0/qwen3b_job3207_2026-04-18_20h30m03s',
#     'cais/mmlu: Mal Agent 2': 'transcripts/MAS/early_stopping/1ESagent/mmlu/agent1/qwen3b_job4767_2026-05-12_20h04m57s',
#     'cais/mmlu: Mal Agent 3': 'transcripts/MAS/early_stopping/1ESagent/mmlu/agent2/qwen3b_job275_2026-05-15_03h52m22s',
# }
# # result = malicious_comparison(transcript_data, round_figure=False, comparison_type='early_stopping', return_data=True, comparison_name='Qwen, MMLU, MAS, both')
# # all_results.append(result)
# result = malicious_comparison(transcript_data, round_figure=True, comparison_type='early_stopping',return_data=False, return_round_data=True, comparison_name='Qwen, MMLU, MAS, both')
# all_results.append(result)

# transcript_data = {
#     'ChilleD/StrategyQA: No Mal Agent': 'transcripts/MAS/both/qwen3b_2026-03-29_15h26m25s',
#     'ChilleD/StrategyQA: Mal Agent 1': 'transcripts/MAS/early_stopping/1ESagent/strategyQA/agent0/qwen3b_job3868_2026-04-19_14h26m08s',
#     'ChilleD/StrategyQA: Mal Agent 2': 'transcripts/MAS/early_stopping/1ESagent/strategyQA/agent1/qwen3b_job4770_2026-05-12_20h04m57s',
#     'ChilleD/StrategyQA: Mal Agent 3': 'transcripts/MAS/early_stopping/1ESagent/strategyQA/agent2/qwen3b_job272_2026-05-15_02h06m16s',
# }
# # result = malicious_comparison(transcript_data, round_figure=False, comparison_type='early_stopping', return_data=True, comparison_name='Qwen, StrategyQA, MAS, both')
# # all_results.append(result)
# result = malicious_comparison(transcript_data, round_figure=True, comparison_type='early_stopping',return_data=False, return_round_data=True, comparison_name='Qwen, StrategyQA, MAS, both')
# all_results.append(result)

# transcript_data = {
#     'tasksource/bigbench: No Mal Agent': 'transcripts/MAS/both/qwen3b_2026-03-29_03h35m18s',
#     'tasksource/bigbench: Mal Agent 1': 'transcripts/MAS/early_stopping/1ESagent/bigbench/agent0/qwen3b_job3869_2026-04-19_14h38m03s',
#     'tasksource/bigbench: Mal Agent 2': 'transcripts/MAS/early_stopping/1ESagent/bigbench/agent1/qwen3b_job6014_2026-05-14_05h12m35s',
#     'tasksource/bigbench: Mal Agent 3': 'transcripts/MAS/early_stopping/1ESagent/bigbench/agent2/qwen3b_job4469_2026-05-15_17h35m55s'
# }
# # result = malicious_comparison(transcript_data, round_figure=False, comparison_type='early_stopping', return_data=True, comparison_name='Qwen, BigBench, MAS, both')
# # all_results.append(result)
# result = malicious_comparison(transcript_data, round_figure=True, comparison_type='early_stopping',return_data=False, return_round_data=True, comparison_name='Qwen, BigBench, MAS, both')
# all_results.append(result)


# # --- Early Stopping comparison for Olmo ---
# transcript_data = {
#     'openai/gsm8k: No Mal Agent': 'transcripts/MAS/both/olmo7b_2026-03-29_15h41m56s',
#     'openai/gsm8k: Mal Agent 1': 'transcripts/MAS/early_stopping/1ESagent/gsm8k/agent0/olmo7b_job3202_2026-04-18_16h04m58s',
#     'openai/gsm8k: Mal Agent 2': 'transcripts/MAS/early_stopping/1ESagent/gsm8k/agent1/olmo7b_job4765_2026-05-12_12h21m16s',
#     'openai/gsm8k: Mal Agent 3': 'transcripts/MAS/early_stopping/1ESagent/gsm8k/agent2/olmo7b_job6018_2026-05-14_08h35m46s'
# }
# # result = malicious_comparison(transcript_data, round_figure=False, comparison_type='early_stopping', return_data=True, comparison_name='Olmo, GSM8K, MAS, both')
# # all_results.append(result)
# result = malicious_comparison(transcript_data, round_figure=True, comparison_type='early_stopping',return_data=False, return_round_data=True, comparison_name='Olmo, GSM8K, MAS, both')
# all_results.append(result)

# transcript_data = {
#     'cais/mmlu: No Mal Agent': 'transcripts/MAS/both/olmo7b_2026-03-29_15h29m53s',
#     'cais/mmlu: Mal Agent 1': 'transcripts/MAS/early_stopping/1ESagent/mmlu/agent0/olmo7b_job3208_2026-04-18_20h30m03s',
#     'cais/mmlu: Mal Agent 2': 'transcripts/MAS/early_stopping/1ESagent/mmlu/agent1/olmo7b_job4768_2026-05-12_17h03m50s',
#     'cais/mmlu: Mal Agent 3': 'transcripts/MAS/early_stopping/1ESagent/mmlu/agent2/olmo7b_job276_2026-05-15_04h36m22s'
# }
# # result = malicious_comparison(transcript_data, round_figure=False, comparison_type='early_stopping', return_data=True, comparison_name='Olmo, MMLU, MAS, both')
# # all_results.append(result)
# result = malicious_comparison(transcript_data, round_figure=True, comparison_type='early_stopping',return_data=False, return_round_data=True, comparison_name='Olmo, MMLU, MAS, both')
# all_results.append(result)

# transcript_data = {
#     'ChilleD/StrategyQA: No Mal Agent': 'transcripts/MAS/both/olmo7b_2026-03-29_15h26m25s',
#     'ChilleD/StrategyQA: Mal Agent 1': 'transcripts/MAS/early_stopping/1ESagent/strategyQA/agent0/olmo7b_job3870_2026-04-19_16h04m28s',
#     'ChilleD/StrategyQA: Mal Agent 2': 'transcripts/MAS/early_stopping/1ESagent/strategyQA/agent1/olmo7b_job4771_2026-05-12_20h04m57s',
#     'ChilleD/StrategyQA: Mal Agent 3': 'transcripts/MAS/early_stopping/1ESagent/strategyQA/agent2/olmo7b_job273_2026-05-15_03h01m56s'
# }
# # result = malicious_comparison(transcript_data, round_figure=False, comparison_type='early_stopping', return_data=True, comparison_name='Olmo, StrategyQA, MAS, both')
# # all_results.append(result)
# result = malicious_comparison(transcript_data, round_figure=True, comparison_type='early_stopping',return_data=False, return_round_data=True, comparison_name='Olmo, StrategyQA, MAS, both')
# all_results.append(result)

# transcript_data = {
#     'tasksource/bigbench: No Mal Agent': 'transcripts/MAS/both/olmo7b_2026-03-29_01h47m46s',
#     'tasksource/bigbench: Mal Agent 1': 'transcripts/MAS/early_stopping/1ESagent/bigbench/agent0/olmo7b_job3871_2026-04-19_19h09m27s',
#     'tasksource/bigbench: Mal Agent 2': 'transcripts/MAS/early_stopping/1ESagent/bigbench/agent1/olmo7b_job6015_2026-05-14_06h57m05s',
#     'tasksource/bigbench: Mal Agent 3': 'transcripts/MAS/early_stopping/1ESagent/bigbench/agent2/olmo7b_job4470_2026-05-15_19h04m13s'
# }
# # result = malicious_comparison(transcript_data, round_figure=False, comparison_type='early_stopping', return_data=True, comparison_name='Olmo, BigBench, MAS, both')
# # all_results.append(result)
# result = malicious_comparison(transcript_data, round_figure=True, comparison_type='early_stopping',return_data=False, return_round_data=True, comparison_name='Olmo, BigBench, MAS, both')
# all_results.append(result)


# # --- Early Stopping comparison for Olmo Think ---
# transcript_data = {
#     'openai/gsm8k: No Mal Agent': 'transcripts/MAS/both/olmo7bThink_job3961_2026-04-25_15h14m16s',
#     'openai/gsm8k: Mal Agent 1': 'transcripts/MAS/early_stopping/1ESagent/gsm8k/agent0/olmo7bThink_job5414_2026-04-13_16h35m24s',
#     # 'openai/gsm8k: Mal Agent 2': 'transcripts/MAS/early_stopping/1ESagent/gsm8k/agent1/...',
#     # 'openai/gsm8k: Mal Agent 3': 'transcripts/MAS/early_stopping/1ESagent/gsm8k/agent2/...'
# }
# result = malicious_comparison(transcript_data, round_figure=False, comparison_type='early_stopping', return_data=True, comparison_name='Olmo Think, GSM8K, MAS, both')
# all_results.append(result)

# transcript_data = {
#     'cais/mmlu: No Mal Agent': 'transcripts/MAS/both/olmo7bThink_job3963_2026-04-25_15h15m10s',
#     'cais/mmlu: Mal Agent 1': 'transcripts/MAS/early_stopping/1ESagent/mmlu/agent0/olmo7bThink_job5415_2026-04-13_17h00m44s',
#     # 'cais/mmlu: Mal Agent 2': 'transcripts/MAS/early_stopping/1ESagent/mmlu/agent1/...',
#     # 'cais/mmlu: Mal Agent 3': 'transcripts/MAS/early_stopping/1ESagent/mmlu/agent2/...'
# }
# result = malicious_comparison(transcript_data, round_figure=False, comparison_type='early_stopping', return_data=True, comparison_name='Olmo Think, MMLU, MAS, both')
# all_results.append(result)

# transcript_data = {
#     'ChilleD/StrategyQA: No Mal Agent': 'transcripts/MAS/both/olmo7bThink_job3965_2026-04-25_15h20m06s',
#     'ChilleD/StrategyQA: Mal Agent 1': 'transcripts/MAS/early_stopping/1ESagent/strategyQA/agent0/olmo7bThink_job5416_2026-04-13_17h00m44s',
#     # 'ChilleD/StrategyQA: Mal Agent 2': 'transcripts/MAS/early_stopping/1ESagent/strategyQA/agent1/...',
#     # 'ChilleD/StrategyQA: Mal Agent 3': 'transcripts/MAS/early_stopping/1ESagent/strategyQA/agent2/...'
# }
# result = malicious_comparison(transcript_data, round_figure=False, comparison_type='early_stopping', return_data=True, comparison_name='Olmo Think, StrategyQA, MAS, both')
# all_results.append(result)

# transcript_data = {
#     'tasksource/bigbench: No Mal Agent': 'transcripts/MAS/both/olmo7bThink_job3967_2026-04-25_15h20m06s',
#     'tasksource/bigbench: Mal Agent 1': 'transcripts/MAS/early_stopping/1ESagent/bigbench/agent0/olmo7bThink_job5417_2026-04-13_17h00m44s',
#     # 'tasksource/bigbench: Mal Agent 2': 'transcripts/MAS/early_stopping/1ESagent/bigbench/agent1/...',
#     # 'tasksource/bigbench: Mal Agent 3': 'transcripts/MAS/early_stopping/1ESagent/bigbench/agent2/...'
# }
# result = malicious_comparison(transcript_data, round_figure=False, comparison_type='early_stopping', return_data=True, comparison_name='Olmo Think, BigBench, MAS, both')
# all_results.append(result)


# # --- Early Stopping comparison for Llama ---
# transcript_data = {
#     'openai/gsm8k: No Mal Agent': 'transcripts/MAS/both/llama3b_2026-03-29_15h19m52s',
#     'openai/gsm8k: Mal Agent 1': 'transcripts/MAS/early_stopping/1ESagent/gsm8k/agent0/llama3b_job3200_2026-04-18_16h02m29s',
#     'openai/gsm8k: Mal Agent 2': 'transcripts/MAS/early_stopping/1ESagent/gsm8k/agent1/llama3b_job4763_2026-05-12_20h04m57s',
#     'openai/gsm8k: Mal Agent 3': 'transcripts/MAS/early_stopping/1ESagent/gsm8k/agent2/llama3b_job6016_2026-05-14_07h57m11s'
# }
# # result = malicious_comparison(transcript_data, round_figure=False, comparison_type='early_stopping', return_data=True, comparison_name='Llama, GSM8K, MAS, both')
# # all_results.append(result)
# result = malicious_comparison(transcript_data, round_figure=True, comparison_type='early_stopping',return_data=False, return_round_data=True, comparison_name='Llama, GSM8K, MAS, both')
# all_results.append(result)

# transcript_data = {
#     'cais/mmlu: No Mal Agent': 'transcripts/MAS/both/llama3b_2026-03-29_00h34m49s',
#     'cais/mmlu: Mal Agent 1': 'transcripts/MAS/early_stopping/1ESagent/mmlu/agent0/llama3b_job3206_2026-04-18_16h09m00s',
#     'cais/mmlu: Mal Agent 2': 'transcripts/MAS/early_stopping/1ESagent/mmlu/agent1/llama3b_job4766_2026-05-12_20h04m57s',
#     'cais/mmlu: Mal Agent 3': 'transcripts/MAS/early_stopping/1ESagent/mmlu/agent2/llama3b_job274_2026-05-15_03h02m27s'
# }
# # result = malicious_comparison(transcript_data, round_figure=False, comparison_type='early_stopping', return_data=True, comparison_name='Llama, MMLU, MAS, both')
# # all_results.append(result)
# result = malicious_comparison(transcript_data, round_figure=True, comparison_type='early_stopping',return_data=False, return_round_data=True, comparison_name='Llama, MMLU, MAS, both')
# all_results.append(result)

# transcript_data = {
#     'ChilleD/StrategyQA: No Mal Agent': 'transcripts/MAS/both/llama3b_2026-03-29_15h21m49s',
#     'ChilleD/StrategyQA: Mal Agent 1': 'transcripts/MAS/early_stopping/1ESagent/strategyQA/agent0/llama3b_job3866_2026-04-19_12h45m46s',
#     'ChilleD/StrategyQA: Mal Agent 2': 'transcripts/MAS/early_stopping/1ESagent/strategyQA/agent1/llama3b_job4769_2026-05-12_17h59m34s',
#     'ChilleD/StrategyQA: Mal Agent 3': 'transcripts/MAS/early_stopping/1ESagent/strategyQA/agent2/llama3b_job271_2026-05-15_01h02m58s'
# }
# # result = malicious_comparison(transcript_data, round_figure=False, comparison_type='early_stopping', return_data=True, comparison_name='Llama, StrategyQA, MAS, both')
# # all_results.append(result)
# result = malicious_comparison(transcript_data, round_figure=True, comparison_type='early_stopping',return_data=False, return_round_data=True, comparison_name='Llama, StrategyQA, MAS, both')
# all_results.append(result)

# transcript_data = {
#     'tasksource/bigbench: No Mal Agent': 'transcripts/MAS/both/llama3b_2026-03-29_03h08m47s',
#     'tasksource/bigbench: Mal Agent 1': 'transcripts/MAS/early_stopping/1ESagent/bigbench/agent0/llama3b_job3867_2026-04-19_12h45m46s',
#     'tasksource/bigbench: Mal Agent 2': 'transcripts/MAS/early_stopping/1ESagent/bigbench/agent1/llama3b_job6013_2026-05-14_03h57m51s',
#     'tasksource/bigbench: Mal Agent 3': 'transcripts/MAS/early_stopping/1ESagent/bigbench/agent2/llama3b_job4468_2026-05-15_17h20m36s'
# }
# # result = malicious_comparison(transcript_data, round_figure=False, comparison_type='early_stopping', return_data=True, comparison_name='Qwen, BigBench, MAS, both')
# # all_results.append(result)
# result = malicious_comparison(transcript_data, round_figure=True, comparison_type='early_stopping',return_data=False, return_round_data=True, comparison_name='Llama, BigBench, MAS, both')
# all_results.append(result)


# # ╔═════════════════════════════════════════════════╗
# # ║       EARLY STOPPING TWO AGENT COMPARISON       ║
# # ╚═════════════════════════════════════════════════╝

# # --- Early Stopping comparison for Qwen ---
# transcript_data = {
#     'openai/gsm8k: No Mal Agent': 'transcripts/MAS/both/qwen3b_2026-03-29_00h31m15s',
#     'openai/gsm8k: Mal Agent 1 & 2': 'transcripts/MAS/early_stopping/2ESagents/gsm8k/agents0_1/qwen3b_job5477_2026-04-20_11h06m32s',
#     'openai/gsm8k: Mal Agent 1 & 3': 'transcripts/MAS/early_stopping/2ESagents/gsm8k/agents0_2/qwen3b_job5444_2026-05-13_02h50m15s',
#     'openai/gsm8k: Mal Agent 2 & 3': 'transcripts/MAS/early_stopping/2ESagents/gsm8k/agents1_2/qwen3b_job7949_2026-05-14_19h15m40s',
# }
# # result = malicious_comparison(transcript_data, round_figure=False, comparison_type='early_stopping', return_data=True, comparison_name='Qwen, GSM8K, MAS, both')
# # all_results.append(result)
# result = malicious_comparison(transcript_data, round_figure=True, comparison_type='early_stopping',return_data=False, return_round_data=True, comparison_name='Qwen, GSM8K, MAS, both')
# all_results.append(result)

# transcript_data = {
#     'cais/mmlu: No Mal Agent': 'transcripts/MAS/both/qwen3b_2026-03-29_15h34m56s',
#     'cais/mmlu: Mal Agent 1 & 2': 'transcripts/MAS/early_stopping/2ESagents/mmlu/agents0_1/qwen3b_job5480_2026-04-20_11h06m58s',
#     'cais/mmlu: Mal Agent 1 & 3': 'transcripts/MAS/early_stopping/2ESagents/mmlu/agents0_2/qwen3b_job5447_2026-05-13_09h36m43s',
#     'cais/mmlu: Mal Agent 2 & 3': 'transcripts/MAS/early_stopping/2ESagents/mmlu/agents1_2/qwen3b_job7840_2026-05-14_17h38m05s',
# }
# # result = malicious_comparison(transcript_data, round_figure=False, comparison_type='early_stopping', return_data=True, comparison_name='Qwen, MMLU, MAS, both')
# # all_results.append(result)
# result = malicious_comparison(transcript_data, round_figure=True, comparison_type='early_stopping',return_data=False, return_round_data=True, comparison_name='Qwen, MMLU, MAS, both')
# all_results.append(result)

# transcript_data = {
#     'ChilleD/StrategyQA: No Mal Agent': 'transcripts/MAS/both/qwen3b_2026-03-29_15h26m25s',
#     'ChilleD/StrategyQA: Mal Agent 1 & 2': 'transcripts/MAS/early_stopping/2ESagents/strategyQA/agents0_1/qwen3b_job544_2026-04-23_23h06m55s',
#     'ChilleD/StrategyQA: Mal Agent 1 & 3': 'transcripts/MAS/early_stopping/2ESagents/strategyQA/agents0_2/qwen3b_job5452_2026-05-13_12h46m46s',
#     'ChilleD/StrategyQA: Mal Agent 2 & 3': 'transcripts/MAS/early_stopping/2ESagents/strategyQA/agents1_2/qwen3b_job2361_2026-05-15_15h53m17s',
# }
# # result = malicious_comparison(transcript_data, round_figure=False, comparison_type='early_stopping', return_data=True, comparison_name='Qwen, StrategyQA, MAS, both')
# # all_results.append(result)
# result = malicious_comparison(transcript_data, round_figure=True, comparison_type='early_stopping',return_data=False, return_round_data=True, comparison_name='Qwen, StrategyQA, MAS, both')
# all_results.append(result)

# transcript_data = {
#     'tasksource/bigbench: No Mal Agent': 'transcripts/MAS/both/qwen3b_2026-03-29_03h35m18s',
#     'tasksource/bigbench: Mal Agent 1 & 2': 'transcripts/MAS/early_stopping/2ESagents/bigbench/agents0_1/qwen3b_job547_2026-04-24_00h48m59s',
#     'tasksource/bigbench: Mal Agent 1 & 3': 'transcripts/MAS/early_stopping/2ESagents/bigbench/agents0_2/qwen3b_job5997_2026-05-14_11h11m49s',
#     'tasksource/bigbench: Mal Agent 2 & 3': 'transcripts/MAS/early_stopping/2ESagents/bigbench/agents1_2/qwen3b_job4789_2026-05-15_22h19m38s'
# }
# # result = malicious_comparison(transcript_data, round_figure=False, comparison_type='early_stopping', return_data=True, comparison_name='Qwen, BigBench, MAS, both')
# # all_results.append(result)
# result = malicious_comparison(transcript_data, round_figure=True, comparison_type='early_stopping',return_data=False, return_round_data=True, comparison_name='Qwen, BigBench, MAS, both')
# all_results.append(result)



# # --- Early Stopping comparison for Olmo ---
# transcript_data = {
#     'openai/gsm8k: No Mal Agent': 'transcripts/MAS/both/olmo7b_2026-03-29_15h41m56s',
#     'openai/gsm8k: Mal Agent 1 & 2': 'transcripts/MAS/early_stopping/2ESagents/gsm8k/agents0_1/olmo7b_job5478_2026-04-20_11h06m32s',
#     'openai/gsm8k: Mal Agent 1 & 3': 'transcripts/MAS/early_stopping/2ESagents/gsm8k/agents0_2/olmo7b_job5445_2026-05-13_04h01m14s',
#     'openai/gsm8k: Mal Agent 2 & 3': 'transcripts/MAS/early_stopping/2ESagents/gsm8k/agents1_2/olmo7b_job7950_2026-05-14_21h24m28s'
# }
# # result = malicious_comparison(transcript_data, round_figure=False, comparison_type='early_stopping', return_data=True, comparison_name='Olmo, GSM8K, MAS, both')
# # all_results.append(result)
# result = malicious_comparison(transcript_data, round_figure=True, comparison_type='early_stopping',return_data=False, return_round_data=True, comparison_name='Olmo, GSM8K, MAS, both')
# all_results.append(result)

# transcript_data = {
#     'cais/mmlu: No Mal Agent': 'transcripts/MAS/both/olmo7b_2026-03-29_15h29m53s',
#     'cais/mmlu: Mal Agent 1 & 2': 'transcripts/MAS/early_stopping/2ESagents/mmlu/agents0_1/olmo7b_job5481_2026-04-20_11h52m10s',
#     'cais/mmlu: Mal Agent 1 & 3': 'transcripts/MAS/early_stopping/2ESagents/mmlu/agents0_2/olmo7b_job5448_2026-05-13_11h50m22s',
#     'cais/mmlu: Mal Agent 2 & 3': 'transcripts/MAS/early_stopping/2ESagents/mmlu/agents1_2/olmo7b_job1015_2026-05-15_06h20m21s'
# }
# # result = malicious_comparison(transcript_data, round_figure=False, comparison_type='early_stopping', return_data=True, comparison_name='Olmo, MMLU, MAS, both')
# # all_results.append(result)
# result = malicious_comparison(transcript_data, round_figure=True, comparison_type='early_stopping',return_data=False, return_round_data=True, comparison_name='Olmo, MMLU, MAS, both')
# all_results.append(result)

# transcript_data = {
#     'ChilleD/StrategyQA: No Mal Agent': 'transcripts/MAS/both/olmo7b_2026-03-29_15h26m25s',
#     'ChilleD/StrategyQA: Mal Agent 1 & 2': 'transcripts/MAS/early_stopping/2ESagents/strategyQA/agents0_1/olmo7b_job545_2026-04-24_00h48m59s',
#     'ChilleD/StrategyQA: Mal Agent 1 & 3': 'transcripts/MAS/early_stopping/2ESagents/strategyQA/agents0_2/olmo7b_job5453_2026-05-13_14h03m37s',
#     'ChilleD/StrategyQA: Mal Agent 2 & 3': 'transcripts/MAS/early_stopping/2ESagents/strategyQA/agents1_2/olmo7b_job4473_2026-05-15_20h05m15s'
# }
# # result = malicious_comparison(transcript_data, round_figure=False, comparison_type='early_stopping', return_data=True, comparison_name='Olmo, StrategyQA, MAS, both')
# # all_results.append(result)
# result = malicious_comparison(transcript_data, round_figure=True, comparison_type='early_stopping',return_data=False, return_round_data=True, comparison_name='Olmo, StrategyQA, MAS, both')
# all_results.append(result)

# transcript_data = {
#     'tasksource/bigbench: No Mal Agent': 'transcripts/MAS/both/olmo7b_2026-03-29_01h47m46s',
#     'tasksource/bigbench: Mal Agent 1 & 2': 'transcripts/MAS/early_stopping/2ESagents/bigbench/agents0_1/olmo7b_job548_2026-04-24_00h52m24s',
#     'tasksource/bigbench: Mal Agent 1 & 3': 'transcripts/MAS/early_stopping/2ESagents/bigbench/agents0_2/olmo7b_job6680_2026-05-14_17h03m26s',
#     'tasksource/bigbench: Mal Agent 2 & 3': 'transcripts/MAS/early_stopping/2ESagents/bigbench/agents1_2/olmo7b_job4790_2026-05-15_23h16m12s'
# }
# # result = malicious_comparison(transcript_data, round_figure=False, comparison_type='early_stopping', return_data=True, comparison_name='Olmo, BigBench, MAS, both')
# # all_results.append(result)
# result = malicious_comparison(transcript_data, round_figure=True, comparison_type='early_stopping',return_data=False, return_round_data=True, comparison_name='Olmo, BigBench, MAS, both')
# all_results.append(result)



# # --- Early Stopping comparison for Olmo Think ---
# transcript_data = {
#     'openai/gsm8k: No Mal Agent': 'transcripts/MAS/both/olmo7bThink_job3961_2026-04-25_15h14m16s',
#     'openai/gsm8k: Mal Agent 1 & 2': 'transcripts/MAS/early_stopping/2ESagents/gsm8k/agents0_1/olmo7bThink_job5418_2026-04-13_21h04m25s',
#     'openai/gsm8k: Mal Agent 1 & 3': 'transcripts/MAS/early_stopping/2ESagents/gsm8k/agents0_2/olmo7bThink_job9386_2026-05-20_18h26m26s',
#     'openai/gsm8k: Mal Agent 2 & 3': 'transcripts/MAS/early_stopping/2ESagents/gsm8k/agents1_2/olmo7bThink_job9880_2026-05-21_07h41m44s'

# }
# result = malicious_comparison(transcript_data, round_figure=False, comparison_type='early_stopping', return_data=True, comparison_name='Olmo Think, GSM8K, MAS, both')
# all_results.append(result)

# transcript_data = {
#     'cais/mmlu: No Mal Agent': 'transcripts/MAS/both/olmo7bThink_job3963_2026-04-25_15h15m10s',
#     'cais/mmlu: Mal Agent 1 & 2': 'transcripts/MAS/early_stopping/2ESagents/mmlu/agents0_1/olmo7bThink_job5419_2026-04-13_23h20m25s',
#     'cais/mmlu: Mal Agent 1 & 3': 'transcripts/MAS/early_stopping/2ESagents/mmlu/agents0_2/olmo7bThink_job9387_2026-05-20_18h26m26s',
#     'cais/mmlu: Mal Agent 2 & 3': 'transcripts/MAS/early_stopping/2ESagents/mmlu/agents1_2/olmo7bThink_job9881_2026-05-21_08h57m21s'
# }
# result = malicious_comparison(transcript_data, round_figure=False, comparison_type='early_stopping', return_data=True, comparison_name='Olmo Think, MMLU, MAS, both')
# all_results.append(result)

# transcript_data = {
#     'ChilleD/StrategyQA: No Mal Agent': 'transcripts/MAS/both/olmo7bThink_job3965_2026-04-25_15h20m06s',
#     'ChilleD/StrategyQA: Mal Agent 1 & 2': 'transcripts/MAS/early_stopping/2ESagents/strategyQA/agents0_1/olmo7bThink_job5420_2026-04-14_07h41m13s',
#     'ChilleD/StrategyQA: Mal Agent 1 & 3': 'transcripts/MAS/early_stopping/2ESagents/strategyQA/agents0_2/olmo7bThink_job9388_2026-05-20_18h42m31s',
#     'ChilleD/StrategyQA: Mal Agent 2 & 3': 'transcripts/MAS/early_stopping/2ESagents/strategyQA/agents1_2/olmo7bThink_job9882_2026-05-21_10h28m05s'
# }
# result = malicious_comparison(transcript_data, round_figure=False, comparison_type='early_stopping', return_data=True, comparison_name='Olmo Think, StrategyQA, MAS, both')
# all_results.append(result)

# transcript_data = {
#     'tasksource/bigbench: No Mal Agent': 'transcripts/MAS/both/olmo7bThink_job3967_2026-04-25_15h20m06s',
#     'tasksource/bigbench: Mal Agent 1 & 2': 'transcripts/MAS/early_stopping/2ESagents/bigbench/agents0_1/olmo7bThink_job5422_2026-04-14_12h38m15s',
#     'tasksource/bigbench: Mal Agent 1 & 3': 'transcripts/MAS/early_stopping/2ESagents/bigbench/agents0_2/olmo7bThink_job9389_2026-05-20_18h45m03s',
#     # 'tasksource/bigbench: Mal Agent 2 & 3': 'transcripts/MAS/early_stopping/2ESagents/bigbench/agents1_2/...'
# }
# result = malicious_comparison(transcript_data, round_figure=False, comparison_type='early_stopping', return_data=True, comparison_name='Olmo Think, BigBench, MAS, both')
# all_results.append(result)



# # --- Early Stopping comparison for Llama ---
# transcript_data = {
#     'openai/gsm8k: No Mal Agent': 'transcripts/MAS/both/llama3b_2026-03-29_15h19m52s',
#     'openai/gsm8k: Mal Agent 1 & 2': 'transcripts/MAS/early_stopping/2ESagents/gsm8k/agents0_1/llama3b_job5476_2026-04-20_11h02m50s',
#     'openai/gsm8k: Mal Agent 1 & 3': 'transcripts/MAS/early_stopping/2ESagents/gsm8k/agents0_2/llama3b_job5443_2026-05-12_23h12m38s',
#     'openai/gsm8k: Mal Agent 2 & 3': 'transcripts/MAS/early_stopping/2ESagents/gsm8k/agents1_2/llama3b_job7948_2026-05-14_18h02m42s'
# }
# # result = malicious_comparison(transcript_data, round_figure=False, comparison_type='early_stopping', return_data=True, comparison_name='Llama, GSM8K, MAS, both')
# # all_results.append(result)
# result = malicious_comparison(transcript_data, round_figure=True, comparison_type='early_stopping',return_data=False, return_round_data=True, comparison_name='Llama, GSM8K, MAS, both')
# all_results.append(result)

# transcript_data = {
#     'cais/mmlu: No Mal Agent': 'transcripts/MAS/both/llama3b_2026-03-29_00h34m49s',
#     'cais/mmlu: Mal Agent 1 & 2': 'transcripts/MAS/early_stopping/2ESagents/mmlu/agents0_1/llama3b_job7227_2026-04-21_12h52m23s',
#     'cais/mmlu: Mal Agent 1 & 3': 'transcripts/MAS/early_stopping/2ESagents/mmlu/agents0_2/llama3b_job5446_2026-05-13_06h29m50s',
#     'cais/mmlu: Mal Agent 2 & 3': 'transcripts/MAS/early_stopping/2ESagents/mmlu/agents1_2/llama3b_job7828_2026-05-14_17h15m06s'
# }
# # result = malicious_comparison(transcript_data, round_figure=False, comparison_type='early_stopping', return_data=True, comparison_name='Llama, MMLU, MAS, both')
# # all_results.append(result)
# result = malicious_comparison(transcript_data, round_figure=True, comparison_type='early_stopping',return_data=False, return_round_data=True, comparison_name='Llama, MMLU, MAS, both')
# all_results.append(result)

# transcript_data = {
#     'ChilleD/StrategyQA: No Mal Agent': 'transcripts/MAS/both/llama3b_2026-03-29_15h21m49s',
#     'ChilleD/StrategyQA: Mal Agent 1 & 2': 'transcripts/MAS/early_stopping/2ESagents/strategyQA/agents0_1/llama3b_job543_2026-04-23_18h59m23s',
#     'ChilleD/StrategyQA: Mal Agent 1 & 3': 'transcripts/MAS/early_stopping/2ESagents/strategyQA/agents0_2/llama3b_job5451_2026-05-13_12h19m58s',
#     'ChilleD/StrategyQA: Mal Agent 2 & 3': 'transcripts/MAS/early_stopping/2ESagents/strategyQA/agents1_2/llama3b_job2358_2026-05-15_07h30m17s'
# }
# # result = malicious_comparison(transcript_data, round_figure=False, comparison_type='early_stopping', return_data=True, comparison_name='Llama, StrategyQA, MAS, both')
# # all_results.append(result)
# result = malicious_comparison(transcript_data, round_figure=True, comparison_type='early_stopping',return_data=False, return_round_data=True, comparison_name='Llama, StrategyQA, MAS, both')
# all_results.append(result)

# transcript_data = {
#     'tasksource/bigbench: No Mal Agent': 'transcripts/MAS/both/llama3b_2026-03-29_03h08m47s',
#     'tasksource/bigbench: Mal Agent 1 & 2': 'transcripts/MAS/early_stopping/2ESagents/bigbench/agents0_1/llama3b_job546_2026-04-24_00h48m59s',
#     'tasksource/bigbench: Mal Agent 1 & 3': 'transcripts/MAS/early_stopping/2ESagents/bigbench/agents0_2/llama3b_job5996_2026-05-14_11h11m49s',
#     'tasksource/bigbench: Mal Agent 2 & 3': 'transcripts/MAS/early_stopping/2ESagents/bigbench/agents1_2/llama3b_job4788_2026-05-15_22h19m38s'
# }
# # result = malicious_comparison(transcript_data, round_figure=False, comparison_type='early_stopping', return_data=True, comparison_name='Llama, BigBench, MAS, both')
# # all_results.append(result)
# result = malicious_comparison(transcript_data, round_figure=True, comparison_type='early_stopping',return_data=False, return_round_data=True, comparison_name='Llama, BigBench, MAS, both')
# all_results.append(result)

# # export_comparisons_to_excel(all_results, 'early_stopping_comparisons.xlsx')

# # export_comparisons_to_excel(all_results, 'dataset_comparison_results_new.xlsx')

# export_round_comparisons_to_excel(all_results, 'ES_round_results.xlsx')

# # ╔═════════════════════════════════════════════════╗
# # ║           IR-CA ONE AGENT COMPARISON            ║
# # ╚═════════════════════════════════════════════════╝

# # --- IR-CA comparison for Qwen ---
# transcript_data = {
#     'openai/gsm8k: No Mal Agent': 'transcripts/MAS/both/qwen3b_2026-03-29_00h31m15s',
#     'openai/gsm8k: Mal Agent 1': 'transcripts/MAS/unfaithfulReasoning/wrongReason_correctAnswer/1agent/gsm8k/agent0/qwen3b_job8439_2026-05-04_20h58m52s',
#     'openai/gsm8k: Mal Agent 2': 'transcripts/MAS/unfaithfulReasoning/wrongReason_correctAnswer/1agent/gsm8k/agent1/qwen3b_job3915_2026-05-06_11h19m58s',
#     'openai/gsm8k: Mal Agent 3': 'transcripts/MAS/unfaithfulReasoning/wrongReason_correctAnswer/1agent/gsm8k/agent2/qwen3b_job7909_2026-05-08_02h18m30s',
# }
# # result = malicious_comparison(transcript_data, round_figure=False, comparison_type='IR-CA', return_data=True, comparison_name='Qwen, GSM8K, MAS, both')
# # all_results.append(result)
# result = malicious_comparison(transcript_data, round_figure=True, comparison_type='IR-CA',return_data=False, return_round_data=True, comparison_name='Qwen, GSM8K, MAS, both')
# all_results.append(result)

# transcript_data = {
#     'cais/mmlu: No Mal Agent': 'transcripts/MAS/both/qwen3b_2026-03-29_15h34m56s',
#     'cais/mmlu: Mal Agent 1': 'transcripts/MAS/unfaithfulReasoning/wrongReason_correctAnswer/1agent/mmlu/agent0/qwen3b_job8438_2026-05-04_20h42m42s',
#     'cais/mmlu: Mal Agent 2': 'transcripts/MAS/unfaithfulReasoning/wrongReason_correctAnswer/1agent/mmlu/agent1/qwen3b_job3918_2026-05-06_12h44m15s',
#     'cais/mmlu: Mal Agent 3': 'transcripts/MAS/unfaithfulReasoning/wrongReason_correctAnswer/1agent/mmlu/agent2/qwen3b_job1053_2026-05-09_20h39m55s',
# }
# # result = malicious_comparison(transcript_data, round_figure=False, comparison_type='IR-CA', return_data=True, comparison_name='Qwen, MMLU, MAS, both')
# # all_results.append(result)
# result = malicious_comparison(transcript_data, round_figure=True, comparison_type='IR-CA',return_data=False, return_round_data=True, comparison_name='Qwen, MMLU, MAS, both')
# all_results.append(result)

# transcript_data = {
#     'ChilleD/StrategyQA: No Mal Agent': 'transcripts/MAS/both/qwen3b_2026-03-29_15h26m25s',
#     'ChilleD/StrategyQA: Mal Agent 1': 'transcripts/MAS/unfaithfulReasoning/wrongReason_correctAnswer/1agent/strategyQA/agent0/qwen3b_job8443_2026-05-05_08h54m11s',
#     'ChilleD/StrategyQA: Mal Agent 2': 'transcripts/MAS/unfaithfulReasoning/wrongReason_correctAnswer/1agent/strategyQA/agent1/qwen3b_job5171_2026-05-06_23h27m51s',
#     'ChilleD/StrategyQA: Mal Agent 3': 'transcripts/MAS/unfaithfulReasoning/wrongReason_correctAnswer/1agent/strategyQA/agent2/qwen3b_job1056_2026-05-10_01h36m02s',
# }
# # result = malicious_comparison(transcript_data, round_figure=False, comparison_type='IR-CA', return_data=True, comparison_name='Qwen, StrategyQA, MAS, both')
# # all_results.append(result)
# result = malicious_comparison(transcript_data, round_figure=True, comparison_type='IR-CA',return_data=False, return_round_data=True, comparison_name='Qwen, StrategyQA, MAS, both')
# all_results.append(result)

# transcript_data = {
#     'tasksource/bigbench: No Mal Agent': 'transcripts/MAS/both/qwen3b_2026-03-29_03h35m18s',
#     'tasksource/bigbench: Mal Agent 1': 'transcripts/MAS/unfaithfulReasoning/wrongReason_correctAnswer/1agent/bigbench/agent0/qwen3b_job253_2026-05-05_18h41m07s',
#     'tasksource/bigbench: Mal Agent 2': 'transcripts/MAS/unfaithfulReasoning/wrongReason_correctAnswer/1agent/bigbench/agent1/qwen3b_job7906_2026-05-08_01h01m56s',
#     'tasksource/bigbench: Mal Agent 3': 'transcripts/MAS/unfaithfulReasoning/wrongReason_correctAnswer/1agent/bigbench/agent2/qwen3b_job1059_2026-05-10_09h31m10s',
# }
# # result = malicious_comparison(transcript_data, round_figure=False, comparison_type='IR-CA', return_data=True, comparison_name='Qwen, BigBench, MAS, both')
# # all_results.append(result)
# result = malicious_comparison(transcript_data, round_figure=True, comparison_type='IR-CA',return_data=False, return_round_data=True, comparison_name='Qwen, BigBench, MAS, both')
# all_results.append(result)


# # --- IR-CA comparison for Olmo ---
# transcript_data = {
#     'openai/gsm8k: No Mal Agent': 'transcripts/MAS/both/olmo7b_2026-03-29_15h41m56s',
#     'openai/gsm8k: Mal Agent 1': 'transcripts/MAS/unfaithfulReasoning/wrongReason_correctAnswer/1agent/gsm8k/agent0/olmo7b_job8441_2026-05-05_02h59m15s',
#     'openai/gsm8k: Mal Agent 2': 'transcripts/MAS/unfaithfulReasoning/wrongReason_correctAnswer/1agent/gsm8k/agent1/olmo7b_job3916_2026-05-06_11h21m25s',
#     'openai/gsm8k: Mal Agent 3': 'transcripts/MAS/unfaithfulReasoning/wrongReason_correctAnswer/1agent/gsm8k/agent2/olmo7b_job7910_2026-05-08_02h28m27s',
# }
# # result = malicious_comparison(transcript_data, round_figure=False, comparison_type='IR-CA', return_data=True, comparison_name='Olmo, GSM8K, MAS, both')
# # all_results.append(result)
# result = malicious_comparison(transcript_data, round_figure=True, comparison_type='IR-CA',return_data=False, return_round_data=True, comparison_name='Olmo, GSM8K, MAS, both')
# all_results.append(result)

# transcript_data = {
#     'cais/mmlu: No Mal Agent': 'transcripts/MAS/both/olmo7b_2026-03-29_15h29m53s',
#     'cais/mmlu: Mal Agent 1': 'transcripts/MAS/unfaithfulReasoning/wrongReason_correctAnswer/1agent/mmlu/agent0/olmo7b_job8440_2026-05-05_00h56m03s',
#     'cais/mmlu: Mal Agent 2': 'transcripts/MAS/unfaithfulReasoning/wrongReason_correctAnswer/1agent/mmlu/agent1/olmo7b_job3919_2026-05-06_13h02m23s',
#     'cais/mmlu: Mal Agent 3': 'transcripts/MAS/unfaithfulReasoning/wrongReason_correctAnswer/1agent/mmlu/agent2/olmo7b_job1054_2026-05-09_20h53m56s',
# }
# # result = malicious_comparison(transcript_data, round_figure=False, comparison_type='IR-CA', return_data=True, comparison_name='Olmo, MMLU, MAS, both')
# # all_results.append(result)
# result = malicious_comparison(transcript_data, round_figure=True, comparison_type='IR-CA',return_data=False, return_round_data=True, comparison_name='Olmo, MMLU, MAS, both')
# all_results.append(result)

# transcript_data = {
#     'ChilleD/StrategyQA: No Mal Agent': 'transcripts/MAS/both/olmo7b_2026-03-29_15h26m25s',
#     'ChilleD/StrategyQA: Mal Agent 1': 'transcripts/MAS/unfaithfulReasoning/wrongReason_correctAnswer/1agent/strategyQA/agent0/olmo7b_job8444_2026-05-05_09h30m47s',
#     'ChilleD/StrategyQA: Mal Agent 2': 'transcripts/MAS/unfaithfulReasoning/wrongReason_correctAnswer/1agent/strategyQA/agent1/olmo7b_job5174_2026-05-06_23h56m08s',
#     'ChilleD/StrategyQA: Mal Agent 3': 'transcripts/MAS/unfaithfulReasoning/wrongReason_correctAnswer/1agent/strategyQA/agent2/olmo7b_job1057_2026-05-10_05h12m38s',
# }
# # result = malicious_comparison(transcript_data, round_figure=False, comparison_type='IR-CA', return_data=True, comparison_name='Olmo, StrategyQA, MAS, both')
# # all_results.append(result)
# result = malicious_comparison(transcript_data, round_figure=True, comparison_type='IR-CA',return_data=False, return_round_data=True, comparison_name='Olmo, StrategyQA, MAS, both')
# all_results.append(result)

# transcript_data = {
#     'tasksource/bigbench: No Mal Agent': 'transcripts/MAS/both/olmo7b_2026-03-29_01h47m46s',
#     'tasksource/bigbench: Mal Agent 1': 'transcripts/MAS/unfaithfulReasoning/wrongReason_correctAnswer/1agent/bigbench/agent0/olmo7b_job254_2026-05-05_22h57m24s',
#     'tasksource/bigbench: Mal Agent 2': 'transcripts/MAS/unfaithfulReasoning/wrongReason_correctAnswer/1agent/bigbench/agent1/olmo7b_job7907_2026-05-08_02h00m32s',
#     'tasksource/bigbench: Mal Agent 3': 'transcripts/MAS/unfaithfulReasoning/wrongReason_correctAnswer/1agent/bigbench/agent2/olmo7b_job1060_2026-05-10_11h18m48s',
# }
# # result = malicious_comparison(transcript_data, round_figure=False, comparison_type='IR-CA', return_data=True, comparison_name='Olmo, BigBench, MAS, both')
# # all_results.append(result)
# result = malicious_comparison(transcript_data, round_figure=True, comparison_type='IR-CA',return_data=False, return_round_data=True, comparison_name='Olmo, BigBench, MAS, both')
# all_results.append(result)


# # --- IR-CA comparison for Llama ---
# transcript_data = {
#     'openai/gsm8k: No Mal Agent': 'transcripts/MAS/both/llama3b_2026-03-29_15h19m52s',
#     'openai/gsm8k: Mal Agent 1': 'transcripts/MAS/unfaithfulReasoning/wrongReason_correctAnswer/1agent/gsm8k/agent0/llama3b_job8437_2026-05-04_18h19m45s',
#     'openai/gsm8k: Mal Agent 2': 'transcripts/MAS/unfaithfulReasoning/wrongReason_correctAnswer/1agent/gsm8k/agent1/llama3b_job3914_2026-05-06_01h23m27s',
#     'openai/gsm8k: Mal Agent 3': 'transcripts/MAS/unfaithfulReasoning/wrongReason_correctAnswer/1agent/gsm8k/agent2/llama3b_job7908_2026-05-08_02h14m28s',
# }
# # result = malicious_comparison(transcript_data, round_figure=False, comparison_type='IR-CA', return_data=True, comparison_name='Llama, GSM8K, MAS, both')
# # all_results.append(result)
# result = malicious_comparison(transcript_data, round_figure=True, comparison_type='IR-CA',return_data=False, return_round_data=True, comparison_name='Llama, GSM8K, MAS, both')
# all_results.append(result)

# transcript_data = {
#     'cais/mmlu: No Mal Agent': 'transcripts/MAS/both/llama3b_2026-03-29_00h34m49s',
#     'cais/mmlu: Mal Agent 1': 'transcripts/MAS/unfaithfulReasoning/wrongReason_correctAnswer/1agent/mmlu/agent0/llama3b_job8436_2026-05-04_18h11m38s',
#     'cais/mmlu: Mal Agent 2': 'transcripts/MAS/unfaithfulReasoning/wrongReason_correctAnswer/1agent/mmlu/agent1/llama3b_job3917_2026-05-06_11h24m12s',
#     'cais/mmlu: Mal Agent 3': 'transcripts/MAS/unfaithfulReasoning/wrongReason_correctAnswer/1agent/mmlu/agent2/llama3b_job1052_2026-05-09_16h13m04s',
# }
# # result = malicious_comparison(transcript_data, round_figure=False, comparison_type='IR-CA', return_data=True, comparison_name='Llama, MMLU, MAS, both')
# # all_results.append(result)
# result = malicious_comparison(transcript_data, round_figure=True, comparison_type='IR-CA',return_data=False, return_round_data=True, comparison_name='Llama, MMLU, MAS, both')
# all_results.append(result)

# transcript_data = {
#     'ChilleD/StrategyQA: No Mal Agent': 'transcripts/MAS/both/llama3b_2026-03-29_15h21m49s',
#     'ChilleD/StrategyQA: Mal Agent 1': 'transcripts/MAS/unfaithfulReasoning/wrongReason_correctAnswer/1agent/strategyQA/agent0/llama3b_job8442_2026-05-05_08h26m01s',
#     'ChilleD/StrategyQA: Mal Agent 2': 'transcripts/MAS/unfaithfulReasoning/wrongReason_correctAnswer/1agent/strategyQA/agent1/llama3b_job5170_2026-05-06_19h02m38s',
#     'ChilleD/StrategyQA: Mal Agent 3': 'transcripts/MAS/unfaithfulReasoning/wrongReason_correctAnswer/1agent/strategyQA/agent2/llama3b_job1055_2026-05-09_21h06m53s',
# }
# # result = malicious_comparison(transcript_data, round_figure=False, comparison_type='IR-CA', return_data=True, comparison_name='Llama, StrategyQA, MAS, both')
# # all_results.append(result)
# result = malicious_comparison(transcript_data, round_figure=True, comparison_type='IR-CA',return_data=False, return_round_data=True, comparison_name='Llama, StrategyQA, MAS, both')
# all_results.append(result)

# transcript_data = {
#     'tasksource/bigbench: No Mal Agent': 'transcripts/MAS/both/llama3b_2026-03-29_03h08m47s',
#     'tasksource/bigbench: Mal Agent 1': 'transcripts/MAS/unfaithfulReasoning/wrongReason_correctAnswer/1agent/bigbench/agent0/llama3b_job252_2026-05-05_18h16m44s',
#     'tasksource/bigbench: Mal Agent 2': 'transcripts/MAS/unfaithfulReasoning/wrongReason_correctAnswer/1agent/bigbench/agent1/llama3b_job7905_2026-05-07_20h54m03s',
#     'tasksource/bigbench: Mal Agent 3': 'transcripts/MAS/unfaithfulReasoning/wrongReason_correctAnswer/1agent/bigbench/agent2/llama3b_job1058_2026-05-10_05h18m36s',
# }
# # result = malicious_comparison(transcript_data, round_figure=False, comparison_type='IR-CA', return_data=True, comparison_name='Llama, BigBench, MAS, both')
# # all_results.append(result)
# result = malicious_comparison(transcript_data, round_figure=True, comparison_type='IR-CA',return_data=False, return_round_data=True, comparison_name='Llama, BigBench, MAS, both')
# all_results.append(result)


# # ╔═════════════════════════════════════════════════╗
# # ║           IR-CA TWO AGENT COMPARISON            ║
# # ╚═════════════════════════════════════════════════╝

# # --- IR-CA comparison for Qwen ---
# transcript_data = {
#     'openai/gsm8k: No Mal Agent': 'transcripts/MAS/both/qwen3b_2026-03-29_00h31m15s',
#     'openai/gsm8k: Mal Agent 1 & 2': 'transcripts/MAS/unfaithfulReasoning/wrongReason_correctAnswer/2agents/gsm8k/agents1_2/qwen3b_job8172_2026-05-08_04h03m51s',
#     'openai/gsm8k: Mal Agent 1 & 3': 'transcripts/MAS/unfaithfulReasoning/wrongReason_correctAnswer/2agents/gsm8k/agents1_3/qwen3b_job2342_2026-05-11_06h38m17s',
#     'openai/gsm8k: Mal Agent 2 & 3': 'transcripts/MAS/unfaithfulReasoning/wrongReason_correctAnswer/2agents/gsm8k/agents2_3/qwen3b_job1719_2026-05-10_15h55m59s',
# }
# result = malicious_comparison(transcript_data, round_figure=False, comparison_type='IR-CA', return_data=True, comparison_name='Qwen, GSM8K, MAS, both')
# # all_results.append(result)
# result = malicious_comparison(transcript_data, round_figure=True, comparison_type='IR-CA',return_data=False, return_round_data=True, comparison_name='Qwen, GSM8K, MAS, both')
# all_results.append(result)

# transcript_data = {
#     'cais/mmlu: No Mal Agent': 'transcripts/MAS/both/qwen3b_2026-03-29_15h34m56s',
#     'cais/mmlu: Mal Agent 1 & 2': 'transcripts/MAS/unfaithfulReasoning/wrongReason_correctAnswer/2agents/mmlu/agents1_2/qwen3b_job1027_2026-05-09_15h24m59s',
#     'cais/mmlu: Mal Agent 1 & 3': 'transcripts/MAS/unfaithfulReasoning/wrongReason_correctAnswer/2agents/mmlu/agents1_3/qwen3b_job2346_2026-05-11_08h55m53s',
#     'cais/mmlu: Mal Agent 2 & 3': 'transcripts/MAS/unfaithfulReasoning/wrongReason_correctAnswer/2agents/mmlu/agents2_3/qwen3b_job1724_2026-05-10_17h04m34s',
# }
# result = malicious_comparison(transcript_data, round_figure=False, comparison_type='IR-CA', return_data=True, comparison_name='Qwen, MMLU, MAS, both')
# # all_results.append(result)
# result = malicious_comparison(transcript_data, round_figure=True, comparison_type='IR-CA',return_data=False, return_round_data=True, comparison_name='Qwen, MMLU, MAS, both')
# all_results.append(result)

# transcript_data = {
#     'ChilleD/StrategyQA: No Mal Agent': 'transcripts/MAS/both/qwen3b_2026-03-29_15h26m25s',
#     'ChilleD/StrategyQA: Mal Agent 1 & 2': 'transcripts/MAS/unfaithfulReasoning/wrongReason_correctAnswer/2agents/strategyQA/agents1_2/qwen3b_job1031_2026-05-09_15h27m58s',
#     'ChilleD/StrategyQA: Mal Agent 1 & 3': 'transcripts/MAS/unfaithfulReasoning/wrongReason_correctAnswer/2agents/strategyQA/agents1_3/qwen3b_job2351_2026-05-11_14h04m30s',
#     'ChilleD/StrategyQA: Mal Agent 2 & 3': 'transcripts/MAS/unfaithfulReasoning/wrongReason_correctAnswer/2agents/strategyQA/agents2_3/qwen3b_job2354_2026-05-11_00h54m40s',
# }
# result = malicious_comparison(transcript_data, round_figure=False, comparison_type='IR-CA', return_data=True, comparison_name='Qwen, StrategyQA, MAS, both')
# # all_results.append(result)
# result = malicious_comparison(transcript_data, round_figure=True, comparison_type='IR-CA',return_data=False, return_round_data=True, comparison_name='Qwen, StrategyQA, MAS, both')
# all_results.append(result)

# transcript_data = {
#     'tasksource/bigbench: No Mal Agent': 'transcripts/MAS/both/qwen3b_2026-03-29_03h35m18s',
#     'tasksource/bigbench: Mal Agent 1 & 2': 'transcripts/MAS/unfaithfulReasoning/wrongReason_correctAnswer/2agents/bigbench/agents1_2/qwen3b_job1156_2026-05-10_12h39m31s',
#     'tasksource/bigbench: Mal Agent 1 & 3': 'transcripts/MAS/unfaithfulReasoning/wrongReason_correctAnswer/2agents/bigbench/agents1_3/qwen3b_job4455_2026-05-12_11h02m17s',
#     'tasksource/bigbench: Mal Agent 2 & 3': 'transcripts/MAS/unfaithfulReasoning/wrongReason_correctAnswer/2agents/bigbench/agents2_3/qwen3b_job2639_2026-05-11_18h34m16s',
# }
# result = malicious_comparison(transcript_data, round_figure=False, comparison_type='IR-CA', return_data=True, comparison_name='Qwen, BigBench, MAS, both')
# # all_results.append(result)
# result = malicious_comparison(transcript_data, round_figure=True, comparison_type='IR-CA',return_data=False, return_round_data=True, comparison_name='Qwen, BigBench, MAS, both')
# all_results.append(result)


# # --- IR-CA comparison for Olmo ---
# transcript_data = {
#     'openai/gsm8k: No Mal Agent': 'transcripts/MAS/both/olmo7b_2026-03-29_15h41m56s',
#     'openai/gsm8k: Mal Agent 1 & 2': 'transcripts/MAS/unfaithfulReasoning/wrongReason_correctAnswer/2agents/gsm8k/agents1_2/olmo7b_job8173_2026-05-08_04h04m01s',
#     'openai/gsm8k: Mal Agent 1 & 3': 'transcripts/MAS/unfaithfulReasoning/wrongReason_correctAnswer/2agents/gsm8k/agents1_3/olmo7b_job2343_2026-05-11_08h38m42s',
#     'openai/gsm8k: Mal Agent 2 & 3': 'transcripts/MAS/unfaithfulReasoning/wrongReason_correctAnswer/2agents/gsm8k/agents2_3/olmo7b_job1720_2026-05-10_16h27m23s',
# }
# # result = malicious_comparison(transcript_data, round_figure=False, comparison_type='IR-CA', return_data=True, comparison_name='Olmo, GSM8K, MAS, both')
# # all_results.append(result)
# result = malicious_comparison(transcript_data, round_figure=True, comparison_type='IR-CA',return_data=False, return_round_data=True, comparison_name='Olmo, GSM8K, MAS, both')
# all_results.append(result)

# transcript_data = {
#     'cais/mmlu: No Mal Agent': 'transcripts/MAS/both/olmo7b_2026-03-29_15h29m53s',
#     'cais/mmlu: Mal Agent 1 & 2': 'transcripts/MAS/unfaithfulReasoning/wrongReason_correctAnswer/2agents/mmlu/agents1_2/olmo7b_job1028_2026-05-09_15h24m59s',
#     'cais/mmlu: Mal Agent 1 & 3': 'transcripts/MAS/unfaithfulReasoning/wrongReason_correctAnswer/2agents/mmlu/agents1_3/olmo7b_job2347_2026-05-11_09h14m50s',
#     'cais/mmlu: Mal Agent 2 & 3': 'transcripts/MAS/unfaithfulReasoning/wrongReason_correctAnswer/2agents/mmlu/agents2_3/olmo7b_job1725_2026-05-10_17h07m30s',
# }
# # result = malicious_comparison(transcript_data, round_figure=False, comparison_type='IR-CA', return_data=True, comparison_name='Olmo, MMLU, MAS, both')
# # all_results.append(result)
# result = malicious_comparison(transcript_data, round_figure=True, comparison_type='IR-CA',return_data=False, return_round_data=True, comparison_name='Olmo, MMLU, MAS, both')
# all_results.append(result)

# transcript_data = {
#     'ChilleD/StrategyQA: No Mal Agent': 'transcripts/MAS/both/olmo7b_2026-03-29_15h26m25s',
#     'ChilleD/StrategyQA: Mal Agent 1 & 2': 'transcripts/MAS/unfaithfulReasoning/wrongReason_correctAnswer/2agents/strategyQA/agents1_2/olmo7b_job1032_2026-05-09_15h26m46s',
#     'ChilleD/StrategyQA: Mal Agent 1 & 3': 'transcripts/MAS/unfaithfulReasoning/wrongReason_correctAnswer/2agents/strategyQA/agents1_3/olmo7b_job2352_2026-05-11_14h21m58s',
#     'ChilleD/StrategyQA: Mal Agent 2 & 3': 'transcripts/MAS/unfaithfulReasoning/wrongReason_correctAnswer/2agents/strategyQA/agents2_3/olmo7b_job2355_2026-05-11_01h02m18s',
# }
# # result = malicious_comparison(transcript_data, round_figure=False, comparison_type='IR-CA', return_data=True, comparison_name='Olmo, StrategyQA, MAS, both')
# # all_results.append(result)
# result = malicious_comparison(transcript_data, round_figure=True, comparison_type='IR-CA',return_data=False, return_round_data=True, comparison_name='Olmo, StrategyQA, MAS, both')
# all_results.append(result)

# transcript_data = {
#     'tasksource/bigbench: No Mal Agent': 'transcripts/MAS/both/olmo7b_2026-03-29_01h47m46s',
#     'tasksource/bigbench: Mal Agent 1 & 2': 'transcripts/MAS/unfaithfulReasoning/wrongReason_correctAnswer/2agents/bigbench/agents1_2/olmo7b_job1157_2026-05-10_14h38m48s',
#     'tasksource/bigbench: Mal Agent 1 & 3': 'transcripts/MAS/unfaithfulReasoning/wrongReason_correctAnswer/2agents/bigbench/agents1_3/olmo7b_job4454_2026-05-12_09h19m47s',
#     'tasksource/bigbench: Mal Agent 2 & 3': 'transcripts/MAS/unfaithfulReasoning/wrongReason_correctAnswer/2agents/bigbench/agents2_3/olmo7b_job2640_2026-05-11_18h42m58s',
# }
# # result = malicious_comparison(transcript_data, round_figure=False, comparison_type='IR-CA', return_data=True, comparison_name='Olmo, BigBench, MAS, both')
# # all_results.append(result)
# result = malicious_comparison(transcript_data, round_figure=True, comparison_type='IR-CA',return_data=False, return_round_data=True, comparison_name='Olmo, BigBench, MAS, both')
# all_results.append(result)


# # --- IR-CA comparison for Llama ---
# transcript_data = {
#     'openai/gsm8k: No Mal Agent': 'transcripts/MAS/both/llama3b_2026-03-29_15h19m52s',
#     'openai/gsm8k: Mal Agent 1 & 2': 'transcripts/MAS/unfaithfulReasoning/wrongReason_correctAnswer/2agents/gsm8k/agents1_2/llama3b_job8171_2026-05-08_03h24m46s',
#     'openai/gsm8k: Mal Agent 1 & 3': 'transcripts/MAS/unfaithfulReasoning/wrongReason_correctAnswer/2agents/gsm8k/agents1_3/llama3b_job2341_2026-05-11_00h36m05s',
#     'openai/gsm8k: Mal Agent 2 & 3': 'transcripts/MAS/unfaithfulReasoning/wrongReason_correctAnswer/2agents/gsm8k/agents2_3/llama3b_job1718_2026-05-10_15h30m17s',
# }
# # result = malicious_comparison(transcript_data, round_figure=False, comparison_type='IR-CA', return_data=True, comparison_name='Llama, GSM8K, MAS, both')
# # all_results.append(result)
# result = malicious_comparison(transcript_data, round_figure=True, comparison_type='IR-CA',return_data=False, return_round_data=True, comparison_name='Llama, GSM8K, MAS, both')
# all_results.append(result)

# transcript_data = {
#     'cais/mmlu: No Mal Agent': 'transcripts/MAS/both/llama3b_2026-03-29_00h34m49s',
#     'cais/mmlu: Mal Agent 1 & 2': 'transcripts/MAS/unfaithfulReasoning/wrongReason_correctAnswer/2agents/mmlu/agents1_2/llama3b_job1026_2026-05-09_15h24m59s',
#     'cais/mmlu: Mal Agent 1 & 3': 'transcripts/MAS/unfaithfulReasoning/wrongReason_correctAnswer/2agents/mmlu/agents1_3/llama3b_job2345_2026-05-11_08h43m38s',
#     'cais/mmlu: Mal Agent 2 & 3': 'transcripts/MAS/unfaithfulReasoning/wrongReason_correctAnswer/2agents/mmlu/agents2_3/llama3b_job1723_2026-05-10_16h58m35s',
# }
# # result = malicious_comparison(transcript_data, round_figure=False, comparison_type='IR-CA', return_data=True, comparison_name='Llama, MMLU, MAS, both')
# # all_results.append(result)
# result = malicious_comparison(transcript_data, round_figure=True, comparison_type='IR-CA',return_data=False, return_round_data=True, comparison_name='Llama, MMLU, MAS, both')
# all_results.append(result)


# transcript_data = {
#     'ChilleD/StrategyQA: No Mal Agent': 'transcripts/MAS/both/llama3b_2026-03-29_15h21m49s',
#     'ChilleD/StrategyQA: Mal Agent 1 & 2': 'transcripts/MAS/unfaithfulReasoning/wrongReason_correctAnswer/2agents/strategyQA/agents1_2/llama3b_job1030_2026-05-09_15h27m58s',
#     'ChilleD/StrategyQA: Mal Agent 1 & 3': 'transcripts/MAS/unfaithfulReasoning/wrongReason_correctAnswer/2agents/strategyQA/agents1_3/llama3b_job2350_2026-05-11_10h32m38s',
#     'ChilleD/StrategyQA: Mal Agent 2 & 3': 'transcripts/MAS/unfaithfulReasoning/wrongReason_correctAnswer/2agents/strategyQA/agents2_3/llama3b_job2353_2026-05-11_00h50m11s',
# }
# # result = malicious_comparison(transcript_data, round_figure=False, comparison_type='IR-CA', return_data=True, comparison_name='Llama, StrategyQA, MAS, both')
# # all_results.append(result)
# result = malicious_comparison(transcript_data, round_figure=True, comparison_type='IR-CA',return_data=False, return_round_data=True, comparison_name='Llama, StrategyQA, MAS, both')
# all_results.append(result)


# transcript_data = {
#     'tasksource/bigbench: No Mal Agent': 'transcripts/MAS/both/llama3b_2026-03-29_03h08m47s',
#     'tasksource/bigbench: Mal Agent 1 & 2': 'transcripts/MAS/unfaithfulReasoning/wrongReason_correctAnswer/2agents/bigbench/agents1_2/llama3b_job1155_2026-05-10_12h18m04s',
#     'tasksource/bigbench: Mal Agent 1 & 3': 'transcripts/MAS/unfaithfulReasoning/wrongReason_correctAnswer/2agents/bigbench/agents1_3/llama3b_job1155_2026-05-10_12h18m04s',
#     'tasksource/bigbench: Mal Agent 2 & 3': 'transcripts/MAS/unfaithfulReasoning/wrongReason_correctAnswer/2agents/bigbench/agents2_3/llama3b_job2638_2026-05-11_15h13m44s',
# }
# # result = malicious_comparison(transcript_data, round_figure=False, comparison_type='IR-CA', return_data=True, comparison_name='Llama, BigBench, MAS, both')
# # all_results.append(result)
# result = malicious_comparison(transcript_data, round_figure=True, comparison_type='IR-CA',return_data=False, return_round_data=True, comparison_name='Llama, BigBench, MAS, both')
# all_results.append(result)


# # export_comparisons_to_excel(all_results, 'IR_CA_results.xlsx')
# export_round_comparisons_to_excel(all_results, 'IR_CA_round_results.xlsx')



# # ╔═════════════════════════════════════════════════╗
# # ║           CR-IA ONE AGENT COMPARISON            ║
# # ╚═════════════════════════════════════════════════╝

# # --- CR-IA comparison for Qwen ---
# transcript_data = {
#     'openai/gsm8k: No Mal Agent': 'transcripts/MAS/both/qwen3b_2026-03-29_00h31m15s',
#     'openai/gsm8k: Mal Agent 1': 'transcripts/MAS/unfaithfulReasoning/correctReasoning_wrongAnswer/1agent/gsm8k/agent0/qwen3b_job4855_2026-05-15_21h13m45s',
#     'openai/gsm8k: Mal Agent 2': 'transcripts/MAS/unfaithfulReasoning/correctReasoning_wrongAnswer/1agent/gsm8k/agent1/qwen3b_job5975_2026-05-18_11h32m17s',
#     'openai/gsm8k: Mal Agent 3': 'transcripts/MAS/unfaithfulReasoning/correctReasoning_wrongAnswer/1agent/gsm8k/agent2/qwen3b_job8192_2026-05-19_15h31m08s',
# }
# # result = malicious_comparison(transcript_data, round_figure=False, comparison_type='CR-IA', return_data=True, comparison_name='Qwen, GSM8K, MAS, both')
# # all_results.append(result)
# result = malicious_comparison(transcript_data, round_figure=True, comparison_type='CR-IA',return_data=False, return_round_data=True, comparison_name='Qwen, GSM8K, MAS, both')
# all_results.append(result)

# transcript_data = {
#     'cais/mmlu: No Mal Agent': 'transcripts/MAS/both/qwen3b_2026-03-29_15h34m56s',
#     'cais/mmlu: Mal Agent 1': 'transcripts/MAS/unfaithfulReasoning/correctReasoning_wrongAnswer/1agent/mmlu/agent0/qwen3b_job3198_2026-05-16_20h45m29s',
#     'cais/mmlu: Mal Agent 2': 'transcripts/MAS/unfaithfulReasoning/correctReasoning_wrongAnswer/1agent/mmlu/agent1/qwen3b_job6677_2026-05-18_15h55m00s',
#     'cais/mmlu: Mal Agent 3': 'transcripts/MAS/unfaithfulReasoning/correctReasoning_wrongAnswer/1agent/mmlu/agent2/qwen3b_job8195_2026-05-19_17h55m04s',
# }
# # result = malicious_comparison(transcript_data, round_figure=False, comparison_type='CR-IA', return_data=True, comparison_name='Qwen, MMLU, MAS, both')
# # all_results.append(result)
# result = malicious_comparison(transcript_data, round_figure=True, comparison_type='CR-IA',return_data=False, return_round_data=True, comparison_name='Qwen, MMLU, MAS, both')
# all_results.append(result)

# transcript_data = {
#     'ChilleD/StrategyQA: No Mal Agent': 'transcripts/MAS/both/qwen3b_2026-03-29_15h26m25s',
#     'ChilleD/StrategyQA: Mal Agent 1': 'transcripts/MAS/unfaithfulReasoning/correctReasoning_wrongAnswer/1agent/strategyQA/agent0/qwen3b_job3201_2026-05-17_00h30m02s',
#     'ChilleD/StrategyQA: Mal Agent 2': 'transcripts/MAS/unfaithfulReasoning/correctReasoning_wrongAnswer/1agent/strategyQA/agent1/qwen3b_job6956_2026-05-18_23h50m46s',
#     'ChilleD/StrategyQA: Mal Agent 3': 'transcripts/MAS/unfaithfulReasoning/correctReasoning_wrongAnswer/1agent/strategyQA/agent2/qwen3b_job9025_2026-05-20_11h21m24s',
# }
# # result = malicious_comparison(transcript_data, round_figure=False, comparison_type='CR-IA', return_data=True, comparison_name='Qwen, StrategyQA, MAS, both')
# # all_results.append(result)
# result = malicious_comparison(transcript_data, round_figure=True, comparison_type='CR-IA',return_data=False, return_round_data=True, comparison_name='Qwen, StrategyQA, MAS, both')
# all_results.append(result)

# transcript_data = {
#     'tasksource/bigbench: No Mal Agent': 'transcripts/MAS/both/qwen3b_2026-03-29_03h35m18s',
#     'tasksource/bigbench: Mal Agent 1': 'transcripts/MAS/unfaithfulReasoning/correctReasoning_wrongAnswer/1agent/bigbench/agent0/qwen3b_job5978_2026-05-18_12h08m30s',
#     'tasksource/bigbench: Mal Agent 2': 'transcripts/MAS/unfaithfulReasoning/correctReasoning_wrongAnswer/1agent/bigbench/agent1/qwen3b_job7476_2026-05-19_12h25m29s',
#     'tasksource/bigbench: Mal Agent 3': 'transcripts/MAS/unfaithfulReasoning/correctReasoning_wrongAnswer/1agent/bigbench/agent2/qwen3b_job9037_2026-05-20_11h47m01s',
# }
# # result = malicious_comparison(transcript_data, round_figure=False, comparison_type='CR-IA', return_data=True, comparison_name='Qwen, BigBench, MAS, both')
# # all_results.append(result)
# result = malicious_comparison(transcript_data, round_figure=True, comparison_type='CR-IA',return_data=False, return_round_data=True, comparison_name='Qwen, BigBench, MAS, both')
# all_results.append(result)


# # --- CR-IA comparison for Olmo ---
# transcript_data = {
#     'openai/gsm8k: No Mal Agent': 'transcripts/MAS/both/olmo7b_2026-03-29_15h41m56s',
#     'openai/gsm8k: Mal Agent 1': 'transcripts/MAS/unfaithfulReasoning/correctReasoning_wrongAnswer/1agent/gsm8k/agent0/olmo7b_job4856_2026-05-15_22h19m33s',
#     'openai/gsm8k: Mal Agent 2': 'transcripts/MAS/unfaithfulReasoning/correctReasoning_wrongAnswer/1agent/gsm8k/agent1/olmo7b_job5976_2026-05-18_12h01m09s',
#     'openai/gsm8k: Mal Agent 3': 'transcripts/MAS/unfaithfulReasoning/correctReasoning_wrongAnswer/1agent/gsm8k/agent2/olmo7b_job8193_2026-05-19_16h13m38s',
# }
# # result = malicious_comparison(transcript_data, round_figure=False, comparison_type='CR-IA', return_data=True, comparison_name='Olmo, GSM8K, MAS, both')
# # all_results.append(result)
# result = malicious_comparison(transcript_data, round_figure=True, comparison_type='CR-IA',return_data=False, return_round_data=True, comparison_name='Olmo, GSM8K, MAS, both')
# all_results.append(result)

# transcript_data = {
#     'cais/mmlu: No Mal Agent': 'transcripts/MAS/both/olmo7b_2026-03-29_15h29m53s',
#     'cais/mmlu: Mal Agent 1': 'transcripts/MAS/unfaithfulReasoning/correctReasoning_wrongAnswer/1agent/mmlu/agent0/olmo7b_job3199_2026-05-16_22h05m34s',
#     'cais/mmlu: Mal Agent 2': 'transcripts/MAS/unfaithfulReasoning/correctReasoning_wrongAnswer/1agent/mmlu/agent1/olmo7b_job6678_2026-05-18_16h39m20s',
#     'cais/mmlu: Mal Agent 3': 'transcripts/MAS/unfaithfulReasoning/correctReasoning_wrongAnswer/1agent/mmlu/agent2/olmo7b_job8196_2026-05-19_18h42m08s',
# }
# # result = malicious_comparison(transcript_data, round_figure=False, comparison_type='CR-IA', return_data=True, comparison_name='Olmo, MMLU, MAS, both')
# # all_results.append(result)
# result = malicious_comparison(transcript_data, round_figure=True, comparison_type='CR-IA',return_data=False, return_round_data=True, comparison_name='Olmo, MMLU, MAS, both')
# all_results.append(result)

# transcript_data = {
#     'ChilleD/StrategyQA: No Mal Agent': 'transcripts/MAS/both/olmo7b_2026-03-29_15h26m25s',
#     'ChilleD/StrategyQA: Mal Agent 1': 'transcripts/MAS/unfaithfulReasoning/correctReasoning_wrongAnswer/1agent/strategyQA/agent0/olmo7b_job3202_2026-05-17_02h37m10s',
#     'ChilleD/StrategyQA: Mal Agent 2': 'transcripts/MAS/unfaithfulReasoning/correctReasoning_wrongAnswer/1agent/strategyQA/agent1/olmo7b_job6957_2026-05-19_01h03m15s',
#     'ChilleD/StrategyQA: Mal Agent 3': 'transcripts/MAS/unfaithfulReasoning/correctReasoning_wrongAnswer/1agent/strategyQA/agent2/olmo7b_job9026_2026-05-20_11h24m36s',
# }
# # result = malicious_comparison(transcript_data, round_figure=False, comparison_type='CR-IA', return_data=True, comparison_name='Olmo, StrategyQA, MAS, both')
# # all_results.append(result)
# result = malicious_comparison(transcript_data, round_figure=True, comparison_type='CR-IA',return_data=False, return_round_data=True, comparison_name='Olmo, StrategyQA, MAS, both')
# all_results.append(result)

# transcript_data = {
#     'tasksource/bigbench: No Mal Agent': 'transcripts/MAS/both/olmo7b_2026-03-29_01h47m46s',
#     'tasksource/bigbench: Mal Agent 1': 'transcripts/MAS/unfaithfulReasoning/correctReasoning_wrongAnswer/1agent/bigbench/agent0/olmo7b_job5979_2026-05-18_12h08m30s',
#     'tasksource/bigbench: Mal Agent 2': 'transcripts/MAS/unfaithfulReasoning/correctReasoning_wrongAnswer/1agent/bigbench/agent1/olmo7b_job7477_2026-05-19_13h04m13s',
#     'tasksource/bigbench: Mal Agent 3': 'transcripts/MAS/unfaithfulReasoning/correctReasoning_wrongAnswer/1agent/bigbench/agent2/olmo7b_job9038_2026-05-20_12h10m39s',
# }
# # result = malicious_comparison(transcript_data, round_figure=False, comparison_type='CR-IA', return_data=True, comparison_name='Olmo, BigBench, MAS, both')
# # all_results.append(result)
# result = malicious_comparison(transcript_data, round_figure=True, comparison_type='CR-IA',return_data=False, return_round_data=True, comparison_name='Olmo, BigBench, MAS, both')
# all_results.append(result)


# # --- CR-IA comparison for Llama ---
# transcript_data = {
#     'openai/gsm8k: No Mal Agent': 'transcripts/MAS/both/llama3b_2026-03-29_15h19m52s',
#     'openai/gsm8k: Mal Agent 1': 'transcripts/MAS/unfaithfulReasoning/correctReasoning_wrongAnswer/1agent/gsm8k/agent0/llama3b_job4854_2026-05-15_20h59m29s',
#     'openai/gsm8k: Mal Agent 2': 'transcripts/MAS/unfaithfulReasoning/correctReasoning_wrongAnswer/1agent/gsm8k/agent1/llama3b_job5974_2026-05-18_10h31m36s',
#     'openai/gsm8k: Mal Agent 3': 'transcripts/MAS/unfaithfulReasoning/correctReasoning_wrongAnswer/1agent/gsm8k/agent2/llama3b_job8191_2026-05-19_15h09m26s',
# }
# # result = malicious_comparison(transcript_data, round_figure=False, comparison_type='CR-IA', return_data=True, comparison_name='Llama, GSM8K, MAS, both')
# # all_results.append(result)
# result = malicious_comparison(transcript_data, round_figure=True, comparison_type='CR-IA',return_data=False, return_round_data=True, comparison_name='Llama, GSM8K, MAS, both')
# all_results.append(result)

# transcript_data = {
#     'cais/mmlu: No Mal Agent': 'transcripts/MAS/both/llama3b_2026-03-29_00h34m49s',
#     'cais/mmlu: Mal Agent 1': 'transcripts/MAS/unfaithfulReasoning/correctReasoning_wrongAnswer/1agent/mmlu/agent0/llama3b_job3197_2026-05-16_19h28m16s',
#     'cais/mmlu: Mal Agent 2': 'transcripts/MAS/unfaithfulReasoning/correctReasoning_wrongAnswer/1agent/mmlu/agent1/llama3b_job6676_2026-05-18_15h53m06s',
#     'cais/mmlu: Mal Agent 3': 'transcripts/MAS/unfaithfulReasoning/correctReasoning_wrongAnswer/1agent/mmlu/agent2/llama3b_job8194_2026-05-19_16h57m45s',
# }
# # result = malicious_comparison(transcript_data, round_figure=False, comparison_type='CR-IA', return_data=True, comparison_name='Llama, MMLU, MAS, both')
# # all_results.append(result)
# result = malicious_comparison(transcript_data, round_figure=True, comparison_type='CR-IA',return_data=False, return_round_data=True, comparison_name='Llama, MMLU, MAS, both')
# all_results.append(result)

# transcript_data = {
#     'ChilleD/StrategyQA: No Mal Agent': 'transcripts/MAS/both/llama3b_2026-03-29_15h21m49s',
#     'ChilleD/StrategyQA: Mal Agent 1': 'transcripts/MAS/unfaithfulReasoning/correctReasoning_wrongAnswer/1agent/strategyQA/agent0/llama3b_job3200_2026-05-17_00h30m02s',
#     'ChilleD/StrategyQA: Mal Agent 2': 'transcripts/MAS/unfaithfulReasoning/correctReasoning_wrongAnswer/1agent/strategyQA/agent1/llama3b_job6955_2026-05-18_22h44m28s',
#     'ChilleD/StrategyQA: Mal Agent 3': 'transcripts/MAS/unfaithfulReasoning/correctReasoning_wrongAnswer/1agent/strategyQA/agent2/llama3b_job9024_2026-05-20_11h07m38s',
# }
# # result = malicious_comparison(transcript_data, round_figure=False, comparison_type='CR-IA', return_data=True, comparison_name='Llama, StrategyQA, MAS, both')
# # all_results.append(result)
# result = malicious_comparison(transcript_data, round_figure=True, comparison_type='CR-IA',return_data=False, return_round_data=True, comparison_name='Llama, StrategyQA, MAS, both')
# all_results.append(result)

# transcript_data = {
#     'tasksource/bigbench: No Mal Agent': 'transcripts/MAS/both/llama3b_2026-03-29_03h08m47s',
#     'tasksource/bigbench: Mal Agent 1': 'transcripts/MAS/unfaithfulReasoning/correctReasoning_wrongAnswer/1agent/bigbench/agent0/llama3b_job5977_2026-05-18_12h01m09s',
#     'tasksource/bigbench: Mal Agent 2': 'transcripts/MAS/unfaithfulReasoning/correctReasoning_wrongAnswer/1agent/bigbench/agent1/llama3b_job7475_2026-05-19_08h49m52s',
#     'tasksource/bigbench: Mal Agent 3': 'transcripts/MAS/unfaithfulReasoning/correctReasoning_wrongAnswer/1agent/bigbench/agent2/llama3b_job9036_2026-05-20_11h25m56s',
# }
# # result = malicious_comparison(transcript_data, round_figure=False, comparison_type='CR-IA', return_data=True, comparison_name='Llama, BigBench, MAS, both')
# # all_results.append(result)
# result = malicious_comparison(transcript_data, round_figure=True, comparison_type='CR-IA',return_data=False, return_round_data=True, comparison_name='Llama, BigBench, MAS, both')
# all_results.append(result)


# # ╔═════════════════════════════════════════════════╗
# # ║           CR-IA TWO AGENT COMPARISON            ║
# # ╚═════════════════════════════════════════════════╝

# # --- CR-IA comparison for Qwen ---
# transcript_data = {
#     'openai/gsm8k: No Mal Agent': 'transcripts/MAS/both/qwen3b_2026-03-29_00h31m15s',
#     'openai/gsm8k: Mal Agent 1 & 2': 'transcripts/MAS/unfaithfulReasoning/correctReasoning_wrongAnswer/2agents/gsm8k/agents0_1/qwen3b_job2954_2026-05-17_06h06m33s',
#     'openai/gsm8k: Mal Agent 1 & 3': 'transcripts/MAS/unfaithfulReasoning/correctReasoning_wrongAnswer/2agents/gsm8k/agents0_2/qwen3b_job6471_2026-05-18_17h32m03s',
#     'openai/gsm8k: Mal Agent 2 & 3': 'transcripts/MAS/unfaithfulReasoning/correctReasoning_wrongAnswer/2agents/gsm8k/agents1_2/qwen3b_job8227_2026-05-19_20h17m22s',
# }
# result = malicious_comparison(transcript_data, round_figure=False, comparison_type='CR-IA', return_data=True, comparison_name='Qwen, GSM8K, MAS, both')
# # all_results.append(result)
# result = malicious_comparison(transcript_data, round_figure=True, comparison_type='CR-IA',return_data=False, return_round_data=True, comparison_name='Qwen, GSM8K, MAS, both')
# all_results.append(result)

# transcript_data = {
#     'cais/mmlu: No Mal Agent': 'transcripts/MAS/both/qwen3b_2026-03-29_15h34m56s',
#     'cais/mmlu: Mal Agent 1 & 2': 'transcripts/MAS/unfaithfulReasoning/correctReasoning_wrongAnswer/2agents/mmlu/agents0_1/qwen3b_job2958_2026-05-17_17h55m47s',
#     'cais/mmlu: Mal Agent 1 & 3': 'transcripts/MAS/unfaithfulReasoning/correctReasoning_wrongAnswer/2agents/mmlu/agents0_2/qwen3b_job6966_2026-05-19_04h43m11s',
#     'cais/mmlu: Mal Agent 2 & 3': 'transcripts/MAS/unfaithfulReasoning/correctReasoning_wrongAnswer/2agents/mmlu/agents1_2/qwen3b_job8515_2026-05-20_04h51m12s',
# }
# result = malicious_comparison(transcript_data, round_figure=False, comparison_type='CR-IA', return_data=True, comparison_name='Qwen, MMLU, MAS, both')
# # all_results.append(result)
# result = malicious_comparison(transcript_data, round_figure=True, comparison_type='CR-IA',return_data=False, return_round_data=True, comparison_name='Qwen, MMLU, MAS, both')
# all_results.append(result)

# transcript_data = {
#     'ChilleD/StrategyQA: No Mal Agent': 'transcripts/MAS/both/qwen3b_2026-03-29_15h26m25s',
#     'ChilleD/StrategyQA: Mal Agent 1 & 2': 'transcripts/MAS/unfaithfulReasoning/correctReasoning_wrongAnswer/2agents/strategyQA/agents0_1/qwen3b_job2965_2026-05-17_19h40m57s',
#     'ChilleD/StrategyQA: Mal Agent 1 & 3': 'transcripts/MAS/unfaithfulReasoning/correctReasoning_wrongAnswer/2agents/strategyQA/agents0_2/qwen3b_job6970_2026-05-19_07h25m13s',
#     'ChilleD/StrategyQA: Mal Agent 2 & 3': 'transcripts/MAS/unfaithfulReasoning/correctReasoning_wrongAnswer/2agents/strategyQA/agents1_2/qwen3b_job8518_2026-05-20_07h39m10s',
# }
# result = malicious_comparison(transcript_data, round_figure=False, comparison_type='CR-IA', return_data=True, comparison_name='Qwen, StrategyQA, MAS, both')
# # all_results.append(result)
# result = malicious_comparison(transcript_data, round_figure=True, comparison_type='CR-IA',return_data=False, return_round_data=True, comparison_name='Qwen, StrategyQA, MAS, both')
# all_results.append(result)

# transcript_data = {
#     'tasksource/bigbench: No Mal Agent': 'transcripts/MAS/both/qwen3b_2026-03-29_03h35m18s',
#     'tasksource/bigbench: Mal Agent 1 & 2': 'transcripts/MAS/unfaithfulReasoning/correctReasoning_wrongAnswer/2agents/bigbench/agents0_1/qwen3b_job5946_2026-05-18_14h01m47s',
#     'tasksource/bigbench: Mal Agent 1 & 3': 'transcripts/MAS/unfaithfulReasoning/correctReasoning_wrongAnswer/2agents/bigbench/agents0_2/qwen3b_job7774_2026-05-19_14h08m25s',
#     'tasksource/bigbench: Mal Agent 2 & 3': 'transcripts/MAS/unfaithfulReasoning/correctReasoning_wrongAnswer/2agents/bigbench/agents1_2/qwen3b_job9124_2026-05-20_13h28m28s',
# }
# result = malicious_comparison(transcript_data, round_figure=False, comparison_type='CR-IA', return_data=True, comparison_name='Qwen, BigBench, MAS, both')
# # all_results.append(result)
# result = malicious_comparison(transcript_data, round_figure=True, comparison_type='CR-IA',return_data=False, return_round_data=True, comparison_name='Qwen, BigBench, MAS, both')
# all_results.append(result)


# # --- CR-IA comparison for Olmo ---
# transcript_data = {
#     'openai/gsm8k: No Mal Agent': 'transcripts/MAS/both/olmo7b_2026-03-29_15h41m56s',
#     'openai/gsm8k: Mal Agent 1 & 2': 'transcripts/MAS/unfaithfulReasoning/correctReasoning_wrongAnswer/2agents/gsm8k/agents0_1/olmo7b_job2955_2026-05-17_06h22m58s',
#     'openai/gsm8k: Mal Agent 1 & 3': 'transcripts/MAS/unfaithfulReasoning/correctReasoning_wrongAnswer/2agents/gsm8k/agents0_2/olmo7b_job6472_2026-05-18_17h43m15s',
#     'openai/gsm8k: Mal Agent 2 & 3': 'transcripts/MAS/unfaithfulReasoning/correctReasoning_wrongAnswer/2agents/gsm8k/agents1_2/olmo7b_job8512_2026-05-20_02h53m54s',
# }
# result = malicious_comparison(transcript_data, round_figure=False, comparison_type='CR-IA', return_data=True, comparison_name='Olmo, GSM8K, MAS, both')
# # all_results.append(result)
# result = malicious_comparison(transcript_data, round_figure=True, comparison_type='CR-IA',return_data=False, return_round_data=True, comparison_name='Olmo, GSM8K, MAS, both')
# all_results.append(result)

# transcript_data = {
#     'cais/mmlu: No Mal Agent': 'transcripts/MAS/both/olmo7b_2026-03-29_15h29m53s',
#     'cais/mmlu: Mal Agent 1 & 2': 'transcripts/MAS/unfaithfulReasoning/correctReasoning_wrongAnswer/2agents/mmlu/agents0_1/olmo7b_job2959_2026-05-17_18h20m47s',
#     'cais/mmlu: Mal Agent 1 & 3': 'transcripts/MAS/unfaithfulReasoning/correctReasoning_wrongAnswer/2agents/mmlu/agents0_2/olmo7b_job6967_2026-05-19_05h06m08s',
#     'cais/mmlu: Mal Agent 2 & 3': 'transcripts/MAS/unfaithfulReasoning/correctReasoning_wrongAnswer/2agents/mmlu/agents1_2/olmo7b_job8516_2026-05-20_05h40m38s',
# }
# result = malicious_comparison(transcript_data, round_figure=False, comparison_type='CR-IA', return_data=True, comparison_name='Olmo, MMLU, MAS, both')
# # all_results.append(result)
# result = malicious_comparison(transcript_data, round_figure=True, comparison_type='CR-IA',return_data=False, return_round_data=True, comparison_name='Olmo, MMLU, MAS, both')
# all_results.append(result)

# transcript_data = {
#     'ChilleD/StrategyQA: No Mal Agent': 'transcripts/MAS/both/olmo7b_2026-03-29_15h26m25s',
#     'ChilleD/StrategyQA: Mal Agent 1 & 2': 'transcripts/MAS/unfaithfulReasoning/correctReasoning_wrongAnswer/2agents/strategyQA/agents0_1/olmo7b_job5943_2026-05-18_12h08m44s',
#     'ChilleD/StrategyQA: Mal Agent 1 & 3': 'transcripts/MAS/unfaithfulReasoning/correctReasoning_wrongAnswer/2agents/strategyQA/agents0_2/olmo7b_job6971_2026-05-19_07h34m23s',
#     'ChilleD/StrategyQA: Mal Agent 2 & 3': 'transcripts/MAS/unfaithfulReasoning/correctReasoning_wrongAnswer/2agents/strategyQA/agents1_2/olmo7b_job9122_2026-05-20_13h12m55s',
# }
# result = malicious_comparison(transcript_data, round_figure=False, comparison_type='CR-IA', return_data=True, comparison_name='Olmo, StrategyQA, MAS, both')
# # all_results.append(result)
# result = malicious_comparison(transcript_data, round_figure=True, comparison_type='CR-IA',return_data=False, return_round_data=True, comparison_name='Olmo, StrategyQA, MAS, both')
# all_results.append(result)

# transcript_data = {
#     'tasksource/bigbench: No Mal Agent': 'transcripts/MAS/both/olmo7b_2026-03-29_01h47m46s',
#     'tasksource/bigbench: Mal Agent 1 & 2': 'transcripts/MAS/unfaithfulReasoning/correctReasoning_wrongAnswer/2agents/bigbench/agents0_1/olmo7b_job5947_2026-05-18_15h10m49s',
#     'tasksource/bigbench: Mal Agent 1 & 3': 'transcripts/MAS/unfaithfulReasoning/correctReasoning_wrongAnswer/2agents/bigbench/agents0_2/olmo7b_job8109_2026-05-19_14h22m46s',
#     'tasksource/bigbench: Mal Agent 2 & 3': 'transcripts/MAS/unfaithfulReasoning/correctReasoning_wrongAnswer/2agents/bigbench/agents1_2/olmo7b_job9370_2026-05-20_16h27m09s',
# }
# result = malicious_comparison(transcript_data, round_figure=False, comparison_type='CR-IA', return_data=True, comparison_name='Olmo, BigBench, MAS, both')
# # all_results.append(result)
# result = malicious_comparison(transcript_data, round_figure=True, comparison_type='CR-IA',return_data=False, return_round_data=True, comparison_name='Olmo, BigBench, MAS, both')
# all_results.append(result)


# # --- CR-IA comparison for Llama ---
# transcript_data = {
#     'openai/gsm8k: No Mal Agent': 'transcripts/MAS/both/llama3b_2026-03-29_15h19m52s',
#     'openai/gsm8k: Mal Agent 1 & 2': 'transcripts/MAS/unfaithfulReasoning/correctReasoning_wrongAnswer/2agents/gsm8k/agents0_1/llama3b_job2953_2026-05-16_14h31m32s',
#     'openai/gsm8k: Mal Agent 1 & 3': 'transcripts/MAS/unfaithfulReasoning/correctReasoning_wrongAnswer/2agents/gsm8k/agents0_2/llama3b_job6470_2026-05-18_17h02m32s',
#     'openai/gsm8k: Mal Agent 2 & 3': 'transcripts/MAS/unfaithfulReasoning/correctReasoning_wrongAnswer/2agents/gsm8k/agents1_2/llama3b_job8226_2026-05-19_18h49m59s',
# }
# result = malicious_comparison(transcript_data, round_figure=False, comparison_type='CR-IA', return_data=True, comparison_name='Llama, GSM8K, MAS, both')
# # all_results.append(result)
# result = malicious_comparison(transcript_data, round_figure=True, comparison_type='CR-IA',return_data=False, return_round_data=True, comparison_name='Llama, GSM8K, MAS, both')
# all_results.append(result)

# transcript_data = {
#     'cais/mmlu: No Mal Agent': 'transcripts/MAS/both/llama3b_2026-03-29_00h34m49s',
#     'cais/mmlu: Mal Agent 1 & 2': 'transcripts/MAS/unfaithfulReasoning/correctReasoning_wrongAnswer/2agents/mmlu/agents0_1/llama3b_job2957_2026-05-17_10h47m15s',
#     'cais/mmlu: Mal Agent 1 & 3': 'transcripts/MAS/unfaithfulReasoning/correctReasoning_wrongAnswer/2agents/mmlu/agents0_2/llama3b_job6473_2026-05-18_18h22m37s',
#     'cais/mmlu: Mal Agent 2 & 3': 'transcripts/MAS/unfaithfulReasoning/correctReasoning_wrongAnswer/2agents/mmlu/agents1_2/llama3b_job8514_2026-05-20_03h34m56s',
# }
# result = malicious_comparison(transcript_data, round_figure=False, comparison_type='CR-IA', return_data=True, comparison_name='Llama, MMLU, MAS, both')
# # all_results.append(result)
# result = malicious_comparison(transcript_data, round_figure=True, comparison_type='CR-IA',return_data=False, return_round_data=True, comparison_name='Llama, MMLU, MAS, both')
# all_results.append(result)

# transcript_data = {
#     'ChilleD/StrategyQA: No Mal Agent': 'transcripts/MAS/both/llama3b_2026-03-29_15h21m49s',
#     'ChilleD/StrategyQA: Mal Agent 1 & 2': 'transcripts/MAS/unfaithfulReasoning/correctReasoning_wrongAnswer/2agents/strategyQA/agents0_1/llama3b_job2964_2026-05-17_18h45m10s',
#     'ChilleD/StrategyQA: Mal Agent 1 & 3': 'transcripts/MAS/unfaithfulReasoning/correctReasoning_wrongAnswer/2agents/strategyQA/agents0_2/llama3b_job6969_2026-05-19_06h41m56s',
#     'ChilleD/StrategyQA: Mal Agent 2 & 3': 'transcripts/MAS/unfaithfulReasoning/correctReasoning_wrongAnswer/2agents/strategyQA/agents1_2/llama3b_job8517_2026-05-20_05h56m50s',
# }
# result = malicious_comparison(transcript_data, round_figure=False, comparison_type='CR-IA', return_data=True, comparison_name='Llama, StrategyQA, MAS, both')
# # all_results.append(result)
# result = malicious_comparison(transcript_data, round_figure=True, comparison_type='CR-IA',return_data=False, return_round_data=True, comparison_name='Llama, StrategyQA, MAS, both')
# all_results.append(result)

# transcript_data = {
#     'tasksource/bigbench: No Mal Agent': 'transcripts/MAS/both/llama3b_2026-03-29_03h08m47s',
#     'tasksource/bigbench: Mal Agent 1 & 2': 'transcripts/MAS/unfaithfulReasoning/correctReasoning_wrongAnswer/2agents/bigbench/agents0_1/llama3b_job5945_2026-05-18_13h13m17s',
#     'tasksource/bigbench: Mal Agent 1 & 3': 'transcripts/MAS/unfaithfulReasoning/correctReasoning_wrongAnswer/2agents/bigbench/agents0_2/llama3b_job7773_2026-05-19_14h07m22s',
#     'tasksource/bigbench: Mal Agent 2 & 3': 'transcripts/MAS/unfaithfulReasoning/correctReasoning_wrongAnswer/2agents/bigbench/agents1_2/llama3b_job9123_2026-05-20_13h18m49s',
# }
# result = malicious_comparison(transcript_data, round_figure=False, comparison_type='CR-IA', return_data=True, comparison_name='Llama, BigBench, MAS, both')
# # all_results.append(result)
# result = malicious_comparison(transcript_data, round_figure=True, comparison_type='CR-IA',return_data=False, return_round_data=True, comparison_name='Llama, BigBench, MAS, both')
# all_results.append(result)


# #export_comparisons_to_excel(all_results, 'CR_IA_results.xlsx')
# export_round_comparisons_to_excel(all_results, 'CR_IA_round_results.xlsx')

# ╔═════════════════════════════════════════════════╗
# ║                       RQ1                       ║
# ╚═════════════════════════════════════════════════╝

# # --- Comparison for openai/gsm8k, both, MAS ---
# transcript_data = {
#     'Qwen/Qwen2.5-3B-Instruct': 'transcripts/MAS/both/qwen3b_2026-03-29_00h31m15s',
#     'allenai/Olmo-3-7B-Instruct': 'transcripts/MAS/both/olmo7b_2026-03-29_15h41m56s',
#     'meta-llama/Llama-3.2-3B-Instruct': 'transcripts/MAS/both/llama3b_2026-03-29_15h19m52s',
# }
# dataset_comparison(transcript_data, round_figure=False, comparison_type='model')


# --- Comparison for cais/mmlu, both, MAS ---
# transcript_data = {
#     'Qwen/Qwen2.5-3B-Instruct': 'transcripts/MAS/both/...',
#     'allenai/Olmo-3-7B-Instruct': 'transcripts/MAS/both/...',
#     'meta-llama/Llama-3.2-3B-Instruct': 'transcripts/MAS/both/...',
# }
# dataset_comparison(transcript_data, round_figure=False, comparison_type='model')

# --- Comparison for ChilleD/StrategyQA, both, MAS ---
# transcript_data = {
#     'Qwen/Qwen2.5-3B-Instruct': 'transcripts/MAS/both/...',
#     'allenai/Olmo-3-7B-Instruct': 'transcripts/MAS/both/...',
#     'meta-llama/Llama-3.2-3B-Instruct': 'transcripts/MAS/both/...',
# }
# dataset_comparison(transcript_data, round_figure=False, comparison_type='model')


# --- Comparison for tasksource/bigbench, both, MAS ---
# transcript_data = {
#     'Qwen/Qwen2.5-3B-Instruct': 'transcripts/MAS/both/...',
#     'allenai/Olmo-3-7B-Instruct': 'transcripts/MAS/both/...',
#     'meta-llama/Llama-3.2-3B-Instruct': 'transcripts/MAS/both/...',
# }
# dataset_comparison(transcript_data, round_figure=False, comparison_type='tasksource/bigbench')


# ╔═════════════════════════════════════════════════╗
# ║                       RQ2                       ║
# ╚═════════════════════════════════════════════════╝

# ------ Comparison for QWEN, both, MAS --------
# transcript_data = {
#     'openai/gsm8k': 'transcripts/MAS/both/qwen3b_2026-03-29_00h31m15s',
#     'cais/mmlu': 'transcripts/MAS/both/qwen3b_2026-03-29_15h34m56s',
#     'ChilleD/StrategyQA': 'transcripts/MAS/both/qwen3b_2026-03-29_15h26m25s',
#     'tasksource/bigbench': 'transcripts/MAS/both/qwen3b_2026-03-29_03h35m18s'
# }
# dataset_comparison(transcript_data, round_figure=True)

# # ------ Comparison for Olmo, both, MAS --------
# transcript_data = {
#     'openai/gsm8k': 'transcripts/MAS/both/olmo7b_2026-03-29_15h41m56s',
#     'cais/mmlu': 'transcripts/MAS/both/olmo7b_2026-03-29_15h29m53s',
#     'ChilleD/StrategyQA': 'transcripts/MAS/both/olmo7b_2026-03-29_15h26m25s',
#     'tasksource/bigbench': 'transcripts/MAS/both/olmo7b_2026-03-29_01h47m46s'
# }
# dataset_comparison(transcript_data, round_figure=True)

# # ------ Comparison for Llama, both, MAS --------
# transcript_data = {
#     'openai/gsm8k': 'transcripts/MAS/both/llama3b_2026-03-29_15h19m52s',
#     'cais/mmlu': 'transcripts/MAS/both/llama3b_2026-03-29_00h34m49s',
#     'ChilleD/StrategyQA': 'transcripts/MAS/both/llama3b_2026-03-29_15h21m49s',
#     'tasksource/bigbench': 'transcripts/MAS/both/llama3b_2026-03-29_03h08m47s'
# }
# dataset_comparison(transcript_data, round_figure=True)


# ╔═════════════════════════════════════════════════╗
# ║                       RQ3                       ║
# ╚═════════════════════════════════════════════════╝

# --- Malicious prompt comparison for Qwen ---
# transcript_data = {
#     'openai/gsm8k: No Mal Agent': 'transcripts/MAS/both/qwen3b_2026-03-29_00h31m15s',
#     'openai/gsm8k: Mal Agent 1 - Both': 'transcripts/MAS/malicious/gsm8k/qwen3b_job345_2026-04-01_01h35m32s',
#     'openai/gsm8k: Mal Agent 1 - Answer': 'transcripts/MAS/malicious/gsm8k/qwen3b_job3963_2026-04-02_02h14m14s',
#     'cais/mmlu: No Mal Agent': 'transcripts/MAS/both/qwen3b_2026-03-29_15h34m56s',
#     'cais/mmlu: Mal Agent 1 - Both': 'transcripts/MAS/malicious/mmlu/qwen3b_job4470_2026-04-02_08h10m34s',
#     'cais/mmlu: Mal Agent 1 - Answer': 'transcripts/MAS/malicious/mmlu/qwen3b_job7400_2026-04-03_01h36m45s',
#     'ChilleD/StrategyQA: No Mal Agent': 'transcripts/MAS/both/qwen3b_2026-03-29_15h26m25s',
#     'ChilleD/StrategyQA: Mal Agent 1 - Both': 'transcripts/MAS/malicious/strategyQA/qwen3b_job7394_2026-04-03_01h31m22s',
#     'ChilleD/StrategyQA: Mal Agent 1 - Answer': 'transcripts/MAS/malicious/strategyQA/qwen3b_job7387_2026-04-03_01h15m56s',
#     'tasksource/bigbench: No Mal Agent': 'transcripts/MAS/both/qwen3b_2026-03-29_03h35m18s',
#     'tasksource/bigbench: Mal Agent 1 - Both': 'transcripts/MAS/malicious/bigbench/qwen3b_job9921_2026-03-31_13h53m46s',
#     'tasksource/bigbench: Mal Agent 1 - Answer': 'transcripts/MAS/malicious/bigbench/qwen3b_job9929_2026-03-31_16h56m02s',
# }
# malicious_comparison(transcript_data, round_figure=False, comparison_type='sharemode_malicious')

# # --- Malicious prompt comparison for Olmo ---
# transcript_data = {
#     'openai/gsm8k: No Mal Agent': 'transcripts/MAS/both/...,
#     'openai/gsm8k: Mal Agent 1 - Answer': 'transcripts/MAS/malicious/gsm8k/...',
#     'openai/gsm8k: Mal Agent 1 - Both': 'transcripts/MAS/malicious/gsm8k/...',
#     'cais/mmlu: No Mal Agent': 'transcripts/MAS/both/...',
#     'cais/mmlu: Mal Agent 1 - Answer': 'transcripts/MAS/malicious/mmlu/...',
#     'cais/mmlu: Mal Agent 1 - Both': 'transcripts/MAS/malicious/mmlu/...',
#     'ChilleD/StrategyQA: No Mal Agent': 'transcripts/MAS/both/...',
#     'ChilleD/StrategyQA: Mal Agent 1 - Answer': 'transcripts/MAS/malicious/strategyQA/...',
#     'ChilleD/StrategyQA: Mal Agent 1 - Both': 'transcripts/MAS/malicious/strategyQA/...',
#     'tasksource/bigbench: No Mal Agent': 'transcripts/MAS/both/...',
#     'tasksource/bigbench: Mal Agent 1 - Answer': 'transcripts/MAS/malicious/bigbench/...',
#     'tasksource/bigbench: Mal Agent 1 - Both': 'transcripts/MAS/malicious/bigbench/...',
# }
# malicious_comparison(transcript_data, round_figure=False, comparison_type='sharemode_malicious')

# # --- Malicious prompt comparison for Llama ---
# transcript_data = {
#     'openai/gsm8k: No Mal Agent': 'transcripts/MAS/both/...,
#     'openai/gsm8k: Mal Agent 1 - Answer': 'transcripts/MAS/malicious/gsm8k/...',
#     'openai/gsm8k: Mal Agent 1 - Both': 'transcripts/MAS/malicious/gsm8k/...',
#     'cais/mmlu: No Mal Agent': 'transcripts/MAS/both/...',
#     'cais/mmlu: Mal Agent 1 - Answer': 'transcripts/MAS/malicious/mmlu/...',
#     'cais/mmlu: Mal Agent 1 - Both': 'transcripts/MAS/malicious/mmlu/...',
#     'ChilleD/StrategyQA: No Mal Agent': 'transcripts/MAS/both/...',
#     'ChilleD/StrategyQA: Mal Agent 1 - Answer': 'transcripts/MAS/malicious/strategyQA/...',
#     'ChilleD/StrategyQA: Mal Agent 1 - Both': 'transcripts/MAS/malicious/strategyQA/...',
#     'tasksource/bigbench: No Mal Agent': 'transcripts/MAS/both/...',
#     'tasksource/bigbench: Mal Agent 1 - Answer': 'transcripts/MAS/malicious/bigbench/...',
#     'tasksource/bigbench: Mal Agent 1 - Both': 'transcripts/MAS/malicious/bigbench/...',
# }
# malicious_comparison(transcript_data, round_figure=False, comparison_type='sharemode_malicious')



# ╔═════════════════════════════════════════════════╗
# ║                       RQ4                       ║
# ╚═════════════════════════════════════════════════╝

# # --- Comparison for QWEN, dataset, both ---
# transcript_data = {
#     'SAS': 'transcripts/MAS/both/...',
#     'MAS': 'transcripts/MAS/both/...',
# }
# dataset_comparison(transcript_data, round_figure=False, comparison_type='system')

# # --- Comparison for Olmo, dataset, both ---
# transcript_data = {
#     'SAS': 'transcripts/MAS/both/...',
#     'MAS': 'transcripts/MAS/both/...',
# }
# dataset_comparison(transcript_data, round_figure=False, comparison_type='system')

# --- Comparison for Llama, tasksource/bigbench, both ---
# transcript_data = {
#     'SAS': 'transcripts/SAS/both/llama3b_job9791_2026-03-31_09h16m39s',
#     'MAS': 'transcripts/MAS/both/llama3b_2026-03-29_03h08m47s',
# }
# dataset_comparison(transcript_data, round_figure=False, comparison_type='system')


# ╔═════════════════════════════════════════════════╗
# ║                       RQ5                       ║
# ╚═════════════════════════════════════════════════╝

# # --- Malicious prompt comparison for Qwen, both ---
# transcript_data = {
#     'openai/gsm8k: No Mal Agent': 'transcripts/MAS/both/qwen3b_2026-03-29_00h31m15s',
#     'openai/gsm8k: Mal Agent 1': 'transcripts/MAS/malicious/gsm8k/qwen3b_job345_2026-04-01_01h35m32s',
#     'openai/gsm8k: Mal Agent 1 & 2': 'transcripts/MAS/malicious/2maliciousAgents/agents1_2/gsm8k/qwen3b_job7913_2026-04-04_02h04m02s',
#     'cais/mmlu: No Mal Agent': 'transcripts/MAS/both/qwen3b_2026-03-29_15h34m56s',
#     'cais/mmlu: Mal Agent 1': 'transcripts/MAS/malicious/mmlu/qwen3b_job4470_2026-04-02_08h10m34s',
#     'cais/mmlu: Mal Agent 1 & 2': 'transcripts/MAS/malicious/2maliciousAgents/agents1_2/mmlu/qwen3b_job7916_2026-04-04_02h03m27s',
#     'ChilleD/StrategyQA: No Mal Agent': 'transcripts/MAS/both/qwen3b_2026-03-29_15h26m25s',
#     'ChilleD/StrategyQA: Mal Agent 1': 'transcripts/MAS/malicious/strategyQA/qwen3b_job7394_2026-04-03_01h31m22s',
#     'ChilleD/StrategyQA: Mal Agent 1 & 2': 'transcripts/MAS/malicious/2maliciousAgents/agents1_2/strategyQA/qwen3b_job7920_2026-04-04_02h08m17s',
#     'tasksource/bigbench: No Mal Agent': 'transcripts/MAS/both/qwen3b_2026-03-29_03h35m18s',
#     'tasksource/bigbench: Mal Agent 1': 'transcripts/MAS/malicious/bigbench/qwen3b_job9921_2026-03-31_13h53m46s',
#     'tasksource/bigbench: Mal Agent 1 & 2': 'transcripts/MAS/malicious/2maliciousAgents/agents1_2/bigbench/qwen3b_job7929_2026-04-04_07h15m35s',
# }
# malicious_comparison(transcript_data, round_figure=False, comparison_type='malicious_convergence')

# # --- Malicious prompt comparison for Qwen, both, 1 & 3 ---
# transcript_data = {
#     'openai/gsm8k: No Mal Agent': 'transcripts/MAS/both/qwen3b_2026-03-29_00h31m15s',
#     'openai/gsm8k: Mal Agent 1': 'transcripts/MAS/malicious/gsm8k/qwen3b_job345_2026-04-01_01h35m32s',
#     'openai/gsm8k: Mal Agent 1 & 3': 'transcripts/MAS/malicious/2maliciousAgents/agents1_3/gsm8k/qwen3b_job5417_2026-04-08_18h19m29s',
#     'cais/mmlu: No Mal Agent': 'transcripts/MAS/both/qwen3b_2026-03-29_15h34m56s',
#     'cais/mmlu: Mal Agent 1': 'transcripts/MAS/malicious/mmlu/qwen3b_job4470_2026-04-02_08h10m34s',
#     'cais/mmlu: Mal Agent 1 & 3': 'transcripts/MAS/malicious/2maliciousAgents/agents1_3/mmlu/qwen3b_job5416_2026-04-08_13h29m51s',
#     'ChilleD/StrategyQA: No Mal Agent': 'transcripts/MAS/both/qwen3b_2026-03-29_15h26m25s',
#     'ChilleD/StrategyQA: Mal Agent 1': 'transcripts/MAS/malicious/strategyQA/qwen3b_job7394_2026-04-03_01h31m22s',
#     'ChilleD/StrategyQA: Mal Agent 1 & 3': 'transcripts/MAS/malicious/2maliciousAgents/agents1_3/strategyQA/qwen3b_job5421_2026-04-09_08h09m11s',
#     'tasksource/bigbench: No Mal Agent': 'transcripts/MAS/both/qwen3b_2026-03-29_03h35m18s',
#     'tasksource/bigbench: Mal Agent 1': 'transcripts/MAS/malicious/bigbench/qwen3b_job9921_2026-03-31_13h53m46s',
#     'tasksource/bigbench: Mal Agent 1 & 3': 'transcripts/MAS/malicious/2maliciousAgents/agents1_3/bigbench/qwen3b_job1760_2026-04-10_15h44m06s',
# }
# malicious_comparison(transcript_data, round_figure=False, comparison_type='malicious_convergence')

# --- Malicious prompt comparison for Llama, both ---
# transcript_data = {
#     'openai/gsm8k: No Mal Agent': 'transcripts/MAS/both/qwen3b_2026-03-29_00h31m15s',
#     'openai/gsm8k: Mal Agent 1': 'transcripts/MAS/malicious/gsm8k/qwen3b_job345_2026-04-01_01h35m32s',
#     'openai/gsm8k: Mal Agent 1 & 3': 'transcripts/MAS/malicious/2maliciousAgents/agents1_3/gsm8k/qwen3b_job5417_2026-04-08_18h19m29s',
#     'cais/mmlu: No Mal Agent': 'transcripts/MAS/both/qwen3b_2026-03-29_15h34m56s',
#     'cais/mmlu: Mal Agent 1': 'transcripts/MAS/malicious/mmlu/qwen3b_job4470_2026-04-02_08h10m34s',
#     'cais/mmlu: Mal Agent 1 & 3': 'transcripts/MAS/malicious/2maliciousAgents/agents1_3/mmlu/qwen3b_job5416_2026-04-08_13h29m51s',
#     'ChilleD/StrategyQA: No Mal Agent': 'transcripts/MAS/both/qwen3b_2026-03-29_15h26m25s',
#     'ChilleD/StrategyQA: Mal Agent 1': 'transcripts/MAS/malicious/strategyQA/qwen3b_job7394_2026-04-03_01h31m22s',
#     'ChilleD/StrategyQA: Mal Agent 1 & 3': 'transcripts/MAS/malicious/2maliciousAgents/agents1_3/strategyQA/qwen3b_job5421_2026-04-09_08h09m11s',
#     'tasksource/bigbench: No Mal Agent': 'transcripts/MAS/both/qwen3b_2026-03-29_03h35m18s',
#     'tasksource/bigbench: Mal Agent 1': 'transcripts/MAS/malicious/bigbench/qwen3b_job9921_2026-03-31_13h53m46s',
#     'tasksource/bigbench: Mal Agent 1 & 3': 'transcripts/MAS/malicious/2maliciousAgents/agents1_3/bigbench/qwen3b_job1760_2026-04-10_15h44m06s',
# }
# malicious_comparison(transcript_data, round_figure=False, comparison_type='malicious_convergence')


# Test normalization function
# print(normalize_answer("1, 4"))
# print(normalize_answer("1,4"))
# print(normalize_answer("no"))
# print(normalize_answer("yes"))
# print(normalize_answer("false"))
# print(normalize_answer("False, true"))
# print(normalize_answer("It is implausible"))
# print(normalize_answer("(1,6)"))
# print(normalize_answer("(1, 6)"))
# print(normalize_answer("(x - 2)(x^2 + 4x + 5)"))
# print(normalize_answer("(x-2)(x^2+4x+5)"))
# print(normalize_answer("infinite, non abelian group"))
# print(normalize_answer("abelian group"))
# print(normalize_answer("no</think>"))
# print(normalize_answer("$no</think>"))
# print(normalize_answer("implausible."))
# print(normalize_answer("implausible!"))
# print(normalize_answer("implausible?"))
# print(normalize_answer("implausible$"))
# print(normalize_answer("implausible#"))
# print(normalize_answer("implausible&"))